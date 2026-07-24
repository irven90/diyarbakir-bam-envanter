# -*- coding: utf-8 -*-
from __future__ import annotations

import io
import json
import os
import re
import traceback
import threading
import time
from dataclasses import dataclass
from datetime import datetime, date
from typing import List, Dict, Optional, Tuple

from flask import (
    Flask, request, redirect, url_for, session,
    render_template_string, send_file, make_response, jsonify
)
from flask_sqlalchemy import SQLAlchemy
from sqlalchemy import func
from werkzeug.security import generate_password_hash, check_password_hash
from fpdf import FPDF
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# =========================================================
# AYARLAR
# =========================================================
# PyInstaller ile dondurulmuş (EXE) çalışma kontrolü
import sys
if getattr(sys, 'frozen', False):
    # EXE'nin bulunduğu klasör (Veritabanı burada kalmalı)
    APP_DIR = os.path.dirname(sys.executable)
    # EXE içine gömülen kaynak dosyalarının çıkarıldığı geçici klasör (_MEIPASS)
    BUNDLE_DIR = sys._MEIPASS
else:
    # Normal Python çalışma ortamı
    APP_DIR = os.path.dirname(os.path.abspath(__file__))
    BUNDLE_DIR = APP_DIR

DB_PATH = os.path.join(APP_DIR, "zimmet.db")

FONT_PATH = os.path.join(BUNDLE_DIR, "DejaVuSans.ttf")
LOGO_PATH = os.path.join(BUNDLE_DIR, "logo.png")
TEMPLATE_PATH = os.path.join(BUNDLE_DIR, "fis_template.png")


USERS = [
    ("Murat İRVEN", "187665", "Adalet21"),
    ("Arda EKER", "327360", "Adalet21"),
    ("Ahmet AYDIN", "187696", "Adalet21"),
    ("Necip Fazıl EKER", "86537", "Adalet21"),
]

ADMIN_SICIL = "187665"  # admin sicil

DEVICE_CATEGORIES = ["Kasa", "Monitör", "Yazıcı", "Tarayıcı"]
DEVICE_STATUSES = ["Depoda", "Zimmetli", "Arızalı / Bakımda", "Hek / Hurda"]

BAM_UNITS_STRUCTURE = {
    "Ceza Daireleri": [f"{i}. Ceza Dairesi" for i in range(1, 18)],
    "Hukuk Daireleri": [f"{i}. Hukuk Dairesi" for i in range(1, 13)],
    "Duruşma Salonları": [f"Duruşma Salonu {i} (Ortak)" for i in range(1, 9)],
    "Başsavcılık ve İdari Bürolar": [
        "BAM Başsavcılığı",
        "BAM Komisyon Başkanlığı",
        "Ceza Daireleri Başkanlığı",
        "Hukuk Daireleri Başkanlığı",
        "Savcılık Bürosu",
        "Adalet Komisyonu Bürosu",
        "Önbüro - Vezne",
        "Polis Noktaları",
        "Veri Temizleme Bürosu",
        "İdari İşler Müdürlüğü",
        "Bilgi İşlem Müdürlüğü",
        "Posta Bürosu",
        "Santral Bürosu",
        "Mutemetlik Bürosu"
    ]
}

STOCK_ITEMS = [
    "Lexmark Ms710-810dn Toner",
    "Lexmark Ms710-810dn Drum",
    "Lexmark Ms823-Mx722 Toner",
    "Lexmark Ms7823-Mx722 Drum",
    "Epson M320dn Toner",
    "Epson M320dn Drum",
    "Brother L-5100 Toner",
    "Brother L-5100 Drum",
    "Pantum BP5100dn Toner",
    "Pantum BP5100dn Drum",
    "F Klavye",
    "Q Klavye",
    "Mouse",
    "USB Çoklayıcı",
    "3 metre Uyap Kablosu",
    "5 metre Uyap Kablosu",
    "10 metre Uyap Kablosu",
]

# =========================================================
# FLASK
# =========================================================
app = Flask(__name__)
app.secret_key = "zimmet-2026-clean-singlefile"
DATABASE_URL = os.environ.get("DATABASE_URL")
if DATABASE_URL:
    if DATABASE_URL.startswith("postgres://"):
        DATABASE_URL = DATABASE_URL.replace("postgres://", "postgresql://", 1)
    app.config["SQLALCHEMY_DATABASE_URI"] = DATABASE_URL
else:
    app.config["SQLALCHEMY_DATABASE_URI"] = "sqlite:///" + DB_PATH.replace("\\", "/")
app.config["SQLALCHEMY_TRACK_MODIFICATIONS"] = False
db = SQLAlchemy(app)

# =========================================================
# MODELLER
# =========================================================
class User(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(120), nullable=False)
    sicil = db.Column(db.String(20), unique=True, nullable=False)
    password_hash = db.Column(db.String(256), nullable=False)

class InventoryItem(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    category = db.Column(db.String(30), nullable=False)
    brand = db.Column(db.String(60), nullable=False)
    # MODEL OPSİYONEL
    model = db.Column(db.String(80), nullable=True)
    serial_no = db.Column(db.String(120), unique=True, nullable=False)

    status = db.Column(db.String(20), default="Depoda")  # Depoda / Zimmetli
    assigned_name = db.Column(db.String(120))
    assigned_sicil = db.Column(db.String(20))
    assigned_title = db.Column(db.String(120))
    assigned_unit = db.Column(db.String(120))
    assigned_at = db.Column(db.String(20))

    last_event = db.Column(db.String(30))
    last_event_at = db.Column(db.String(30))

class Stock(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(120), unique=True, nullable=False)
    qty = db.Column(db.Integer, default=0)

class History(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    at = db.Column(db.String(30), nullable=False)       # batch timestamp
    action = db.Column(db.String(30), nullable=False)   # ZIMMET / TESLIM_TESELLUM

    inv_id = db.Column(db.Integer)
    category = db.Column(db.String(30))
    brand = db.Column(db.String(60))
    model = db.Column(db.String(80))
    serial_no = db.Column(db.String(120))

    alan_ad = db.Column(db.String(120))
    alan_sicil = db.Column(db.String(20))
    alan_unvan = db.Column(db.String(120))
    alan_birim = db.Column(db.String(120))
    bilgisayar_adi = db.Column(db.String(120))

    teslim_ad = db.Column(db.String(120))
    teslim_sicil = db.Column(db.String(20))
    teslim_unvan = db.Column(db.String(120))

class AuditLog(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    at = db.Column(db.String(40), nullable=False)
    user_name = db.Column(db.String(120), nullable=True)
    user_sicil = db.Column(db.String(40), nullable=True)
    action = db.Column(db.String(100), nullable=False)
    details = db.Column(db.Text, nullable=True)
    ip_address = db.Column(db.String(50), nullable=True)

def log_audit(action: str, details: str = ""):
    try:
        u_name = session.get("user") or "Sistem"
        u_sicil = session.get("sicil") or "-"
        ip = "-"
        if request:
            ip = request.headers.get("X-Forwarded-For", request.remote_addr or "-")
        db.session.add(AuditLog(
            at=now_str(),
            user_name=u_name,
            user_sicil=u_sicil,
            action=action,
            details=details,
            ip_address=ip
        ))
        db.session.commit()
    except Exception as e:
        db.session.rollback()

# =========================================================
# YARDIMCILAR
# =========================================================
def now_str() -> str:
    return datetime.now().strftime("%d.%m.%Y %H:%M:%S")

def today_str() -> str:
    return date.today().strftime("%d.%m.%Y")

def require_login() -> bool:
    return bool(session.get("user"))


def is_admin() -> bool:
    return bool(session.get("sicil") == ADMIN_SICIL)

def normalize_spaces(s: str) -> str:
    if s is None:
        return ""
    return re.sub(r"\s+", " ", str(s)).strip()

def get_local_ip() -> str:
    import socket
    try:
        s = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
        s.connect(("8.8.8.8", 80))
        ip = s.getsockname()[0]
        s.close()
        return ip
    except Exception:
        return "127.0.0.1"


def seed_data():
    # Kullanıcılar
    if User.query.count() == 0:
        for name, sicil, pw in USERS:
            db.session.add(User(name=name, sicil=sicil, password_hash=generate_password_hash(pw)))
        db.session.commit()

    # Stok kalemleri: ekle + liste dışını sil
    for s in STOCK_ITEMS:
        if not Stock.query.filter_by(name=s).first():
            db.session.add(Stock(name=s, qty=0))

    for st in Stock.query.all():
        if st.name not in STOCK_ITEMS:
            db.session.delete(st)

    # Legacy DB temizlik (Eski birleşik personel isimlerini ayrıştır)
    try:
        for h in History.query.all():
            if h.alan_ad and "ARDA EKER" in h.alan_ad and "AHMET AYDIN" in h.alan_ad:
                h.alan_ad = "Ahmet AYDIN"
                h.alan_sicil = "187696"
        for it in InventoryItem.query.all():
            if it.assigned_name and "ARDA EKER" in it.assigned_name and "AHMET AYDIN" in it.assigned_name:
                it.assigned_name = "Ahmet AYDIN"
                it.assigned_sicil = "187696"
        db.session.commit()
    except Exception:
        pass

TR_MAP = str.maketrans({
    "İ":"I","ı":"i","Ş":"S","ş":"s","Ğ":"G","ğ":"g","Ü":"U","ü":"u","Ö":"O","ö":"o","Ç":"C","ç":"c"
})

def pdf_font_setup(p: FPDF) -> Tuple[str, bool]:
    if os.path.exists(FONT_PATH):
        try:
            p.add_font("DejaVu", "", FONT_PATH, uni=True)
            p.add_font("DejaVu", "B", FONT_PATH, uni=True)
            return "DejaVu", True
        except Exception:
            pass
    return "Helvetica", False

# Kurumsal PDF stil ayarları
CORP_HEADER_RGB = (0, 51, 102)  # koyu lacivert
CORP_TEXT_RGB = (0, 0, 0)       # siyah


def safe_text(s: str, unicode_ok: bool) -> str:
    s = "" if s is None else str(s)
    return s if unicode_ok else s.translate(TR_MAP)

def clip_text(p: FPDF, s: str, max_w: float) -> str:
    s = "" if s is None else str(s)
    if p.get_string_width(s) <= max_w:
        return s
    ell = "..."
    while s and p.get_string_width(s + ell) > max_w:
        s = s[:-1]
    return (s + ell) if s else ""

def parse_checked_ids(prefix: str) -> List[int]:
    ids = request.form.getlist(f"{prefix}_inv_ids")
    out: List[int] = []
    for x in ids:
        try:
            out.append(int(x))
        except Exception:
            pass
    return out

def parse_manual_devices() -> List[Dict]:
    raw = (request.form.get("manual_json") or "[]").strip()
    try:
        arr = json.loads(raw)
        return arr if isinstance(arr, list) else []
    except Exception:
        return []


# =========================================================
# Helper for Turkish name formatting
def format_tr_name(name: str) -> str:
    if not name:
        return ""
    words = []
    for w in str(name).strip().split():
        w_tr = w.translate(str.maketrans({"i":"İ", "ı":"I", "ş":"Ş", "ğ":"Ğ", "ü":"Ü", "ö":"Ö", "ç":"Ç"})).upper()
        words.append(w_tr)
    return " ".join(words)

def style_excel_sheet(ws):
    header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    data_font = Font(name="Calibri", size=10, color="000000")
    thin_border = Border(
        left=Side(style='thin', color='CBD5E1'),
        right=Side(style='thin', color='CBD5E1'),
        top=Side(style='thin', color='CBD5E1'),
        bottom=Side(style='thin', color='CBD5E1')
    )
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.font = data_font
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center")
    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = col[0].column_letter
        ws.column_dimensions[col_letter].width = max(max_len + 5, 14)

def get_unit_header_color(unit_name: str):
    unit_upper = (unit_name or "").upper()
    if "CEZA DAİRESİ" in unit_upper or "CEZA" in unit_upper:
        return (153, 27, 27), "991B1B"
    elif "HUKUK DAİRESİ" in unit_upper or "HUKUK" in unit_upper:
        return (30, 58, 138), "1E3A8A"
    elif "DURUŞMA SALONU" in unit_upper or "DURUŞMA" in unit_upper:
        return (15, 118, 110), "0F766E"
    else:
        return (51, 65, 85), "334155"


def export_history_xlsx(action_filter: str) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Zimmet Kayitlari" if action_filter == "ZIMMET" else "Teslim Tesellum"
    headers = [
        "Tarih / Saat", "İşlem", "Kategori", "Marka", "Model", "Seri No",
        "Zimmet Alan", "Sicil No", "Ünvan", "Birim", "Bilgisayar Adı",
        "Teslim Eden", "Teslim Sicil", "Teslim Ünvan"
    ]
    ws.append(headers)
    rows = History.query.filter_by(action=action_filter).order_by(History.id.asc()).all()

    # Daire/Birim bazında grupla
    grouped = {}
    for h in rows:
        unit_key = (h.alan_birim or "").strip() or "Bireysel / Diğer Kayıtlar"
        grouped.setdefault(unit_key, []).append(h)

    from openpyxl.styles import PatternFill, Font, Alignment
    current_row = 2

    for unit_name, unit_items in grouped.items():
        rgb, hex_code = get_unit_header_color(unit_name)
        banner_cell = ws.cell(row=current_row, column=1, value=f"  DAİRE / BİRİM: {unit_name.upper()} ({len(unit_items)} Kayıt)")
        banner_cell.fill = PatternFill(start_color=hex_code, end_color=hex_code, fill_type="solid")
        banner_cell.font = Font(color="FFFFFF", bold=True, size=11)
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=len(headers))
        current_row += 1

        for h in unit_items:
            ws.append([
                h.at or "", h.action or "", h.category or "", h.brand or "", h.model or "", h.serial_no or "",
                format_tr_name(h.alan_ad or ""), h.alan_sicil or "", h.alan_unvan or "", h.alan_birim or "", h.bilgisayar_adi or "",
                format_tr_name(h.teslim_ad or ""), h.teslim_sicil or "", h.teslim_unvan or "",
            ])
            current_row += 1

    style_excel_sheet(ws)
    bio = io.BytesIO()
    wb.save(bio)
    return bio.getvalue()


def export_history_pdf(action_filter: str, title: str) -> bytes:
    p = FPDF("L", "mm", "A4")
    p.set_auto_page_break(auto=True, margin=15)
    p.add_page()
    font, unicode_ok = pdf_font_setup(p)
    txt = lambda x: safe_text(x, unicode_ok)

    if os.path.exists(LOGO_PATH):
        p.image(LOGO_PATH, x=15, y=10, w=18, h=18)

    p.set_font(font, "B", 15)
    p.set_text_color(30, 41, 59)
    p.cell(0, 8, txt("T.C. DİYARBAKIR BÖLGE ADLİYE MAHKEMESİ"), ln=1, align="C")
    p.set_font(font, "B", 13)
    p.set_text_color(71, 85, 105)
    p.cell(0, 6, txt(f"{title} (Daire Gruplu Rapor)"), ln=1, align="C")
    p.ln(4)

    rows = History.query.filter_by(action=action_filter).order_by(History.id.asc()).all()
    if not rows:
        p.set_font(font, "", 10)
        p.set_text_color(100, 116, 139)
        p.cell(277, 8, txt("Henüz kayıtlı rapor bulunamadı."), 1, 1, "C")
    else:
        grouped = {}
        for h in rows:
            unit_key = (h.alan_birim or "").strip() or "Bireysel / Diğer Kayıtlar"
            grouped.setdefault(unit_key, []).append(h)

        for unit_name, unit_items in grouped.items():
            rgb, hex_code = get_unit_header_color(unit_name)
            p.set_font(font, "B", 10.5)
            p.set_fill_color(rgb[0], rgb[1], rgb[2])
            p.set_text_color(255, 255, 255)
            p.cell(277, 7.5, txt(f"  DAİRE / BİRİM: {unit_name.upper()} ({len(unit_items)} Kayıt)"), 1, 1, "L", fill=True)

            p.set_font(font, "B", 9)
            p.set_fill_color(30, 41, 59)
            p.set_text_color(255, 255, 255)
            headers = [("Tarih ve Saat", 42), ("İşlem Türü", 35), ("Personel (Ad Soyad / Sicil)", 65), ("Cihaz / Marka Model", 75), ("Seri Numarası", 60)]
            for t_head, w in headers:
                p.cell(w, 7, txt(t_head), 1, 0, "C", fill=True)
            p.ln(7)

            p.set_font(font, "", 8.5)
            for idx, h in enumerate(unit_items):
                person = f"{format_tr_name(h.alan_ad or '')} ({h.alan_sicil or ''})"
                dev_str = f"{h.category or ''} - {h.brand or ''} {h.model or ''}".strip(" -")
                action_name = "Zimmet Alındı" if h.action == "ZIMMET" else "Teslim Edildi"

                fill = idx % 2 == 1
                p.set_fill_color(248, 250, 252) if fill else p.set_fill_color(255, 255, 255)
                p.set_text_color(15, 23, 42)
                p.cell(42, 7, clip_text(p, txt(h.at or ""), 40), 1, 0, "C", fill=fill)
                p.cell(35, 7, txt(action_name), 1, 0, "C", fill=fill)
                p.cell(65, 7, clip_text(p, txt(person), 63), 1, 0, "L", fill=fill)
                p.cell(75, 7, clip_text(p, txt(dev_str), 73), 1, 0, "L", fill=fill)
                p.cell(60, 7, clip_text(p, txt(h.serial_no or ""), 58), 1, 1, "L", fill=fill)

            p.ln(4)

    out = p.output(dest="S")
    return out if isinstance(out, (bytes, bytearray)) else out.encode("latin-1", errors="ignore")

def export_envanter_xlsx() -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Envanter Kayitlari"
    headers = [
        "Kategori", "Marka", "Model", "Seri No", "Durum",
        "Kime Zimmetli (Ad / Sicil / Ünvan)", "Zimmet / İşlem Tarihi ve Saati"
    ]
    ws.append(headers)
    items = InventoryItem.query.order_by(InventoryItem.id.asc()).all()
    for it in items:
        who = "-"
        if it.status == "Zimmetli":
            who = f"{format_tr_name(it.assigned_name or '')} / {it.assigned_sicil or ''} / {it.assigned_title or ''}".strip(" /")
        dt_val = it.last_event_at or "-"
        ws.append([it.category or "", it.brand or "", it.model or "", it.serial_no or "", it.status or "", who, dt_val])
    style_excel_sheet(ws)
    bio = io.BytesIO()
    wb.save(bio)
    return bio.getvalue()

def export_envanter_pdf() -> bytes:
    p = FPDF("L", "mm", "A4")  # YATAY RAPOR (Landscape: 277 mm genişlik)
    p.set_auto_page_break(auto=True, margin=15)
    p.add_page()
    font, unicode_ok = pdf_font_setup(p)
    txt = lambda x: safe_text(x, unicode_ok)

    if os.path.exists(LOGO_PATH):
        p.image(LOGO_PATH, x=15, y=10, w=18, h=18)

    p.set_font(font, "B", 15)
    p.set_text_color(30, 41, 59)
    p.cell(0, 8, txt("T.C. DİYARBAKIR BÖLGE ADLİYE MAHKEMESİ"), ln=1, align="C")
    p.set_font(font, "B", 13)
    p.set_text_color(71, 85, 105)
    p.cell(0, 6, txt("Tüm Envanter Cihaz Raporu"), ln=1, align="C")
    p.ln(4)

    # Başlıklar (277 mm Toplam)
    p.set_font(font, "B", 9)
    p.set_fill_color(30, 41, 59)
    p.set_text_color(255, 255, 255)
    headers = [("Kategori", 32), ("Marka", 35), ("Model", 35), ("Seri No", 55), ("Durum", 25), ("Zimmetli Personel", 55), ("İşlem Tarihi", 40)]
    for title, w in headers:
        p.cell(w, 8, txt(title), 1, 0, "C", fill=True)
    p.ln(8)

    p.set_font(font, "", 8.5)
    items = InventoryItem.query.order_by(InventoryItem.id.asc()).all()
    if not items:
        p.set_text_color(100, 116, 139)
        p.cell(277, 8, txt("Envanterde kayıtlı cihaz bulunamadı."), 1, 1, "C")
    else:
        for idx, it in enumerate(items):
            who = f"{format_tr_name(it.assigned_name or '')}" if it.status == "Zimmetli" else "-"
            dt_val = it.last_event_at or "-"

            if idx % 2 == 1:
                p.set_fill_color(248, 250, 252)
                fill = True
            else:
                p.set_fill_color(255, 255, 255)
                fill = True

            p.set_text_color(15, 23, 42)
            p.cell(32, 7.5, clip_text(p, txt(it.category or ""), 30), 1, 0, "L", fill=fill)
            p.cell(35, 7.5, clip_text(p, txt(it.brand or ""), 33), 1, 0, "L", fill=fill)
            p.cell(35, 7.5, clip_text(p, txt(it.model or ""), 33), 1, 0, "L", fill=fill)
            p.cell(55, 7.5, clip_text(p, txt(it.serial_no or ""), 53), 1, 0, "L", fill=fill)
            p.cell(25, 7.5, txt(it.status or ""), 1, 0, "C", fill=fill)
            p.cell(55, 7.5, clip_text(p, txt(who), 53), 1, 0, "L", fill=fill)
            p.cell(40, 7.5, clip_text(p, txt(dt_val), 38), 1, 1, "C", fill=fill)

    out = p.output(dest="S")
    return out if isinstance(out, (bytes, bytearray)) else out.encode("latin-1", errors="ignore")

def export_depo_xlsx() -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Depo Stok Kayitlari"
    headers = ["Malzeme / Sarf Kalemi", "Mevcut Stok Adedi"]
    ws.append(headers)
    items = Stock.query.order_by(Stock.name.asc()).all()
    for st in items:
        ws.append([st.name or "", st.qty or 0])
    style_excel_sheet(ws)
    bio = io.BytesIO()
    wb.save(bio)
    return bio.getvalue()

def export_depo_pdf() -> bytes:
    p = FPDF("L", "mm", "A4")  # YATAY RAPOR (Landscape: 277 mm genişlik)
    p.set_auto_page_break(auto=True, margin=15)
    p.add_page()
    font, unicode_ok = pdf_font_setup(p)
    txt = lambda x: safe_text(x, unicode_ok)

    if os.path.exists(LOGO_PATH):
        p.image(LOGO_PATH, x=15, y=10, w=18, h=18)

    p.set_font(font, "B", 15)
    p.set_text_color(30, 41, 59)
    p.cell(0, 8, txt("T.C. DİYARBAKIR BÖLGE ADLİYE MAHKEMESİ"), ln=1, align="C")
    p.set_font(font, "B", 13)
    p.set_text_color(71, 85, 105)
    p.cell(0, 6, txt("Depo Sarf Malzeme Stok Raporu"), ln=1, align="C")
    p.ln(4)

    p.set_font(font, "B", 10)
    p.set_fill_color(30, 41, 59)
    p.set_text_color(255, 255, 255)
    p.cell(190, 8, txt("Malzeme / Sarf Kalemi Adı"), 1, 0, "L", fill=True)
    p.cell(87, 8, txt("Mevcut Stok Miktarı"), 1, 1, "C", fill=True)

    p.set_font(font, "", 9)
    items = Stock.query.order_by(Stock.name.asc()).all()
    if not items:
        p.set_text_color(100, 116, 139)
        p.cell(277, 8, txt("Depoda kayıtlı malzeme bulunamadı."), 1, 1, "C")
    else:
        for idx, st in enumerate(items):
            if idx % 2 == 1:
                p.set_fill_color(248, 250, 252)
                fill = True
            else:
                p.set_fill_color(255, 255, 255)
                fill = True
            p.set_text_color(15, 23, 42)
            p.cell(190, 7.5, clip_text(p, txt(st.name or ""), 185), 1, 0, "L", fill=fill)
            p.cell(87, 7.5, txt(f"{st.qty or 0} Adet"), 1, 1, "C", fill=fill)

    out = p.output(dest="S")
    return out if isinstance(out, (bytes, bytearray)) else out.encode("latin-1", errors="ignore")

    if os.path.exists(LOGO_PATH):
        p.image(LOGO_PATH, x=15, y=10, w=18, h=18)

    p.set_font(font, "B", 14)
    p.cell(0, 8, txt("T.C. DİYARBAKIR BÖLGE ADLİYE MAHKEMESİ"), ln=1, align="C")
    p.set_font(font, "B", 12)
    p.cell(0, 6, txt("Depo Sarf Malzeme Stok Raporu"), ln=1, align="C")
    p.ln(4)

    p.set_font(font, "B", 10)
    p.set_fill_color(30, 41, 59)
    p.set_text_color(255, 255, 255)
    p.cell(120, 8, txt("Malzeme / Sarf Kalemi Adı"), 1, 0, "L", fill=True)
    p.cell(60, 8, txt("Mevcut Stok Miktarı"), 1, 1, "C", fill=True)

    p.set_font(font, "", 9)
    p.set_text_color(0, 0, 0)
    items = Stock.query.order_by(Stock.name.asc()).all()
    for st in items:
        p.cell(120, 7, clip_text(p, txt(st.name or ""), 116), 1)
        p.cell(60, 7, txt(f"{st.qty or 0} Adet"), 1, 1, "C")

    out = p.output(dest="S")
    return out if isinstance(out, (bytes, bytearray)) else out.encode("latin-1", errors="ignore")

# =========================================================
# PDF: ZİMMET FİŞİ (fis_template.png üzerine yazar)
# =========================================================
def build_zimmet_fis_pdf(
    *,
    alan_ad: str,
    alan_sicil: str,
    alan_unvan: str,
    alan_birim: str,
    bilgisayar_adi: str,
    teslim_ad: str,
    teslim_sicil: str,
    teslim_unvan: str,
    devices: List[Dict],
) -> bytes:
    """
    fis_template.png A4 arka planı üzerine alanları yazar.
    NOT: Sen template'te SIRA NO rakamlarını sildiğin için burada clear/kapama yok.
    Bu yüzden tablo çizgileri asla "kaybolmaz".
    """
    p = FPDF("P", "mm", "A4")
    p.set_auto_page_break(auto=False)
    font, unicode_ok = pdf_font_setup(p)
    txt = lambda x: safe_text(x, unicode_ok)

    # Koordinatlar (senin şablona uyumlu)
    X_L = 10.1
    X_IL_SPLIT = 49.2
    X_MID_SPLIT = 139.1
    X_RLBL_SPLIT = 162.2
    X_R = 195.0

    Y_TITLE_BOT = 68.2
    Y_IL_BOT = 76.0
    Y_PERSON_HDR_BOT = 82.7
    Y_NAME_BOT = 89.7
    Y_UNVAN_BOT = 96.8

    X_SIRA_BOT = 30.9
    X_CINSI_BOT = 77.1
    X_MARKA_BOT = 133.6

    Y_TABLE_HDR_BOT = 111.3
    ROW_H = 8.6
    PER_PAGE = 8

    def put_in_cell(x0, y0, x1, y1, s, align="L", size=10, bold=False, pad=1.5):
        p.set_text_color(*(CORP_HEADER_RGB if bold else CORP_TEXT_RGB))
        p.set_font(font, "B" if bold else "", size)
        w = (x1 - x0) - 2 * pad
        h = (y1 - y0)
        s2 = clip_text(p, txt(s), w)
        p.set_xy(x0 + pad, y0)
        p.cell(w, h, s2, 0, 0, align)

    # Satırlar
    lines = []
    for d in devices or []:
        cat = normalize_spaces(d.get("category", "")).upper()
        brand = normalize_spaces(d.get("brand", ""))
        model = normalize_spaces(d.get("model", ""))
        ser = normalize_spaces(d.get("serial_no", "")) or "-"
        mm = normalize_spaces(f"{brand} {model}")
        lines.append({"cinsi": cat, "mm": mm, "seri": ser})

    pages = [lines[i:i+PER_PAGE] for i in range(0, len(lines), PER_PAGE)] or [[]]

    person_unvan_birim = normalize_spaces(f"{alan_unvan} - {alan_birim}").strip(" -")

    for page_items in pages:
        p.add_page()
        if os.path.exists(TEMPLATE_PATH):
            p.image(TEMPLATE_PATH, x=0, y=0, w=210, h=297)

        alan_ad_fmt = format_tr_name(alan_ad)
        teslim_ad_fmt = format_tr_name(teslim_ad)

        # Üst alanlar
        put_in_cell(X_IL_SPLIT, Y_TITLE_BOT, X_MID_SPLIT, Y_IL_BOT, "Diyarbakır", size=10)
        put_in_cell(X_RLBL_SPLIT, Y_TITLE_BOT, X_R, Y_IL_BOT, today_str(), size=10, align="C")

        put_in_cell(X_IL_SPLIT, Y_PERSON_HDR_BOT, X_MID_SPLIT, Y_NAME_BOT, alan_ad_fmt, size=10)
        put_in_cell(X_RLBL_SPLIT, Y_PERSON_HDR_BOT, X_R, Y_NAME_BOT, alan_sicil, size=10, align="C")

        put_in_cell(X_IL_SPLIT, Y_NAME_BOT, X_MID_SPLIT, Y_UNVAN_BOT, person_unvan_birim, size=10)
        put_in_cell(X_RLBL_SPLIT, Y_NAME_BOT, X_R, Y_UNVAN_BOT, bilgisayar_adi or "", size=10)

        # Tablo satırları
        for i in range(1, PER_PAGE + 1):
            row_top = Y_TABLE_HDR_BOT + (i - 1) * ROW_H
            row_bot = row_top + ROW_H
            item = page_items[i-1] if (i-1) < len(page_items) else {"cinsi": "", "mm": "", "seri": ""}

            put_in_cell(X_L, row_top, X_SIRA_BOT, row_bot, str(i), align="C", size=10, bold=True)
            put_in_cell(X_SIRA_BOT, row_top, X_CINSI_BOT, row_bot, item["cinsi"], size=10)
            put_in_cell(X_CINSI_BOT, row_top, X_MARKA_BOT, row_bot, item["mm"], size=10)
            put_in_cell(X_MARKA_BOT, row_top, X_R, row_bot, item["seri"], size=10)

        # İmza blokları (":" sonrası boşluğu minimum yaptık)
        y_sig0 = 210.5
        dy = 5.8

        X_SIG_L0, X_SIG_L1 = 31.0, 110.0  # sola daha yakın
        X_SIG_R0, X_SIG_R1 = 133.0, 200.0

        # Teslim eden
        put_in_cell(X_SIG_L0, y_sig0,         X_SIG_L1, y_sig0 + dy,   teslim_ad_fmt, size=10)
        put_in_cell(X_SIG_L0, y_sig0 + dy,    X_SIG_L1, y_sig0 + 2*dy, teslim_sicil, size=10)
        put_in_cell(X_SIG_L0, y_sig0 + 2*dy,  X_SIG_L1, y_sig0 + 3*dy, teslim_unvan, size=10)

        # Teslim alan
        put_in_cell(X_SIG_R0, y_sig0,         X_SIG_R1, y_sig0 + dy,   alan_ad_fmt, size=10)
        put_in_cell(X_SIG_R0, y_sig0 + dy,    X_SIG_R1, y_sig0 + 2*dy, alan_sicil, size=10)
        put_in_cell(X_SIG_R0, y_sig0 + 2*dy,  X_SIG_R1, y_sig0 + 3*dy, alan_unvan, size=10)

    out = p.output(dest="S")
    return out if isinstance(out, (bytes, bytearray)) else out.encode("latin-1", errors="ignore")

# =========================================================
# PDF: TESLİM TESELLÜM BELGESİ
# =========================================================
def build_teslim_tesellum_pdf(
    *,
    iade_eden_ad: str,
    iade_eden_sicil: str,
    teslim_alan_ad: str,
    teslim_alan_sicil: str,
    devices: List[Dict],
) -> bytes:
    p = FPDF("P", "mm", "A4")
    p.set_auto_page_break(auto=True, margin=18)
    p.add_page()

    font, unicode_ok = pdf_font_setup(p)
    txt = lambda x: safe_text(x, unicode_ok)

    # Antet ve Logolar (Sol ve Sağ Köşe)
    if os.path.exists(LOGO_PATH):
        p.image(LOGO_PATH, x=16, y=12, w=24, h=24)
        p.image(LOGO_PATH, x=170, y=12, w=24, h=24)

    # Başlık Metinleri (Kalın, Kurumsal Tasarım)
    p.set_xy(10, 14)
    p.set_text_color(15, 23, 42)  # Koyu Lacivert (Dark Corporate Navy)
    p.set_font(font, "B", 13)
    p.cell(0, 6, txt("T.C."), ln=1, align="C")
    p.set_font(font, "B", 15)
    p.cell(0, 7, txt("DİYARBAKIR BÖLGE ADLİYE MAHKEMESİ"), ln=1, align="C")
    p.set_font(font, "B", 12)
    p.set_text_color(51, 65, 85)
    p.cell(0, 6, txt("Bilgi İşlem Müdürlüğü"), ln=1, align="C")

    p.ln(6)
    # Doküman Ana Başlığı (Kalın, Büyük Kurumsal Başlık & Çizgi)
    p.set_text_color(15, 23, 42)
    p.set_font(font, "B", 17)
    p.cell(0, 9, txt("TESLİM - TESELLÜM BELGESİ"), ln=1, align="C")

    # Çizgi Ayırıcı
    p.set_draw_color(30, 41, 59)
    p.set_line_width(0.8)
    p.line(60, p.get_y() + 1, 150, p.get_y() + 1)
    p.ln(10)

    tarih = datetime.now().strftime("%d/%m/%Y")
    eden_sicil = normalize_spaces(iade_eden_sicil) or "……………"
    eden_ad_fmt = format_tr_name(iade_eden_ad) or "……………"
    teslim_ad_fmt = format_tr_name(teslim_alan_ad) or "……………"

    # Tutanak Paragrafı (SATIR BAŞI YAPILDI: x=22mm)
    p.set_font(font, "", 11)
    p.set_text_color(15, 23, 42)
    body = (
        f"Diyarbakır Bölge Adliye Mahkemesinde {eden_sicil} Sicil numaralı {eden_ad_fmt} (veya ilgili daire/büro) "
        f"kullanımına sunulmuş olan aşağıda detayları belirtilen bilgi işlem donanımlarının çalışır durumda teslim alındığına "
        f"dair düzenlenen işbu teslim tesellüm tutanağı birlikte imza altına alınmıştır."
    )
    p.set_x(22)  # SATIR BAŞI (Paragraph Indentation)
    p.multi_cell(0, 7.0, txt(body), align="L")

    p.ln(3)
    p.set_font(font, "B", 10.5)
    p.cell(0, 6, txt(f"Tarih: {tarih}"), ln=1, align="R")
    p.ln(6)

    # Donanım Listesi Tablosu
    p.set_font(font, "B", 12)
    p.set_text_color(15, 23, 42)
    p.cell(0, 7, txt("TESLİM EDİLEN DONANIMLAR :"), ln=1, align="L")
    p.ln(2)

    # Donanım Tablo Başlıkları
    p.set_font(font, "B", 10)
    p.set_fill_color(30, 41, 59)
    p.set_text_color(255, 255, 255)
    p.cell(45, 8.5, txt("CİNSİ / KATEGORİ"), 1, 0, "C", fill=True)
    p.cell(80, 8.5, txt("MARKA VE MODEL"), 1, 0, "C", fill=True)
    p.cell(65, 8.5, txt("SERİ NUMARASI"), 1, 1, "C", fill=True)

    p.set_font(font, "", 9.5)
    clean_devs = devices or []
    if not clean_devs:
        p.set_text_color(100, 116, 139)
        p.cell(190, 8.5, txt("Teslim edilen donanım bulunmamaktadır."), 1, 1, "C")
    else:
        for idx, d in enumerate(clean_devs):
            cat = normalize_spaces(d.get("category", "")).upper()
            brand = normalize_spaces(d.get("brand", ""))
            model = normalize_spaces(d.get("model", ""))
            serial = normalize_spaces(d.get("serial_no", "")) or "-"
            mm = normalize_spaces(f"{brand} {model}")

            if idx % 2 == 1:
                p.set_fill_color(248, 250, 252)
                fill = True
            else:
                p.set_fill_color(255, 255, 255)
                fill = True

            p.set_text_color(15, 23, 42)
            p.cell(45, 8, clip_text(p, txt(cat), 43), 1, 0, "L", fill=fill)
            p.cell(80, 8, clip_text(p, txt(mm), 78), 1, 0, "L", fill=fill)
            p.cell(65, 8, clip_text(p, txt(serial), 63), 1, 1, "L", fill=fill)

    # İmza Blokları
    p.ln(18)
    y_sig = p.get_y()
    col_w = 90
    gap = 10
    lx = 10
    rx = lx + col_w + gap

    # İmza Kutuları Başlığı (Hazirun / İade Eden)
    p.set_font(font, "B", 10)
    p.set_text_color(15, 23, 42)
    p.set_xy(lx, y_sig)
    p.cell(col_w, 7, txt("Teslim Eden (Hazirun / Birim Yetkilisi)"), align="C")
    p.set_xy(rx, y_sig)
    p.cell(col_w, 7, txt("Teslim Alan (Bilgi İşlem Müdürlüğü)"), align="C")

    # İmza Çizgisi
    y_line = y_sig + 14
    p.set_draw_color(30, 41, 59)
    p.set_line_width(0.5)
    p.line(lx + 10, y_line, lx + col_w - 10, y_line)
    p.line(rx + 10, y_line, rx + col_w - 10, y_line)

    # Ad Soyad & Sicil
    y_name = y_line + 4
    p.set_font(font, "B", 10.5)
    p.set_text_color(15, 23, 42)
    p.set_xy(lx, y_name)
    p.cell(col_w, 6, txt(eden_ad_fmt), align="C")
    p.set_xy(rx, y_name)
    p.cell(col_w, 6, txt(teslim_ad_fmt), align="C")

    y_sicil = y_name + 5.5
    p.set_font(font, "", 9.5)
    p.set_text_color(71, 85, 105)
    p.set_xy(lx, y_sicil)
    p.cell(col_w, 6, txt(f"Sicil No: {iade_eden_sicil or '-'}"), align="C")
    p.set_xy(rx, y_sicil)
    p.cell(col_w, 6, txt(f"Sicil No: {teslim_alan_sicil or '-'}"), align="C")

    out = p.output(dest="S")
    return out if isinstance(out, (bytes, bytearray)) else out.encode("latin-1", errors="ignore")


def build_personel_zimmet_karti_pdf(sicil: str) -> bytes:
    p = FPDF("P", "mm", "A4")
    p.set_auto_page_break(auto=True, margin=15)
    p.add_page()
    font, unicode_ok = pdf_font_setup(p)
    txt = lambda x: safe_text(x, unicode_ok)

    if os.path.exists(LOGO_PATH):
        p.image(LOGO_PATH, x=16, y=12, w=24, h=24)
        p.image(LOGO_PATH, x=170, y=12, w=24, h=24)

    # Antet
    p.set_xy(10, 14)
    p.set_text_color(15, 23, 42)
    p.set_font(font, "B", 13)
    p.cell(0, 6, txt("T.C."), ln=1, align="C")
    p.set_font(font, "B", 15)
    p.cell(0, 7, txt("DİYARBAKIR BÖLGE ADLİYE MAHKEMESİ"), ln=1, align="C")
    p.set_font(font, "B", 12)
    p.set_text_color(51, 65, 85)
    p.cell(0, 6, txt("Bilgi İşlem Müdürlüğü"), ln=1, align="C")

    p.ln(6)
    p.set_text_color(15, 23, 42)
    p.set_font(font, "B", 17)
    p.cell(0, 9, txt("PERSONEL ZİMMET VE DONANIM KARTI"), ln=1, align="C")
    p.set_draw_color(30, 41, 59)
    p.set_line_width(0.8)
    p.line(55, p.get_y() + 1, 155, p.get_y() + 1)
    p.ln(10)

    # Personel Bilgisi Bul
    active_items = InventoryItem.query.filter_by(assigned_sicil=sicil, status="Zimmetli").all()
    history_items = History.query.filter_by(alan_sicil=sicil).order_by(History.id.desc()).all()
    user_rec = User.query.filter_by(sicil=sicil).first()

    person_name = ""
    person_title = ""
    person_unit = ""
    if user_rec:
        person_name = format_tr_name(user_rec.name)
    elif active_items and active_items[0].assigned_name:
        person_name = format_tr_name(active_items[0].assigned_name)
    elif history_items and history_items[0].alan_ad:
        person_name = format_tr_name(history_items[0].alan_ad)
    else:
        person_name = f"Sicil No: {sicil}"

    if active_items:
        person_title = active_items[0].assigned_title or ""
        person_unit = active_items[0].assigned_unit or ""
    elif history_items:
        person_title = history_items[0].alan_unvan or ""
        person_unit = history_items[0].alan_birim or ""

    # Personel Bilgi Kartı Box
    p.set_fill_color(248, 250, 252)
    p.set_draw_color(203, 213, 225)
    p.rect(10, p.get_y(), 190, 24, style="FD")

    cur_y = p.get_y() + 3
    p.set_xy(14, cur_y)
    p.set_font(font, "B", 10.5)
    p.set_text_color(15, 23, 42)
    p.cell(100, 6, txt(f"Personel Adı Soyadı : {person_name}"), 0, 0)
    p.cell(80, 6, txt(f"Sicil No : {sicil}"), 0, 1)

    p.set_xy(14, cur_y + 6.5)
    p.set_font(font, "", 10)
    p.set_text_color(51, 65, 85)
    p.cell(100, 6, txt(f"Ünvan / Birim : {person_title} {person_unit}".strip()), 0, 0)
    p.cell(80, 6, txt(f"Rapor Tarihi : {datetime.now().strftime('%d/%m/%Y')}"), 0, 1)

    p.ln(14)

    # 1. BÖLÜM: Halen Üzerinde Bulunan Zimmetli Cihazlar
    p.set_font(font, "B", 11.5)
    p.set_text_color(15, 23, 42)
    p.cell(0, 6, txt("1. HÂLEN ÜZERİNDE BULUNAN AKTİF ZİMMETLİ CİHAZLAR"), ln=1)
    p.ln(2)

    p.set_font(font, "B", 9.5)
    p.set_fill_color(30, 41, 59)
    p.set_text_color(255, 255, 255)
    p.cell(35, 8, txt("KATEGORİ"), 1, 0, "C", fill=True)
    p.cell(75, 8, txt("MARKA VE MODEL"), 1, 0, "C", fill=True)
    p.cell(45, 8, txt("SERİ NUMARASI"), 1, 0, "C", fill=True)
    p.cell(35, 8, txt("ZİMMET TARİHİ"), 1, 1, "C", fill=True)

    p.set_font(font, "", 9)
    if not active_items:
        p.set_text_color(100, 116, 139)
        p.cell(190, 8, txt("Halen üzerinde zimmetli aktif cihaz bulunmamaktadır."), 1, 1, "C")
    else:
        for idx, it in enumerate(active_items):
            fill = idx % 2 == 1
            p.set_fill_color(248, 250, 252) if fill else p.set_fill_color(255, 255, 255)
            p.set_text_color(15, 23, 42)
            mm = f"{it.brand or ''} {it.model or ''}".strip()
            p.cell(35, 7.5, clip_text(p, txt(it.category or ""), 33), 1, 0, "L", fill=fill)
            p.cell(75, 7.5, clip_text(p, txt(mm), 73), 1, 0, "L", fill=fill)
            p.cell(45, 7.5, clip_text(p, txt(it.serial_no or ""), 43), 1, 0, "L", fill=fill)
            p.cell(35, 7.5, clip_text(p, txt(it.assigned_at or "-"), 33), 1, 1, "C", fill=fill)

    p.ln(10)

    # 2. BÖLÜM: Geçmiş İşlem / Teslim Tesellüm Geçmişi
    p.set_font(font, "B", 11)
    p.set_text_color(15, 23, 42)
    p.cell(0, 6, txt("2. GEÇMİŞ ZİMMET VE TESLİM İŞLEM GEÇMİŞİ"), ln=1)
    p.ln(1)

    p.set_font(font, "B", 9)
    p.set_fill_color(30, 41, 59)
    p.set_text_color(255, 255, 255)
    p.cell(35, 7, txt("İŞLEM TARİHİ"), 1, 0, "C", fill=True)
    p.cell(30, 7, txt("İŞLEM TÜRÜ"), 1, 0, "C", fill=True)
    p.cell(75, 7, txt("CİHAZ / MARKA MODEL"), 1, 0, "C", fill=True)
    p.cell(50, 7, txt("SERİ NUMARASI"), 1, 1, "C", fill=True)

    p.set_font(font, "", 8.5)
    if not history_items:
        p.set_text_color(100, 116, 139)
        p.cell(190, 7, txt("Geçmiş işlem kaydı bulunamadı."), 1, 1, "C")
    else:
        for idx, h in enumerate(history_items[:14]):
            fill = idx % 2 == 1
            p.set_fill_color(248, 250, 252) if fill else p.set_fill_color(255, 255, 255)
            p.set_text_color(15, 23, 42)
            act_lbl = "Zimmet Alındı" if h.action == "ZIMMET" else "Teslim Edildi"
            dev_str = f"{h.category or ''} - {h.brand or ''} {h.model or ''}".strip(" -")
            p.cell(35, 6.2, clip_text(p, txt(h.at or ""), 33), 1, 0, "C", fill=fill)
            p.cell(30, 6.2, txt(act_lbl), 1, 0, "C", fill=fill)
            p.cell(75, 6.2, clip_text(p, txt(dev_str), 73), 1, 0, "L", fill=fill)
            p.cell(50, 6.2, clip_text(p, txt(h.serial_no or ""), 48), 1, 1, "L", fill=fill)

    # İmza Bloğu (Sayfa taşmasını engelle)
    p.set_auto_page_break(auto=False)
    if p.get_y() > 235:
        p.add_page()
    else:
        p.ln(6)

    y_sig = p.get_y()
    col_w = 90
    gap = 10
    lx = 10
    rx = lx + col_w + gap

    p.set_font(font, "B", 10)
    p.set_text_color(15, 23, 42)
    p.set_xy(lx, y_sig)
    p.cell(col_w, 6, txt("Düzenleyen (Bilgi İşlem)"), align="C")
    p.set_xy(rx, y_sig)
    p.cell(col_w, 6, txt("Personel İmza"), align="C")

    y_line = y_sig + 12
    p.set_draw_color(30, 41, 59)
    p.set_line_width(0.5)
    p.line(lx + 10, y_line, lx + col_w - 10, y_line)
    p.line(rx + 10, y_line, rx + col_w - 10, y_line)

def build_daire_envanter_defteri_pdf(daire_adi: str) -> bytes:
    p = FPDF("L", "mm", "A4")  # YATAY (LANDSCAPE) 297mm GENİŞLİK
    p.set_auto_page_break(auto=True, margin=15)
    p.add_page()
    font, unicode_ok = pdf_font_setup(p)
    txt = lambda x: safe_text(x, unicode_ok)

    if os.path.exists(LOGO_PATH):
        p.image(LOGO_PATH, x=16, y=10, w=24, h=24)
        p.image(LOGO_PATH, x=257, y=10, w=24, h=24)

    # Antet
    p.set_xy(10, 12)
    p.set_text_color(15, 23, 42)
    p.set_font(font, "B", 13)
    p.cell(0, 6, txt("T.C."), ln=1, align="C")
    p.set_font(font, "B", 15)
    p.cell(0, 7, txt("DİYARBAKIR BÖLGE ADLİYE MAHKEMESİ"), ln=1, align="C")
    p.set_font(font, "B", 12)
    p.set_text_color(51, 65, 85)
    p.cell(0, 6, txt("Bilgi İşlem Müdürlüğü"), ln=1, align="C")

    p.ln(5)
    p.set_text_color(15, 23, 42)
    p.set_font(font, "B", 16)
    p.cell(0, 8, txt(f"{daire_adi.upper()} ENVANTER DEFTERİ"), ln=1, align="C")
    p.set_draw_color(30, 41, 59)
    p.set_line_width(0.8)
    p.line(60, p.get_y() + 1, 237, p.get_y() + 1)
    p.ln(8)

    # Daireye ait aktif cihazlar
    all_invs = InventoryItem.query.all()
    items = [it for it in all_invs if (it.assigned_unit and it.assigned_unit.strip().lower() == daire_adi.strip().lower())]
    if not items:
        # History kaydında son zimmet birimi bu daire olanları bul
        hist_serials = [h.serial_no for h in History.query.filter(History.alan_birim.ilike(f"%{daire_adi}%")).all()]
        items = [it for it in all_invs if it.serial_no in hist_serials]

    # Bilgi Kutusu (Yatay Format 277mm Tam Genişlik)
    box_start_y = p.get_y()
    p.set_fill_color(248, 250, 252)
    p.set_draw_color(203, 213, 225)
    p.rect(10, box_start_y, 277, 18, style="FD")

    p.set_xy(14, box_start_y + 2.5)
    p.set_font(font, "B", 10.5)
    p.set_text_color(15, 23, 42)
    p.cell(170, 6, txt(f"Birim / Daire Adı : {daire_adi}"), 0, 0)
    p.cell(100, 6, txt(f"Rapor Tarihi : {datetime.now().strftime('%d/%m/%Y')}"), 0, 1, align="R")

    p.set_xy(14, box_start_y + 9.5)
    p.set_font(font, "", 10)
    p.set_text_color(51, 65, 85)
    p.cell(270, 6, txt(f"Dairede Kayıtlı Toplam Cihaz Sayısı : {len(items)} Adet"), 0, 1)

    p.set_y(box_start_y + 22)

    # Tablo Başlıkları (Yatay Format - Toplam 277mm Genişlik)
    p.set_font(font, "B", 9.5)
    p.set_fill_color(30, 41, 59)
    p.set_text_color(255, 255, 255)
    p.cell(14, 7.5, txt("S.NO"), 1, 0, "C", fill=True)
    p.cell(38, 7.5, txt("KATEGORİ"), 1, 0, "C", fill=True)
    p.cell(75, 7.5, txt("MARKA VE MODEL"), 1, 0, "C", fill=True)
    p.cell(55, 7.5, txt("SERİ NUMARASI"), 1, 0, "C", fill=True)
    p.cell(95, 7.5, txt("ZİMMETLENEN MAKAM / VERİLEN YER"), 1, 1, "C", fill=True)

    p.set_font(font, "", 9)
    if not items:
        p.set_text_color(100, 116, 139)
        p.cell(277, 8, txt("Bu daireye zimmetli aktif cihaz bulunamadı."), 1, 1, "C")
    else:
        for idx, it in enumerate(items, start=1):
            fill = idx % 2 == 1
            p.set_fill_color(248, 250, 252) if fill else p.set_fill_color(255, 255, 255)
            p.set_text_color(15, 23, 42)
            mm = f"{it.brand or ''} {it.model or ''}".strip()
            who_str = (it.assigned_title or it.assigned_name or "-")
            p.cell(14, 7, str(idx), 1, 0, "C", fill=fill)
            p.cell(38, 7, clip_text(p, txt(it.category or ""), 36), 1, 0, "L", fill=fill)
            p.cell(75, 7, clip_text(p, txt(mm), 73), 1, 0, "L", fill=fill)
            p.cell(55, 7, clip_text(p, txt(it.serial_no or ""), 53), 1, 0, "L", fill=fill)
            p.cell(95, 7, clip_text(p, txt(who_str), 93), 1, 1, "L", fill=fill)

    out = p.output(dest="S")
    return out if isinstance(out, (bytes, bytearray)) else out.encode("latin-1", errors="ignore")


def build_hek_hurda_pdf(devices: List[Dict], tutanak_no: str = "", tutanak_tarihi: str = "", gerekce_text: str = "") -> bytes:
    p = FPDF("P", "mm", "A4")
    p.set_auto_page_break(auto=True, margin=15)
    p.add_page()
    font, unicode_ok = pdf_font_setup(p)
    txt = lambda x: safe_text(x, unicode_ok)

    if os.path.exists(LOGO_PATH):
        p.image(LOGO_PATH, x=15, y=10, w=22, h=22)
        p.image(LOGO_PATH, x=173, y=10, w=22, h=22)

    p.set_xy(10, 11)
    p.set_text_color(15, 23, 42)
    p.set_font(font, "B", 12)
    p.cell(0, 5.5, txt("T.C."), ln=1, align="C")
    p.set_font(font, "B", 14)
    p.cell(0, 6.5, txt("DİYARBAKIR BÖLGE ADLİYE MAHKEMESİ"), ln=1, align="C")
    p.set_font(font, "B", 11)
    p.set_text_color(51, 65, 85)
    p.cell(0, 5.5, txt("Bilgi İşlem Müdürlüğü"), ln=1, align="C")

    p.ln(4)
    p.set_text_color(15, 23, 42)
    p.set_font(font, "B", 13)
    p.cell(0, 7, txt("TEKNİK DEĞERLENDİRME VE HEK / HURDA AYIRMA TUTANAĞI"), ln=1, align="C")
    p.set_draw_color(30, 41, 59)
    p.set_line_width(0.8)
    p.line(25, p.get_y() + 1, 185, p.get_y() + 1)
    p.ln(6)

    t_tarih = tutanak_tarihi or today_str()
    p.set_font(font, "B", 9)
    p.set_text_color(15, 23, 42)
    p.cell(0, 5, txt(f"Tutanak Tarihi: {t_tarih}"), 0, 1, "R")
    p.ln(2)

    p.set_font(font, "B", 10)
    p.set_text_color(15, 23, 42)
    p.cell(0, 6, txt("I. GEREKÇE VE TEKNİK DEĞERLENDİRME"), ln=1)
    p.set_font(font, "", 8.5)
    p.set_text_color(51, 65, 85)
    g_text = gerekce_text or (
        "Diyarbakır Bölge Adliye Mahkemesi bünyesinde kullanılmakta iken arızalanan ve aşağıda marka, model ile "
        "seri numaraları belirtilen bilgi işlem donanımları, Bilgi İşlem Müdürlüğü teknik personellerince fiziki ve teknik "
        "incelemeye tabi tutulmıştır. Yapılan incelemeler neticesinde; söz konusu donanımların kullanım ömürlerini "
        "(ekonomik ömrünü) tamamladığı, tamir maliyetlerinin güncel donanım değerini aştığı / yedek parça temininin imkansız "
        "olduğu tespit edilmiş olup, kamu yararı ve tasarruf tedbirleri gözetilerek HEK (Kullanılamaz/Hurda) durumuna ayrılmasına karar verilmiştir."
    )
    p.multi_cell(0, 4.5, txt(g_text))
    p.ln(4)

    p.set_font(font, "B", 10)
    p.set_text_color(15, 23, 42)
    p.cell(0, 6, txt("II. HEK / HURDAYA AYRILAN DONANIM LİSTESİ"), ln=1)
    p.ln(1)

    p.set_font(font, "B", 8.5)
    p.set_fill_color(30, 41, 59)
    p.set_text_color(255, 255, 255)
    p.cell(10, 6.5, txt("No"), 1, 0, "C", fill=True)
    p.cell(30, 6.5, txt("Kategori"), 1, 0, "C", fill=True)
    p.cell(50, 6.5, txt("Marka / Model"), 1, 0, "C", fill=True)
    p.cell(45, 6.5, txt("Seri Numarası"), 1, 0, "C", fill=True)
    p.cell(55, 6.5, txt("Arıza / Hek Gerekçesi"), 1, 1, "C", fill=True)

    p.set_font(font, "", 8)
    p.set_text_color(15, 23, 42)
    if not devices:
        p.cell(190, 6.5, txt("Heke ayrılan kayıtlı cihaz bulunamadı."), 1, 1, "C")
    else:
        for idx, d in enumerate(devices, 1):
            fill = idx % 2 == 0
            p.set_fill_color(248, 250, 252) if fill else p.set_fill_color(255, 255, 255)
            cat = d.get("category") or "Donanım"
            brand_mdl = f"{d.get('brand','') } {d.get('model','')}".strip()
            ser = d.get("serial_no") or "-"
            reas = d.get("reason") or "Ekonomik Ömrü Dolmuş / Arızalı"
            
            p.cell(10, 6, str(idx), 1, 0, "C", fill=fill)
            p.cell(30, 6, clip_text(p, txt(cat), 28), 1, 0, "L", fill=fill)
            p.cell(50, 6, clip_text(p, txt(brand_mdl), 48), 1, 0, "L", fill=fill)
            p.cell(45, 6, clip_text(p, txt(ser), 43), 1, 0, "L", fill=fill)
            p.cell(55, 6, clip_text(p, txt(reas), 53), 1, 1, "L", fill=fill)

    p.ln(4)
    p.set_font(font, "B", 9.5)
    p.cell(0, 5, txt("III. SONUÇ VE İMZA"), ln=1)
    p.set_font(font, "", 8)
    p.set_text_color(71, 85, 105)
    p.multi_cell(0, 4, txt("İşbu tutanak, yukarıda detayları verilen donanımların envanter kayıtlarından 'Hek / Hurda' statüsüne geçirilerek Taşınır Mal Yönetmeliği hükümlerince imha / hurda deposuna devir işlemlerinin başlatılması amacıyla tanzim edilmiş ve birlikte imza altına alınmıştır."))

    p.set_auto_page_break(auto=False)
    if p.get_y() > 235:
        p.add_page()
    else:
        p.ln(6)

    y_sig = p.get_y()
    col_w = 58
    gap = 8
    x1 = 10
    x2 = x1 + col_w + gap
    x3 = x2 + col_w + gap

    p.set_font(font, "B", 8.5)
    p.set_text_color(15, 23, 42)
    p.set_xy(x1, y_sig)
    p.cell(col_w, 5, txt("İnceleyen Teknik Personel"), align="C")
    p.set_xy(x2, y_sig)
    p.cell(col_w, 5, txt("Bilgi İşlem Şefi"), align="C")
    p.set_xy(x3, y_sig)
    p.cell(col_w, 5, txt("İdari İşler Müdürü"), align="C")

    p.set_font(font, "", 8)
    p.set_xy(x1, y_sig + 5)
    p.cell(col_w, 4, txt("Murat İRVEN (187665)"), align="C")
    p.set_xy(x2, y_sig + 5)
    p.cell(col_w, 4, txt("Ahmet AYDIN (187696)"), align="C")
    p.set_xy(x3, y_sig + 5)
    p.cell(col_w, 4, txt("Necip Fazıl EKER (86537)"), align="C")

    y_line = y_sig + 16
    p.set_draw_color(30, 41, 59)
    p.set_line_width(0.4)
    p.line(x1 + 5, y_line, x1 + col_w - 5, y_line)
    p.line(x2 + 5, y_line, x2 + col_w - 5, y_line)
    p.line(x3 + 5, y_line, x3 + col_w - 5, y_line)

    out = p.output(dest="S")
    return out if isinstance(out, (bytes, bytearray)) else out.encode("latin-1", errors="ignore")


def purge_dirty_concat_records():
    try:
        # DB'de kayıtlı kirli/birleşik isim kullanıcılarını temizle
        dirty_users = User.query.filter(
            (User.name.like("%ARDA EKER%")) | 
            (User.name.like("%AHMET AYDIN%")) | 
            (User.name.like("%Sicil:%"))
        ).all()
        for u in dirty_users:
            db.session.delete(u)
        db.session.commit()

        # Temiz tekil kullanıcıları ekle
        clean_defaults = [
            ("Ahmet AYDIN", "187696"),
            ("Arda EKER", "327360"),
            ("Murat İRVEN", "187665"),
        ]
        for name, sicil in clean_defaults:
            if not User.query.filter_by(sicil=sicil).first():
                db.session.add(User(name=name, sicil=sicil, password_hash=generate_password_hash("123456")))
        db.session.commit()
    except Exception:
        db.session.rollback()


def get_all_personnel_list():
    res = {}
    try:
        purge_dirty_concat_records()
        for u in User.query.all():
            if u.sicil and str(u.sicil).strip():
                s = str(u.sicil).strip()
                n = format_tr_name(u.name)
                if len(n) < 40 and "AHMET AYDIN" not in n and "ARDA EKER" not in n and "Sicil:" not in n:
                    res[s] = n

        defaults = [
            ("187696", "Ahmet AYDIN"),
            ("327360", "Arda EKER"),
            ("187665", "Murat İRVEN"),
        ]
        for s, n in defaults:
            if s not in res:
                res[s] = n
    except Exception:
        pass
    return sorted(res.items(), key=lambda x: x[1])


def build_daire_makam_zimmet_pdf(
    daire_adi: str,
    makam_title: str,
    yazi_isleri_muduru_ad: str,
    yazi_isleri_muduru_sicil: str,
    teslim_alan_ad: str,
    teslim_alan_sicil: str,
    devices: List[Dict]
) -> bytes:
    p = FPDF("P", "mm", "A4")
    p.set_auto_page_break(auto=True, margin=15)
    p.add_page()
    font, unicode_ok = pdf_font_setup(p)
    txt = lambda x: safe_text(x, unicode_ok)

    if os.path.exists(LOGO_PATH):
        p.image(LOGO_PATH, x=16, y=12, w=24, h=24)
        p.image(LOGO_PATH, x=170, y=12, w=24, h=24)

    # Antet
    p.set_xy(10, 14)
    p.set_text_color(15, 23, 42)
    p.set_font(font, "B", 13)
    p.cell(0, 6, txt("T.C."), ln=1, align="C")
    p.set_font(font, "B", 15)
    p.cell(0, 7, txt("DİYARBAKIR BÖLGE ADLİYE MAHKEMESİ"), ln=1, align="C")
    p.set_font(font, "B", 12)
    p.set_text_color(51, 65, 85)
    p.cell(0, 6, txt("Bilgi İşlem Müdürlüğü"), ln=1, align="C")

    p.ln(6)
    p.set_text_color(15, 23, 42)
    p.set_font(font, "B", 16)
    p.cell(0, 8, txt("DAİRE / MAKAM ZİMMET VE TESLİM TUTANAĞI"), ln=1, align="C")
    p.set_draw_color(30, 41, 59)
    p.set_line_width(0.8)
    p.line(40, p.get_y() + 1, 170, p.get_y() + 1)
    p.ln(9)

    # Kurumsal Bilgi Kutusu (32mm yüksekliğinde 3 satırlı temiz yapı - Belge No kaldırıldı)
    box_start_y = p.get_y()
    p.set_fill_color(248, 250, 252)
    p.set_draw_color(203, 213, 225)
    p.rect(10, box_start_y, 190, 26, style="FD")

    cur_y = box_start_y + 3.5
    p.set_xy(14, cur_y)
    p.set_font(font, "B", 10.5)
    p.set_text_color(15, 23, 42)
    p.cell(115, 6, clip_text(p, txt(f"Teslim Edilen Daire / Birim : {daire_adi}"), 112), 0, 0)
    p.cell(65, 6, txt(f"Tarih : {datetime.now().strftime('%d/%m/%Y')}"), 0, 1, align="R")

    cur_y += 7.0
    p.set_xy(14, cur_y)
    p.set_font(font, "B", 10.5)
    p.cell(180, 6, clip_text(p, txt(f"Zimmetlenen Makam / Envanter : {makam_title}"), 175), 0, 1)

    cur_y += 7.0
    p.set_xy(14, cur_y)
    p.set_font(font, "", 10)
    p.set_text_color(51, 65, 85)
    p.cell(180, 6, txt(f"Daire Yazı İşleri Müdürü : {format_tr_name(yazi_isleri_muduru_ad)} (Sicil: {yazi_isleri_muduru_sicil or '-'})"), 0, 1)

    p.set_y(box_start_y + 31)

    # Beyan Metni (Satır Başı sekme girintisi)
    p.set_font(font, "", 10.5)
    p.set_text_color(15, 23, 42)
    p.set_x(22)
    desc_txt = f"Diyarbakır Bölge Adliye Mahkemesi {daire_adi} bünyesinde {makam_title} kullanımına sunulmak üzere, aşağıda marka, model ve seri numarası belirtilen bilgi işlem donanımları çalışır durumda Daire Yazı İşleri Müdürü {format_tr_name(yazi_isleri_muduru_ad)} hazır olduğu halde makam odasına, kullanılmak üzere bırakılmış olup işbu tutanak tanzim edilerek birlikte imza altına alınmıştır."
    p.multi_cell(178, 6, txt(desc_txt), align="J")
    p.ln(6)

    # Cihazlar Tablosu
    p.set_font(font, "B", 9.5)
    p.set_fill_color(30, 41, 59)
    p.set_text_color(255, 255, 255)
    p.cell(15, 8, txt("S.NO"), 1, 0, "C", fill=True)
    p.cell(40, 8, txt("KATEGORİ"), 1, 0, "C", fill=True)
    p.cell(75, 8, txt("MARKA VE MODEL"), 1, 0, "C", fill=True)
    p.cell(60, 8, txt("SERİ NUMARASI"), 1, 1, "C", fill=True)

    p.set_font(font, "", 9)
    for idx, d in enumerate(devices, start=1):
        fill = idx % 2 == 0
        p.set_fill_color(248, 250, 252) if fill else p.set_fill_color(255, 255, 255)
        p.set_text_color(15, 23, 42)
        mm = f"{d.get('brand','') or ''} {d.get('model','') or ''}".strip()
        p.cell(15, 7.5, str(idx), 1, 0, "C", fill=fill)
        p.cell(40, 7.5, clip_text(p, txt(d.get('category','') or ""), 38), 1, 0, "L", fill=fill)
        p.cell(75, 7.5, clip_text(p, txt(mm), 73), 1, 0, "L", fill=fill)
        p.cell(60, 7.5, clip_text(p, txt(d.get('serial_no','') or ""), 58), 1, 1, "L", fill=fill)

    # İmza Bloğu (Hazirun)
    p.ln(16)
    y_sig = p.get_y()
    col_w = 90
    gap = 10
    lx = 10
    rx = lx + col_w + gap

    p.set_font(font, "B", 10)
    p.set_text_color(15, 23, 42)
    p.set_xy(lx, y_sig)
    p.cell(col_w, 6, txt("Teslim Eden (Bilgi İşlem Müdürlüğü)"), align="C")
    p.set_xy(rx, y_sig)
    p.cell(col_w, 6, txt("Hazirun (Daire Yazı İşleri Müdürü)"), align="C")

    y_line = y_sig + 12
    p.set_draw_color(30, 41, 59)
    p.set_line_width(0.5)
    p.line(lx + 10, y_line, lx + col_w - 10, y_line)
    p.line(rx + 10, y_line, rx + col_w - 10, y_line)

    y_name = y_line + 3
    p.set_font(font, "B", 10)
    p.set_xy(lx, y_name)
    p.cell(col_w, 6, txt(format_tr_name(teslim_alan_ad)), align="C")
    p.set_xy(rx, y_name)
    p.cell(col_w, 6, txt(format_tr_name(yazi_isleri_muduru_ad)), align="C")

    y_sicil = y_name + 5.5
    p.set_font(font, "", 9.5)
    p.set_text_color(71, 85, 105)
    p.set_xy(lx, y_sicil)
    p.cell(col_w, 6, txt(f"Sicil No: {teslim_alan_sicil or '-'}"), align="C")
    p.set_xy(rx, y_sicil)
    p.cell(col_w, 6, txt(f"Sicil No: {yazi_isleri_muduru_sicil or '-'}"), align="C")

    out = p.output(dest="S")
    return out if isinstance(out, (bytes, bytearray)) else out.encode("latin-1", errors="ignore")


# =========================================================
# OTOMATİK ZAMANLANMIŞ YEDEKLEME DÖNGÜSÜ (Her gün 17:00)
# =========================================================
_LAST_AUTO_BACKUP_DATE = ""

def auto_scheduled_backup_loop():
    global _LAST_AUTO_BACKUP_DATE
    while True:
        try:
            now = datetime.now()
            today = now.strftime("%Y-%m-%d")
            # Her gün saat 17:00'de 1 kez çalışır
            if now.hour == 17 and _LAST_AUTO_BACKUP_DATE != today:
                with app.app_context():
                    fname = create_db_backup(prefix="otomatik_1700")
                    print(f"[ZAMANLANMIŞ YEDEK 17:00] Başarıyla yedek oluşturuldu: {fname}")
                _LAST_AUTO_BACKUP_DATE = today
        except Exception as e:
            print(f"[ZAMANLANMIŞ YEDEK HATA] {e}")
        time.sleep(30)

threading.Thread(target=auto_scheduled_backup_loop, daemon=True).start()

# =========================================================
# HTML (RAW STRING! -> JS asla Python dışına taşmaz)
# =========================================================
BASE_HTML = r"""<!doctype html>
<html lang="tr" data-theme="dark">
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>Diyarbakır Bölge Adliye Mahkemesi Envanter Sistemi</title>
<link href="https://cdn.jsdelivr.net/npm/bootstrap@5.3.2/dist/css/bootstrap.min.css" rel="stylesheet">
<link href="https://cdnjs.cloudflare.com/ajax/libs/font-awesome/6.5.1/css/all.min.css" rel="stylesheet">
<link href="https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600;700;800;900&display=swap" rel="stylesheet">
<script src="https://cdn.jsdelivr.net/npm/chart.js@4.4.1/dist/chart.umd.min.js"></script>
<style>
:root, html[data-theme="dark"] {
  --bg: #090d16;
  --card: #121827;
  --sidebar-bg: #0d121f;
  --border: #1e293b;
  --muted: #94a3b8;
  --text: #f8fafc;
  --heading: #ffffff;
  --primary: #6366f1;
  --accent: #6366f1;
  --input-bg: #090d16;
  --input-text: #ffffff;
}

html[data-theme="light"] {
  --bg: #f8fafc;
  --card: #ffffff;
  --sidebar-bg: #ffffff;
  --border: #e2e8f0;
  --muted: #64748b;
  --text: #0f172a;
  --heading: #0f172a;
  --primary: #4f46e5;
  --accent: #4f46e5;
  --input-bg: #ffffff;
  --input-text: #0f172a;
}

html[data-theme="navy"] {
  --bg: #1c132b;
  --card: #2d1e40;
  --sidebar-bg: #160d23;
  --border: #4a2c6d;
  --muted: #d8b4fe;
  --text: #ffffff;
  --heading: #ffffff;
  --primary: #f97316;
  --accent: #ec4899;
  --input-bg: #160d23;
  --input-text: #ffffff;
}

html[data-theme="navy"] body {
  background: radial-gradient(circle at 15% 15%, #3b144c 0%, #1c132b 50%, #0e0716 100%) !important;
  background-attachment: fixed !important;
}

html[data-theme="navy"] .card {
  background: rgba(45, 30, 64, 0.85) !important;
  border: 1px solid rgba(249, 115, 22, 0.35) !important;
  box-shadow: 0 8px 32px rgba(236, 72, 153, 0.18) !important;
}

html[data-theme="navy"] .navlink:hover {
  background: rgba(249, 115, 22, 0.15) !important;
  color: #f97316 !important;
}

html[data-theme="navy"] .navlink.active {
  background: linear-gradient(135deg, #f97316 0%, #ec4899 50%, #8b5cf6 100%) !important;
  color: #ffffff !important;
  box-shadow: 0 4px 18px rgba(249, 115, 22, 0.45) !important;
}

html[data-theme="navy"] .btn-primary {
  background: linear-gradient(135deg, #f97316 0%, #ec4899 100%) !important;
  border: none !important;
  box-shadow: 0 4px 15px rgba(249, 115, 22, 0.4) !important;
}

html[data-theme="navy"] .badge-soft {
  background: rgba(236, 72, 153, 0.2) !important;
  border: 1px solid rgba(236, 72, 153, 0.4) !important;
  color: #f472b6 !important;
}

body {
  background: var(--bg);
  color: var(--text);
  font-family: 'Inter', system-ui, -apple-system, sans-serif;
  margin: 0;
  transition: background 0.15s ease, color 0.15s ease;
  -webkit-font-smoothing: antialiased;
}

a { color: inherit; text-decoration: none; }

h1, h2, h3, h4, h5, h6 {
  color: var(--heading) !important;
  font-weight: 800;
  letter-spacing: -0.3px;
}

.brand {
  font-weight: 900;
  font-size: 14px;
  color: var(--heading);
  border-bottom: 1px solid var(--border);
  padding-bottom: 16px;
  margin-bottom: 20px;
}

/* --- Sidebar --- */
.sidebar {
  width: 260px;
  position: fixed;
  top: 0;
  left: 0;
  bottom: 0;
  background: var(--sidebar-bg);
  border-right: 1px solid var(--border);
  padding: 20px 16px;
  display: flex;
  flex-direction: column;
  z-index: 1000;
}
.navlink {
  display: flex;
  align-items: center;
  gap: 12px;
  padding: 11px 16px;
  border-radius: 12px;
  color: var(--muted);
  text-decoration: none;
  font-weight: 600;
  font-size: 13.5px;
  transition: all 0.2s cubic-bezier(0.4, 0, 0.2, 1);
  margin-bottom: 3px;
}
.navlink:hover {
  background: rgba(255, 255, 255, 0.08);
  color: var(--heading) !important;
  transform: translateX(4px);
}
.navlink.active {
  background: linear-gradient(135deg, var(--primary) 0%, var(--accent) 100%) !important;
  color: #ffffff !important;
  box-shadow: 0 6px 20px rgba(99, 102, 241, 0.35) !important;
  font-weight: 800;
}
.navlink i {
  font-size: 16px;
  width: 22px;
  text-align: center;
  transition: transform 0.2s ease;
}
.navlink:hover i {
  transform: scale(1.15);
}

/* --- Content --- */
.content {
  margin-left: 260px;
  padding: 28px 32px;
}

/* --- Card --- */
.card {
  background: var(--card);
  border: 1px solid var(--border);
  border-radius: 16px;
  box-shadow: 0 4px 20px rgba(0, 0, 0, 0.08);
  margin-bottom: 20px;
  transition: border-color 0.15s ease, box-shadow 0.15s ease;
}
.label {
  font-size: 11px;
  color: var(--muted);
  font-weight: 700;
  text-transform: uppercase;
  letter-spacing: 0.5px;
}
hr {
  border-color: var(--border);
  opacity: 1;
}

/* --- Buttons --- */
.btn-primary {
  background: var(--primary);
  border: none;
  font-weight: 600;
  border-radius: 12px;
  padding: 10px 20px;
  color: #ffffff;
  transition: all 0.15s ease;
}
.btn-primary:hover {
  background: var(--primary);
  opacity: 0.92;
  transform: translateY(-1px);
}
.btn-outline-primary {
  border-color: var(--primary);
  color: var(--primary);
  font-weight: 600;
  border-radius: 12px;
  padding: 8px 16px;
}
.btn-outline-primary:hover {
  background: var(--primary);
  border-color: var(--primary);
  color: #fff;
}
.btn-outline-secondary {
  border-color: var(--border);
  color: var(--muted);
  font-weight: 600;
  border-radius: 12px;
}
.btn-outline-secondary:hover {
  background: var(--border);
  color: var(--heading);
}

/* --- Form --- */
.form-control, .form-select {
  background: var(--input-bg) !important;
  border: 1px solid var(--border) !important;
  color: var(--input-text) !important;
  border-radius: 12px;
  padding: 10px 14px;
  transition: border-color 0.15s ease, box-shadow 0.15s ease;
}
.form-select option {
  background: var(--card) !important;
  color: var(--input-text) !important;
}
.form-control::placeholder {
  color: var(--muted);
}
.form-control:focus, .form-select:focus {
  border-color: var(--primary) !important;
  box-shadow: 0 0 0 3px rgba(99, 102, 241, 0.2);
}

/* --- Table --- */
.table {
  color: var(--text) !important;
}
.table td, .table th {
  border-color: var(--border) !important;
  vertical-align: middle;
  padding: 14px 16px;
}
.table thead th {
  color: var(--heading) !important;
  background: rgba(0, 0, 0, 0.05) !important;
  font-size: 13px;
  font-weight: 700;
}
html[data-theme="dark"] .table thead th, html[data-theme="navy"] .table thead th {
  background: rgba(255, 255, 255, 0.03) !important;
}
.table tbody td {
  font-size: 14px;
}
.badge-soft {
  background: rgba(99, 102, 241, 0.1);
  border: 1px solid rgba(99, 102, 241, 0.3);
  color: var(--accent);
  padding: 6px 12px;
  border-radius: 999px;
  font-weight: 700;
  font-size: 12px;
  display: inline-block;
}
.badge-yellow {
  background: rgba(245, 158, 11, 0.1);
  border: 1px solid rgba(245, 158, 11, 0.3);
  color: #fbbf24;
}

.depo-table tbody tr:hover {
  background: rgba(0, 0, 0, 0.03) !important;
}
html[data-theme="dark"] .depo-table tbody tr:hover, html[data-theme="navy"] .depo-table tbody tr:hover {
  background: rgba(255, 255, 255, 0.02) !important;
}

/* --- Form checks & text visibility --- */
.form-check.check-tight {
  display: flex;
  align-items: center;
  gap: 10px;
  padding-left: 0;
  margin-bottom: 8px;
}
.form-check.check-tight .form-check-input {
  float: none;
  margin: 0;
  width: 18px;
  height: 18px;
  border-color: var(--border);
  background-color: var(--input-bg);
}
.form-check.check-tight .form-check-input:checked {
  background-color: var(--primary);
  border-color: var(--primary);
}
.form-check-label {
  color: var(--text) !important;
}
details summary {
  color: var(--heading) !important;
  font-weight: 700;
  cursor: pointer;
}

.footer-right {
  position: fixed;
  right: 180px;
  bottom: 22px;
  color: var(--muted);
  font-size: 11px;
  font-weight: 600;
}

/* Floating Theme Switcher Pill */
.theme-switcher-pill {
  position: fixed;
  right: 20px;
  bottom: 16px;
  background: var(--card);
  border: 1px solid var(--border);
  border-radius: 30px;
  padding: 4px;
  display: flex;
  gap: 4px;
  box-shadow: 0 8px 24px rgba(0,0,0,0.3);
  z-index: 9999;
}
.theme-btn {
  background: transparent;
  border: none;
  color: var(--muted);
  padding: 6px 12px;
  border-radius: 20px;
  font-size: 11px;
  font-weight: 700;
  cursor: pointer;
  transition: all 0.15s ease;
  display: flex;
  align-items: center;
  gap: 5px;
}
.theme-btn:hover {
  color: var(--heading);
}
.theme-btn.active {
  background: var(--primary);
  color: #ffffff !important;
}

/* --- Mobil Header ve Alt Menü --- */
.mobile-header {
  display: none;
  background: var(--sidebar-bg);
  border-bottom: 1px solid var(--border);
  height: 56px;
  position: fixed;
  top: 0;
  left: 0;
  right: 0;
  z-index: 999;
  padding: 0 16px;
  align-items: center;
  justify-content: space-between;
}
.mobile-brand {
  font-weight: 800;
  font-size: 15px;
  color: var(--heading);
}
.mobile-logout {
  color: #ef4444;
  font-size: 18px;
}
.mobile-logout:hover {
  color: #ef4444;
}

.bottom-nav {
  display: none;
  position: fixed;
  bottom: 0;
  left: 0;
  right: 0;
  height: 64px;
  background: var(--sidebar-bg);
  border-top: 1px solid var(--border);
  z-index: 999;
  justify-content: space-around;
  align-items: center;
}
.bottom-nav-link {
  display: flex;
  flex-direction: column;
  align-items: center;
  gap: 4px;
  color: var(--muted);
  font-size: 11px;
  font-weight: 600;
  text-decoration: none;
}
.bottom-nav-link i {
  font-size: 18px;
}
.bottom-nav-link.active {
  color: var(--accent);
}

@media (max-width: 991px) {
  .sidebar { display: none; }
  .mobile-header { display: flex; }
  .content {
    margin-left: 0 !important;
    padding: 16px !important;
    padding-top: 76px !important;
    padding-bottom: 88px !important;
  }
  .bottom-nav { display: flex; }
  .footer-right { display: none; }
}
/* Yanıp sönen kırmızı nokta animasyonu */
.pulse-red-dot {
  display: inline-block;
  width: 10px;
  height: 10px;
  border-radius: 50%;
  background-color: #ef4444;
  box-shadow: 0 0 0 0 rgba(239, 68, 68, 0.7);
  animation: pulse-red 1.5s infinite;
  vertical-align: middle;
}
@keyframes pulse-red {
  0% { transform: scale(0.95); box-shadow: 0 0 0 0 rgba(239, 68, 68, 0.7); }
  70% { transform: scale(1.1); box-shadow: 0 0 0 8px rgba(239, 68, 68, 0); }
  100% { transform: scale(0.95); box-shadow: 0 0 0 0 rgba(239, 68, 68, 0); }
}
</style>
</head>
<body>

<div class="mobile-header">
  <div class="mobile-brand d-flex align-items-center gap-2">
    <img src="/logo.png" style="width: 32px; height: 32px; border-radius: 50%; object-fit: cover; border: 1px solid var(--accent);" />
    <div style="font-size: 13px; font-weight: 800;">Diyarbakır BAM <span style="color: var(--accent);">Envanter Sistemi</span></div>
  </div>
  <div>
    <a href="#" onclick="window.location.reload(); return false;" class="text-warning me-3" title="Yenile"><i class="fa fa-rotate-right fs-5"></i></a>
    <a href="/logout" class="mobile-logout"><i class="fa fa-sign-out-alt"></i></a>
  </div>
</div>

<div class="sidebar">
  <div class="brand d-flex align-items-center gap-3 text-start pb-3 mb-3" style="border-bottom: 1px solid var(--border);">
    <img src="/logo.png" style="width: 44px; height: 44px; border-radius: 50%; object-fit: cover; border: 2px solid var(--primary); box-shadow: 0 4px 12px rgba(99, 102, 241, 0.3);" />
    <div>
      <div style="font-weight: 800; font-size: 13px; color: var(--heading); line-height: 1.25;">Diyarbakır BAM</div>
      <small style="color: var(--muted); font-size: 10px; font-weight: 700; text-transform: uppercase; letter-spacing: 0.5px;">Envanter Sistemi</small>
    </div>
  </div>
  <div class="d-flex flex-column gap-1">
    <a class="navlink" href="/dashboard" id="s_dash"><i class="fa-solid fa-chart-pie"></i> Yönetim</a>
    <a class="navlink" href="/envanter" id="s_env"><i class="fa-solid fa-boxes-stacked"></i> Envanter</a>
    <a class="navlink" href="/zimmet" id="s_zim"><i class="fa-solid fa-file-contract"></i> Zimmet / Teslim</a>
    <a class="navlink" href="/daire_incele" id="s_dai"><i class="fa-solid fa-building-user"></i> Daire İnceleme</a>
    <a class="navlink" href="/depo" id="s_dep"><i class="fa-solid fa-warehouse"></i> Depo</a>
    <a class="navlink" href="/gecmis" id="s_gec"><i class="fa-solid fa-clock-rotate-left"></i> Geçmiş</a>
    {% if is_admin %}<a class="navlink" href="/ayarlar" id="s_set"><i class="fa-solid fa-sliders"></i> Ayarlar</a>{% endif %}
    {% if (request.path == '/dashboard' or request.path == '/') and critical_count and critical_count > 0 %}
    <button type="button" onclick="onCriticalBellClick()" class="btn btn-sm w-100 mt-2 fw-bold text-start d-flex align-items-center justify-content-between py-2 px-3" style="border-radius: 12px; font-size: 12.5px; background: rgba(239, 68, 68, 0.12); border: 1px solid rgba(239, 68, 68, 0.35) !important; color: #fca5a5; backdrop-filter: blur(8px);" data-bs-toggle="modal" data-bs-target="#criticalStockModal">
      <span><span id="critical_bell_dot" class="pulse-red-dot me-2"></span>Kritik Stok Uyarısı</span>
      <span class="badge bg-danger rounded-pill px-2 py-1" style="font-size: 11px;">{{ critical_count }}</span>
    </button>
    {% endif %}
    <a class="navlink mt-2" href="/logout" style="color: #fca5a5;"><i class="fa-solid fa-right-from-bracket" style="color: #fca5a5;"></i> Çıkış</a>
  </div>
  <div style="margin-top: auto; padding-top: 16px; border-top: 1px solid var(--border); color: var(--muted); font-size: 11px; font-weight: 700; opacity: 0.8;">
    By Murat İRVEN (187665)
  </div>
</div>

<div class="content">
  {% if message %}<div class="alert alert-warning border-0 rounded-3" style="background: rgba(245, 158, 11, 0.15); border: 1px solid rgba(245, 158, 11, 0.3) !important; color: #fbbf24;">{{ message }}</div>{% endif %}
  {{ content|safe }}
</div>

<div class="bottom-nav">
  <a class="bottom-nav-link" href="/dashboard" id="b_dash"><i class="fa fa-chart-line"></i><span>Yönetim</span></a>
  <a class="bottom-nav-link" href="/envanter" id="b_env"><i class="fa fa-desktop"></i><span>Envanter</span></a>
  <a class="bottom-nav-link" href="/zimmet" id="b_zim"><i class="fa fa-file-signature"></i><span>Zimmet</span></a>
  <a class="bottom-nav-link" href="/depo" id="b_dep"><i class="fa fa-warehouse"></i><span>Depo</span></a>
  <a class="bottom-nav-link" href="/gecmis" id="b_gec"><i class="fa fa-clock-rotate-left"></i><span>Geçmiş</span></a>
</div>

<!-- Floating Theme Switcher Pill -->
<div class="theme-switcher-pill">
  <button type="button" onclick="setAppTheme('dark')" class="theme-btn" id="tb_dark" title="Koyu Tema"><i class="fa-solid fa-moon"></i> Koyu</button>
  <button type="button" onclick="setAppTheme('light')" class="theme-btn" id="tb_light" title="Aydınlık Tema"><i class="fa-solid fa-sun"></i> Beyaz</button>
  <button type="button" onclick="setAppTheme('navy')" class="theme-btn" id="tb_navy" title="2026 Sunset Glass Tema"><i class="fa-solid fa-wand-magic-sparkles text-warning me-1"></i> 2026 Sunset Glass</button>
</div>

<!-- Kritik Stok Pop-up Modal -->
{% if critical_count and critical_count > 0 %}
<div class="modal fade" id="criticalStockModal" tabindex="-1" aria-hidden="true">
  <div class="modal-dialog modal-lg modal-dialog-centered">
    <div class="modal-content" style="background: var(--card); border: 2px solid #ef4444; color: var(--text); border-radius: 16px;">
      <div class="modal-header" style="background: rgba(239, 68, 68, 0.15); border-bottom: 1px solid rgba(239, 68, 68, 0.3);">
        <h5 class="modal-title fw-bold text-danger d-flex align-items-center gap-2 mb-0">
          <span class="pulse-red-dot"></span>
          <i class="fa fa-triangle-exclamation text-danger fs-4"></i>
          KRİTİK STOK UYARISI
        </h5>
        <button type="button" class="btn-close btn-close-white" data-bs-dismiss="modal"></button>
      </div>
      <div class="modal-body p-4">
        <div class="alert alert-danger border-0 mb-3" style="background: rgba(239, 68, 68, 0.1); color: #fca5a5; font-size: 13px;">
          <i class="fa fa-circle-info me-2"></i> Aşağıda listelenen <strong>{{ critical_count }} kalem</strong> sarf malzemesinin stoğu tükenmek üzeredir (2 veya altında). Lütfen tedarik sürecini başlatınız.
        </div>
        <div class="table-responsive">
          <table class="table table-striped align-middle mb-0" style="font-size: 13px;">
            <thead>
              <tr class="table-danger">
                <th>Malzeme / Kalem Adı</th>
                <th class="text-center">Mevcut Stok</th>
                <th class="text-center">Durum</th>
                <th class="text-end">İşlem</th>
              </tr>
            </thead>
            <tbody>
              {% for st in critical_items_list %}
              <tr>
                <td><span class="pulse-red-dot me-2"></span><strong>{{ st.name }}</strong></td>
                <td class="text-center"><span class="badge bg-danger fs-6 px-3 py-1">{{ st.qty }} Adet Kalan</span></td>
                <td class="text-center">
                  {% if st.qty == 0 %}
                  <span class="badge bg-dark text-danger border border-danger">TÜKENDİ (0 Adet)</span>
                  {% else %}
                  <span class="badge bg-warning text-dark">TÜKENMEK ÜZERE</span>
                  {% endif %}
                </td>
                <td class="text-end">
                  <a href="/depo?q={{ st.name }}" class="btn btn-sm btn-outline-primary fw-bold" data-bs-dismiss="modal">Depoda Düzenle</a>
                </td>
              </tr>
              {% endfor %}
            </tbody>
          </table>
        </div>
      </div>
      <div class="modal-footer" style="border-top: 1px solid var(--border);">
        <a href="/depo" class="btn btn-primary fw-bold"><i class="fa fa-warehouse me-1"></i> Depo Yönetimine Git</a>
        <button type="button" class="btn btn-secondary fw-bold" data-bs-dismiss="modal">Kapat</button>
      </div>
    </div>
  </div>
</div>

<script>
function onCriticalBellClick() {
  sessionStorage.setItem("critical_bell_seen", "1");
  dimBellPulse();
}

function dimBellPulse() {
  var dot = document.getElementById("critical_bell_dot");
  if (dot) {
    dot.classList.remove("pulse-red-dot");
    dot.style.backgroundColor = "#94a3b8";
    dot.style.boxShadow = "none";
    dot.style.animation = "none";
  }
}

document.addEventListener("DOMContentLoaded", function() {
  // Tıklandıysa veya sönük ayarlandıysa nokta yanıp sönmeyi bırakır
  if (sessionStorage.getItem("critical_bell_seen")) {
    dimBellPulse();
  }
});
</script>
{% endif %}

<script src="https://cdn.jsdelivr.net/npm/bootstrap@5.3.2/dist/js/bootstrap.bundle.min.js"></script>

<script>
function setAppTheme(t) {
  document.documentElement.setAttribute('data-theme', t);
  localStorage.setItem('zimmet_theme', t);
  document.querySelectorAll('.theme-btn').forEach(b => b.classList.remove('active'));
  const activeBtn = document.getElementById('tb_' + t);
  if(activeBtn) activeBtn.classList.add('active');
}

(function(){
  const saved = localStorage.getItem('zimmet_theme') || 'dark';
  setAppTheme(saved);

  const path = window.location.pathname || "/";
  const root = document.documentElement;
  
  const map = [
    {re:/^\/envanter/, accent:"#10b981", s_id:"s_env", b_id:"b_env"},
    {re:/^\/zimmet/, accent:"#06b6d4", s_id:"s_zim", b_id:"b_zim"},
    {re:/^\/depo/, accent:"#f59e0b", s_id:"s_dep", b_id:"b_dep"},
    {re:/^\/gecmis/, accent:"#ec4899", s_id:"s_gec", b_id:"b_gec"},
    {re:/^\/ayarlar/, accent:"#8b5cf6", s_id:"s_set", b_id:""},
    {re:/^\/dashboard/, accent:"#6366f1", s_id:"s_dash", b_id:"b_dash"},
    {re:/^\//, accent:"#6366f1", s_id:"s_dash", b_id:"b_dash"},
  ];
  
  const hit = map.find(x=>x.re.test(path)) || map[map.length-1];
  root.style.setProperty("--accent", hit.accent);
  root.style.setProperty("--primary", hit.accent);
  
  if(hit.s_id) {
    const s_el = document.getElementById(hit.s_id);
    if(s_el) s_el.classList.add("active");
  }
  if(hit.b_id) {
    const b_el = document.getElementById(hit.b_id);
    if(b_el) b_el.classList.add("active");
  }
})();
</script>

</body>
</html>
"""

LOGIN_HTML = r"""<!doctype html>
<html lang="tr">
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>Diyarbakır Bölge Adliye Mahkemesi Envanter Sistemi - Giriş</title>
<link href="https://cdn.jsdelivr.net/npm/bootstrap@5.3.2/dist/css/bootstrap.min.css" rel="stylesheet">
<link href="https://cdnjs.cloudflare.com/ajax/libs/font-awesome/6.5.1/css/all.min.css" rel="stylesheet">
<link href="https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600;700;800;900&display=swap" rel="stylesheet">
<style>
body{
  background: radial-gradient(circle at 50% 50%, #151b2c 0%, #0b0f19 100%);
  color: #e2e8f0;
  height: 100vh;
  display: flex;
  align-items: center;
  justify-content: center;
  font-family: 'Inter', sans-serif;
  margin: 0;
}
.card{
  background: #111726;
  border: 1px solid #232d45;
  border-radius: 20px;
  width: 470px;
  box-shadow: 0 15px 35px rgba(0,0,0,0.4);
  padding: 32px !important;
}
.form-control, .form-select{
  background: #0d121f;
  border: 1px solid #232d45;
  color: #e2e8f0;
  border-radius: 12px;
  padding: 12px;
}
.form-control:focus, .form-select:focus{
  border-color: #6366f1;
  box-shadow: 0 0 0 2px rgba(99, 102, 241, 0.2);
  background: #0d121f;
  color: #fff;
}
h4{
  font-weight: 900;
  color: #fff;
  letter-spacing: 0.5px;
}
.muted{
  color: #64748b;
  font-weight: 700;
  font-size: 11px;
  text-transform: uppercase;
  letter-spacing: 0.5px;
}
.btn-login {
  background: #6366f1;
  border: none;
  color: #fff;
  font-weight: 600;
  padding: 12px;
  border-radius: 12px;
  transition: all 0.2s;
}
.btn-login:hover {
  background: #4f46e5;
  transform: translateY(-1px);
}
</style>
</head>
<body>
<div class="card" style="border-top: 6px solid #6366f1">
  <div class="text-center mb-4">
    <img src="/logo.png" style="width: 150px; height: 150px; border-radius: 50%; object-fit: cover; border: 4px solid rgba(99, 102, 241, 0.5); margin-bottom: 16px; box-shadow: 0 10px 30px rgba(0,0,0,0.5);" />
    <h4 style="font-weight: 900; letter-spacing: -0.5px; margin-bottom: 2px; font-size: 18px;">Diyarbakır Bölge Adliye Mahkemesi</h4>
    <div class="muted" style="color: #818cf8; font-size: 11px; font-weight: 800; letter-spacing: 0.5px;">ENVANTER VE ZİMMET SİSTEMİ</div>
  </div>
  {% if err %}<div class="alert alert-danger py-2 border-0 mb-3" style="background: rgba(239, 68, 68, 0.15); border: 1px solid rgba(239, 68, 68, 0.3) !important; color: #fca5a5; font-size: 13px;">{{ err }}</div>{% endif %}
  {% if message %}<div class="alert alert-success py-2 border-0 mb-3" style="background: rgba(16, 185, 129, 0.15); border: 1px solid rgba(16, 185, 129, 0.3) !important; color: #6ee7b7; font-size: 13px;">{{ message }}</div>{% endif %}
  <form method="post">
    <div class="mb-3">
      <div class="muted mb-1">Kullanıcı</div>
      <select class="form-select" name="sicil" required>
        <option value="" selected disabled>Seçiniz...</option>
        {% for u in users %}
          <option value="{{ u.sicil }}" {% if remembered==u.sicil %}selected{% endif %}>
            {{ u.name }} ({{ u.sicil }})
          </option>
        {% endfor %}
      </select>
    </div>
    <div class="mb-3">
      <div class="muted mb-1">Şifre</div>
      <input type="password" class="form-control" name="password" required>
    </div>
    <div class="d-flex justify-content-between align-items-center mb-4">
      <div class="form-check mb-0">
        <input class="form-check-input" type="checkbox" name="remember" id="remember" checked style="background-color: #0d121f; border-color: #232d45;">
        <label class="form-check-label muted" for="remember" style="padding-top: 2px;">Beni hatırla</label>
      </div>
      <a href="#" data-bs-toggle="modal" data-bs-target="#changePwModal" style="color: #818cf8; text-decoration: none; font-size: 12px; font-weight: 600;"><i class="fa fa-key me-1"></i> Şifremi Değiştir</a>
    </div>
    <button class="btn btn-login w-100">Giriş Yap</button>
  </form>
</div>

<!-- Şifre Değiştir Modal -->
<div class="modal fade" id="changePwModal" tabindex="-1" aria-hidden="true">
  <div class="modal-dialog modal-dialog-centered">
    <div class="modal-content" style="background: #111726; border: 1px solid #232d45; color: #e2e8f0; border-radius: 16px;">
      <div class="modal-header" style="border-bottom: 1px solid #232d45;">
        <h5 class="modal-title fw-bold text-white"><i class="fa fa-key me-2 text-primary"></i>Şifre Değiştir</h5>
        <button type="button" class="btn-close btn-close-white" data-bs-dismiss="modal"></button>
      </div>
      <form action="/change_password" method="post">
        <div class="modal-body">
          <div class="mb-3">
            <div class="muted mb-1">Kullanıcı</div>
            <select class="form-select" name="sicil" required>
              <option value="" selected disabled>Kullanıcı Seçiniz...</option>
              {% for u in users %}
                <option value="{{ u.sicil }}">{{ u.name }} ({{ u.sicil }})</option>
              {% endfor %}
            </select>
          </div>
          <div class="mb-3">
            <div class="muted mb-1">Mevcut Şifre</div>
            <input type="password" class="form-control" name="old_password" required placeholder="Mevcut şifreniz">
          </div>
          <div class="mb-3">
            <div class="muted mb-1">Yeni Şifre</div>
            <input type="password" class="form-control" name="new_password" required placeholder="Yeni şifreniz">
          </div>
          <div class="mb-3">
            <div class="muted mb-1">Yeni Şifre (Tekrar)</div>
            <input type="password" class="form-control" name="new_password2" required placeholder="Yeni şifreniz tekrar">
          </div>
        </div>
        <div class="modal-footer" style="border-top: 1px solid #232d45;">
          <button type="button" class="btn btn-secondary" data-bs-dismiss="modal">İptal</button>
          <button type="submit" class="btn btn-primary fw-bold">Şifreyi Güncelle</button>
        </div>
      </form>
    </div>
  </div>
</div>

<script src="https://cdn.jsdelivr.net/npm/bootstrap@5.3.2/dist/js/bootstrap.bundle.min.js"></script>
</body>
</html>
"""


# =========================================================
# DB INIT (Flask 3.x Uyumlu)
# - Flask 3 ile @app.before_first_request kaldırıldı.
# - Bu yüzden DB init işlemini process başına 1 kez çalıştırıyoruz.
# =========================================================
_DB_INITIALIZED = False

def init_db_once() -> None:
    global _DB_INITIALIZED
    if _DB_INITIALIZED:
        return
    # App context altında tabloları oluştur ve seed'i uygula
    with app.app_context():
        db.create_all()
        seed_data()
    _DB_INITIALIZED = True

@app.before_request
def _ensure_db_initialized():
    # Her request'te çağrılır ama sadece ilkinde init yapar (process başına).
    init_db_once()

@app.route("/logo.png")
def serve_logo():
    if os.path.exists(LOGO_PATH):
        return send_file(LOGO_PATH, mimetype="image/png")
    return "", 404


# =========================================================
# AUTH
# =========================================================
@app.route("/", methods=["GET", "POST"])
def login():
    remembered = request.cookies.get("remember_sicil", "")
    err = None
    if request.method == "POST":
        try:
            sicil = (request.form.get("sicil") or "").strip()
            pw = request.form.get("password") or ""
            if not sicil or not pw:
                err = "Lütfen kullanıcı ve şifre giriniz."
            else:
                user = User.query.filter_by(sicil=sicil).first()
                if user and check_password_hash(user.password_hash, pw):
                    session["user"] = user.name
                    session["sicil"] = user.sicil
                    resp = make_response(redirect(url_for("dashboard")))
                    if request.form.get("remember") == "on":
                        resp.set_cookie("remember_sicil", user.sicil, max_age=60*60*24*30)
                    else:
                        resp.set_cookie("remember_sicil", "", expires=0)
                    return resp
                err = "Hatalı sicil numarası veya şifre!"
        except Exception as e:
            err = f"Giriş hatası: {e}"

    try:
        users = User.query.order_by(User.name.asc()).all()
    except Exception:
        users = []

    return render_template_string(LOGIN_HTML, err=err, message=None, users=users, remembered=remembered)


@app.route("/change_password", methods=["POST"])
def change_password():
    sicil = (request.form.get("sicil") or "").strip()
    old_pw = request.form.get("old_password") or ""
    new_pw = request.form.get("new_password") or ""
    new_pw2 = request.form.get("new_password2") or ""

    try:
        users = User.query.order_by(User.name.asc()).all()
    except Exception:
        users = []

    if not sicil or not old_pw or not new_pw:
        return render_template_string(LOGIN_HTML, err="Lütfen tüm şifre alanlarını doldurunuz.", message=None, users=users, remembered="")
    if new_pw != new_pw2:
        return render_template_string(LOGIN_HTML, err="Yeni şifreler birbiriyle eşleşmiyor!", message=None, users=users, remembered="")

    u = User.query.filter_by(sicil=sicil).first()
    if not u or not check_password_hash(u.password_hash, old_pw):
        return render_template_string(LOGIN_HTML, err="Mevcut şifre veya sicil numarası hatalı!", message=None, users=users, remembered="")

    try:
        u.password_hash = generate_password_hash(new_pw)
        db.session.commit()
        return render_template_string(LOGIN_HTML, err=None, message="Şifreniz başarıyla değiştirildi! Yeni şifrenizle giriş yapabilirsiniz.", users=users, remembered=sicil)
    except Exception as e:
        db.session.rollback()
        return render_template_string(LOGIN_HTML, err=f"Şifre değiştirme hatası: {e}", message=None, users=users, remembered="")

@app.route("/logout")
def logout():
    session.clear()
    return redirect(url_for("login"))

# =========================================================
# DASHBOARD
# =========================================================
def render_base(content: str, message: str = None):
    try:
        critical_items = Stock.query.filter(Stock.qty <= 2).order_by(Stock.qty.asc()).all()
    except Exception:
        critical_items = []
    
    return render_template_string(
        BASE_HTML,
        content=content,
        message=message,
        is_admin=is_admin(),
        critical_items_list=critical_items,
        critical_count=len(critical_items)
    )


@app.route("/dashboard")
def dashboard():
    if not require_login():
        return redirect(url_for("login"))

    toplam = InventoryItem.query.count()
    depoda = InventoryItem.query.filter_by(status="Depoda").count()
    zimmetli = InventoryItem.query.filter_by(status="Zimmetli").count()
    arizali = InventoryItem.query.filter_by(status="Arızalı / Bakımda").count()
    hurda = InventoryItem.query.filter_by(status="Hek / Hurda").count()
    stok = Stock.query.count()
    local_ip = get_local_ip()

    # Canlı Daire Dağılım İstatistikleri ve Alt Birim Listeleri
    unit_stats = {}
    unit_details_html = {}
    all_invs = InventoryItem.query.all()

    for group_name, units in BAM_UNITS_STRUCTURE.items():
        unit_stats[group_name] = 0
        grid_items_html = ""
        for u_name in units:
            cnt = sum(1 for it in all_invs if (it.assigned_unit and u_name.lower() in it.assigned_unit.lower()))
            if cnt > 0:
                unit_stats[group_name] += cnt

            badge_cls = "bg-primary" if cnt > 0 else "bg-secondary"
            grid_items_html += f"""
            <div class="col-md-3 col-6">
              <div class="card p-2 text-center" style="background:#1e293b; border:1px solid #334155;">
                <div style="font-size:12px; font-weight:700; color:#f8fafc;" class="text-truncate">{u_name}</div>
                <div class="d-flex justify-content-between align-items-center mt-2 px-1">
                  <span class="badge {badge_cls}" style="font-size:11px;">{cnt} Cihaz</span>
                  <a href="/admin/export/daire_defteri.pdf?daire_adi={u_name}" class="btn btn-sm btn-outline-warning py-0 px-2" style="font-size:11px;" title="PDF Defter İndir">
                    <i class="fa fa-file-pdf"></i> Defter
                  </a>
                </div>
              </div>
            </div>
            """
        unit_details_html[group_name] = grid_items_html

    idari_units_total = unit_stats.get('İdari Bürolar ve Müdürlükler', 0) + unit_stats.get('Başsavcılık ve Komisyon', 0)
    idari_grid = unit_details_html.get('İdari Bürolar ve Müdürlükler', '') + unit_details_html.get('Başsavcılık ve Komisyon', '')

    daire_cards_html = f"""
    <div class="card p-4 mt-3">
      <div class="d-flex justify-content-between align-items-center mb-2">
        <h5 class="mb-0"><i class="fa fa-sitemap text-warning me-2"></i>Diyarbakır BAM Daire & Birim Canlı Envanter Haritası</h5>
        <a href="/gecmis" class="btn btn-sm btn-outline-primary fw-bold"><i class="fa fa-book me-1"></i> Tüm Daire Defterleri</a>
      </div>
      <div class="label mb-3">Detaylarını görmek ve envanter defterini çıkarmak istediğiniz daire grubuna tıklayınız.</div>

      <div class="row g-3 mb-3">
        <!-- 17 Ceza Dairesi Kartı -->
        <div class="col-md-3 col-6">
          <div class="card p-3 cursor-pointer" onclick="showGroupDetails('ceza_group')" style="border-left:4px solid #ef4444; background:#0f172a; cursor:pointer; transition:transform 0.2s;" onmouseover="this.style.transform='scale(1.02)'" onmouseout="this.style.transform='scale(1)'">
            <div class="d-flex justify-content-between align-items-center">
              <div class="label mb-1 text-danger fw-bold">17 CEZA DAİRESİ</div>
              <i class="fa fa-chevron-down text-white"></i>
            </div>
            <div class="fs-4 fw-bold text-white">{unit_stats.get('Ceza Daireleri', 0)} <span style="font-size:13px; font-weight:700; color:#e2e8f0;" class="ms-1">Aktif Cihaz</span></div>
          </div>
        </div>

        <!-- 12 Hukuk Dairesi Kartı -->
        <div class="col-md-3 col-6">
          <div class="card p-3 cursor-pointer" onclick="showGroupDetails('hukuk_group')" style="border-left:4px solid #3b82f6; background:#0f172a; cursor:pointer; transition:transform 0.2s;" onmouseover="this.style.transform='scale(1.02)'" onmouseout="this.style.transform='scale(1)'">
            <div class="d-flex justify-content-between align-items-center">
              <div class="label mb-1 text-primary fw-bold">12 HUKUK DAİRESİ</div>
              <i class="fa fa-chevron-down text-white"></i>
            </div>
            <div class="fs-4 fw-bold text-white">{unit_stats.get('Hukuk Daireleri', 0)} <span style="font-size:13px; font-weight:700; color:#e2e8f0;" class="ms-1">Aktif Cihaz</span></div>
          </div>
        </div>

        <!-- 8 Duruşma Salonu Kartı -->
        <div class="col-md-3 col-6">
          <div class="card p-3 cursor-pointer" onclick="showGroupDetails('durusma_group')" style="border-left:4px solid #10b981; background:#0f172a; cursor:pointer; transition:transform 0.2s;" onmouseover="this.style.transform='scale(1.02)'" onmouseout="this.style.transform='scale(1)'">
            <div class="d-flex justify-content-between align-items-center">
              <div class="label mb-1 text-success fw-bold">8 DURUŞMA SALONU</div>
              <i class="fa fa-chevron-down text-white"></i>
            </div>
            <div class="fs-4 fw-bold text-white">{unit_stats.get('Duruşma Salonları (Ortak Kullanım)', 0)} <span style="font-size:13px; font-weight:700; color:#e2e8f0;" class="ms-1">Aktif Cihaz</span></div>
          </div>
        </div>

        <!-- İdari Bürolar Kartı -->
        <div class="col-md-3 col-6">
          <div class="card p-3 cursor-pointer" onclick="showGroupDetails('idari_group')" style="border-left:4px solid #f59e0b; background:#0f172a; cursor:pointer; transition:transform 0.2s;" onmouseover="this.style.transform='scale(1.02)'" onmouseout="this.style.transform='scale(1)'">
            <div class="d-flex justify-content-between align-items-center">
              <div class="label mb-1 text-warning fw-bold">İDARİ BÜROLAR & MAKAM</div>
              <i class="fa fa-chevron-down text-white"></i>
            </div>
            <div class="fs-4 fw-bold text-white">{idari_units_total} <span style="font-size:13px; font-weight:700; color:#e2e8f0;" class="ms-1">Aktif Cihaz</span></div>
          </div>
        </div>
      </div>

      <!-- Tıklanınca Açılan Canlı Daire Detay Panelleri -->
      <div id="ceza_group" class="group-detail-panel p-3 rounded mb-3" style="display:none; background:#0b1120; border:1px solid #ef4444;">
        <h6 class="text-danger fw-bold mb-3"><i class="fa fa-gavel me-2"></i>17 Ceza Dairesi Canlı Envanter Haritası</h6>
        <div class="row g-2">{unit_details_html.get('Ceza Daireleri', '')}</div>
      </div>

      <div id="hukuk_group" class="group-detail-panel p-3 rounded mb-3" style="display:none; background:#0b1120; border:1px solid #3b82f6;">
        <h6 class="text-primary fw-bold mb-3"><i class="fa fa-scale-balanced me-2"></i>12 Hukuk Dairesi Canlı Envanter Haritası</h6>
        <div class="row g-2">{unit_details_html.get('Hukuk Daireleri', '')}</div>
      </div>

      <div id="durusma_group" class="group-detail-panel p-3 rounded mb-3" style="display:none; background:#0b1120; border:1px solid #10b981;">
        <h6 class="text-success fw-bold mb-3"><i class="fa fa-chair me-2"></i>8 Ortak Duruşma Salonu Canlı Envanter Haritası</h6>
        <div class="row g-2">{unit_details_html.get('Duruşma Salonları (Ortak Kullanım)', '')}</div>
      </div>

      <div id="idari_group" class="group-detail-panel p-3 rounded mb-3" style="display:none; background:#0b1120; border:1px solid #f59e0b;">
        <h6 class="text-warning fw-bold mb-3"><i class="fa fa-building me-2"></i>Başsavcılık, Komisyon ve İdari Bürolar Canlı Envanter Haritası</h6>
        <div class="row g-2">{idari_grid}</div>
      </div>

      <script>
      function showGroupDetails(panelId) {{
        var panels = document.querySelectorAll('.group-detail-panel');
        panels.forEach(function(p) {{
          if (p.id === panelId) {{
            p.style.display = (p.style.display === 'none' || p.style.display === '') ? 'block' : 'none';
          }} else {{
            p.style.display = 'none';
          }}
        }});
      }}
      </script>
    </div>
    """

    content = f"""
    <div class="card p-4 mb-3" style="background: linear-gradient(135deg, rgba(99, 102, 241, 0.1) 0%, rgba(15, 23, 42, 0.4) 100%); border: 1px solid rgba(99, 102, 241, 0.25);">
      <div class="d-flex align-items-center gap-3 mb-2 flex-wrap">
        <img src="/logo.png" style="width: 54px; height: 54px; border-radius: 50%; object-fit: cover; border: 2px solid var(--primary); box-shadow: 0 4px 16px rgba(99, 102, 241, 0.35);" />
        <div>
          <h4 class="mb-0 fw-bold" style="letter-spacing: -0.5px;">Hoş geldiniz, {session.get('user')}</h4>
          <div class="label mt-1" style="color: #a5b4fc; text-transform: uppercase; font-size: 11px; letter-spacing: 0.5px;">Diyarbakır Bölge Adliye Mahkemesi Bilgi İşlem Müdürlüğü Envanter Paneli</div>
        </div>
        <button type="button" onclick="window.location.reload()" class="btn btn-sm btn-outline-warning ms-auto fw-bold"><i class="fa fa-rotate-right me-1"></i> Sayfayı Yenile</button>
      </div>
    </div>

    <div class="row g-3">
      <div class="col-md-2 col-6"><div class="card p-3 text-center"><div class="label mb-1">Toplam Cihaz</div><h5 class="mb-0 fw-bold">{toplam}</h5></div></div>
      <div class="col-md-2 col-6"><div class="card p-3 text-center"><div class="label mb-1">Depoda</div><h5 class="mb-0 fw-bold text-success">{depoda}</h5></div></div>
      <div class="col-md-2 col-6"><div class="card p-3 text-center"><div class="label mb-1">Zimmetli</div><h5 class="mb-0 fw-bold text-primary">{zimmetli}</h5></div></div>
      <div class="col-md-2 col-6"><div class="card p-3 text-center"><div class="label mb-1">Arızalı / Bakımda</div><h5 class="mb-0 fw-bold text-warning">{arizali}</h5></div></div>
      <div class="col-md-2 col-6"><div class="card p-3 text-center"><div class="label mb-1">Hek / Hurda</div><h5 class="mb-0 fw-bold text-danger">{hurda}</h5></div></div>
      <div class="col-md-2 col-6"><div class="card p-3 text-center"><div class="label mb-1">Depo Kalemi</div><h5 class="mb-0 fw-bold text-info">{stok}</h5></div></div>
    </div>

    <!-- Chart.js Analiz Kartları -->
    <div class="row g-3 my-1">
      <div class="col-lg-7">
        <div class="card p-4 h-100" style="background: rgba(30, 41, 59, 0.35); border: 1px solid var(--border);">
          <div class="d-flex align-items-center justify-content-between mb-3">
            <h6 class="mb-0 fw-bold" style="color: var(--heading);"><i class="fa fa-chart-bar me-2 text-primary"></i> Daire / Birim Bazlı Cihaz Dağılım Grafiği</h6>
            <span class="badge bg-primary-subtle text-primary border border-primary-subtle rounded-pill">Top Birimler</span>
          </div>
          <div style="position: relative; height: 260px;">
            <canvas id="chartUnits"></canvas>
          </div>
        </div>
      </div>
      <div class="col-lg-5">
        <div class="card p-4 h-100" style="background: rgba(30, 41, 59, 0.35); border: 1px solid var(--border);">
          <div class="d-flex align-items-center justify-content-between mb-3">
            <h6 class="mb-0 fw-bold" style="color: var(--heading);"><i class="fa fa-chart-line me-2 text-warning"></i> Sarf Malzeme & Zimmet Tüketim Analizi</h6>
            <span class="badge bg-warning-subtle text-warning border border-warning-subtle rounded-pill">Aylık Trend</span>
          </div>
          <div style="position: relative; height: 260px;">
            <canvas id="chartMonthly"></canvas>
          </div>
        </div>
      </div>
    </div>

    <script>
    document.addEventListener("DOMContentLoaded", function(){{
      fetch("/api/analytics")
        .then(r => r.json())
        .then(data => {{
          if(!data.ok) return;
          
          // Chart 1: Units
          const ctx1 = document.getElementById("chartUnits");
          if(ctx1) {{
            new Chart(ctx1.getContext("2d"), {{
              type: "bar",
              data: {{
                labels: data.units.labels,
                datasets: [
                  {{ label: "Kasa", data: data.units.kasa, backgroundColor: "#6366f1", borderRadius: 4 }},
                  {{ label: "Monitör", data: data.units.monitor, backgroundColor: "#3b82f6", borderRadius: 4 }},
                  {{ label: "Yazıcı", data: data.units.yazici, backgroundColor: "#f59e0b", borderRadius: 4 }},
                  {{ label: "Tarayıcı", data: data.units.tarayici, backgroundColor: "#10b981", borderRadius: 4 }}
                ]
              }},
              options: {{
                responsive: true,
                maintainAspectRatio: false,
                scales: {{
                  x: {{ stacked: true, grid: {{ display: false }}, ticks: {{ color: "#94a3b8", font: {{ size: 10 }} }} }},
                  y: {{ stacked: true, grid: {{ color: "rgba(255,255,255,0.05)" }}, ticks: {{ color: "#94a3b8" }} }}
                }},
                plugins: {{
                  legend: {{ labels: {{ color: "#e2e8f0", font: {{ size: 11 }} }} }}
                }}
              }}
            }});
          }}

          // Chart 2: Monthly
          const ctx2 = document.getElementById("chartMonthly");
          if(ctx2) {{
            new Chart(ctx2.getContext("2d"), {{
              type: "line",
              data: {{
                labels: data.monthly.labels,
                datasets: [{{
                  label: "Aylık İşlem Hacmi",
                  data: data.monthly.data,
                  borderColor: "#f97316",
                  backgroundColor: "rgba(249, 115, 22, 0.15)",
                  fill: true,
                  tension: 0.4,
                  pointRadius: 4,
                  pointBackgroundColor: "#ec4899"
                }}]
              }},
              options: {{
                responsive: true,
                maintainAspectRatio: false,
                scales: {{
                  x: {{ grid: {{ display: false }}, ticks: {{ color: "#94a3b8" }} }},
                  y: {{ grid: {{ color: "rgba(255,255,255,0.05)" }}, ticks: {{ color: "#94a3b8" }} }}
                }},
                plugins: {{
                  legend: {{ labels: {{ color: "#e2e8f0", font: {{ size: 11 }} }} }}
                }}
              }}
            }});
          }}
        }});
    }});
    </script>

    {daire_cards_html}
    """
    return render_base(content)


# =========================================================
# ENVANTER
# =========================================================
@app.route("/envanter", methods=["GET", "POST"])
def envanter():
    if not require_login():
        return redirect(url_for("login"))

    message = None
    if request.method == "POST":
        action = request.form.get("action")
        if action == "add":
            try:
                category = normalize_spaces(request.form.get("category") or "")
                brand = normalize_spaces(request.form.get("brand") or "")
                model = normalize_spaces(request.form.get("model") or "")  # opsiyonel
                serial = normalize_spaces(request.form.get("serial_no") or "")

                if category not in DEVICE_CATEGORIES:
                    raise ValueError("Kategori hatalı.")
                if not (brand and serial):
                    raise ValueError("Marka ve seri no zorunlu. Model opsiyonel.")
                if InventoryItem.query.filter_by(serial_no=serial).first():
                    raise ValueError("Bu seri numarası zaten kayıtlı.")

                db.session.add(InventoryItem(
                    category=category, brand=brand, model=(model or None), serial_no=serial,
                    status="Depoda", last_event="Envantere Eklendi", last_event_at=now_str()
                ))
                db.session.commit()
                log_audit("ENVANTER_EKLE", f"{brand} {model or ''} ({serial}) - {category}")
                message = "Cihaz envantere eklendi."
            except Exception as e:
                db.session.rollback()
                message = f"Hata: {e}"

        elif action == "delete":
            try:
                iid = int(request.form.get("id"))
                it = InventoryItem.query.get(iid)
                if not it:
                    raise ValueError("Kayıt bulunamadı.")
                ser_num = it.serial_no
                db.session.delete(it)
                db.session.commit()
                log_audit("ENVANTER_SIL", f"{ser_num} seri nolu cihaz silindi")
                message = f"'{ser_num}' seri numaralı cihaz envanterden silindi."
            except Exception as e:
                db.session.rollback()
                message = f"Hata: {e}"

        elif action == "change_status":
            try:
                iid = int(request.form.get("id"))
                new_st = (request.form.get("new_status") or "").strip()
                it = InventoryItem.query.get(iid)
                if not it:
                    raise ValueError("Kayıt bulunamadı.")
                if new_st not in DEVICE_STATUSES:
                    raise ValueError("Geçersiz durum seçimi.")
                old_st = it.status
                it.status = new_st
                it.last_event = f"Durum: {new_st}"
                it.last_event_at = now_str()
                if new_st in ["Depoda", "Arızalı / Bakımda", "Hek / Hurda"]:
                    it.assigned_name = None
                    it.assigned_sicil = None
                    it.assigned_title = None
                    it.assigned_unit = None
                    it.assigned_at = None
                db.session.commit()
                log_audit("DURUM_DEGISTI", f"Cihaz #{it.id} ({it.serial_no}) yeni durum: {new_st}")
                message = f"Cihaz durumu '{old_st}' -> '{new_st}' olarak güncellendi."
            except Exception as e:
                db.session.rollback()
                message = f"Hata: {e}"

    items = InventoryItem.query.order_by(InventoryItem.id.desc()).all()

    rows = ""
    for it in items:
        if it.status == "Depoda":
            status_badge = '<span class="badge bg-success">Depoda</span>'
        elif it.status == "Zimmetli":
            status_badge = '<span class="badge bg-primary">Zimmetli</span>'
        elif it.status == "Arızalı / Bakımda":
            status_badge = '<span class="badge bg-warning text-dark">Arızalı / Bakımda</span>'
        else:
            status_badge = '<span class="badge bg-danger">Hek / Hurda</span>'

        who = "-"
        if it.status == "Zimmetli" and it.assigned_name:
            who = f"{format_tr_name(it.assigned_name)} / {it.assigned_sicil or ''}"

        status_opts = ""
        for st_val in DEVICE_STATUSES:
            sel = "selected" if it.status == st_val else ""
            status_opts += f'<option value="{st_val}" {sel}>{st_val}</option>'

        rows += f"""
        <tr>
          <td>{it.category}</td>
          <td>{it.brand}</td>
          <td>{it.model or ''}</td>
          <td><strong>{it.serial_no}</strong></td>
          <td>{status_badge}</td>
          <td>{who}</td>
          <td>
            <div class="d-flex gap-1 align-items-center">
              <form method="post" style="display:inline">
                <input type="hidden" name="action" value="change_status">
                <input type="hidden" name="id" value="{it.id}">
                <select name="new_status" onchange="this.form.submit()" class="form-select form-select-sm" style="width: 140px; font-size: 11px;">
                  {status_opts}
                </select>
              </form>
              <form method="post" style="display:inline" onsubmit="return confirm('{it.brand} {it.serial_no} seri nolu cihazı silmek istediğinize emin misiniz?');">
                <input type="hidden" name="action" value="delete">
                <input type="hidden" name="id" value="{it.id}">
                <button class="btn btn-sm btn-outline-danger" title="Sil"><i class="fa fa-trash"></i></button>
              </form>
            </div>
          </td>
        </tr>
        """

    msg_param = request.args.get("msg")
    if msg_param:
        message = msg_param

    content = f"""
    <div class="card p-4 mb-3">
      <h3>Envanter</h3>
      <div class="label">Model opsiyonel. Seri No benzersizdir.</div>
    </div>

    <!-- Excel İle Toplu Cihaz Yükleme Kartı -->
    <div class="card p-4 mb-3">
      <div class="d-flex justify-content-between align-items-center flex-wrap gap-2 mb-2">
        <h6><i class="fa fa-file-excel text-success me-2"></i>Toplu Excel İle Cihaz Yükle (.xlsx)</h6>
        <div>
          <a href="/admin/download_envanter_sablon" class="btn btn-sm btn-outline-success fw-bold me-2"><i class="fa fa-file-arrow-down me-1"></i> Örnek Şablon Excel İndir</a>
          <a href="/admin/export/hek_tutanak.pdf" class="btn btn-sm btn-outline-danger fw-bold"><i class="fa fa-file-pdf me-1"></i> Resmi Hek / Hurda Tutanağı (PDF) Üret</a>
        </div>
      </div>
      <div class="label mb-3">Sistem sütun başlıklarını (Kategori, Marka, Model, Seri No) otomatik algılar. Şablon Excel dosyasını indirip verilerinizi yapıştırabilirsiniz.</div>
      <form method="post" action="/admin/import_envanter_excel" enctype="multipart/form-data" class="row g-2 align-items-center">
        <div class="col-md-9">
          <input type="file" name="excel_file" class="form-control" accept=".xlsx, .xls" required>
        </div>
        <div class="col-md-3">
          <button class="btn btn-success w-100 fw-bold"><i class="fa fa-upload me-1"></i> Excel'den Aktar</button>
        </div>
      </form>
    </div>

    <div class="card p-4 mb-3">
      <h6>Cihaz Ekle</h6>
      <form method="post" class="row g-3">
        <input type="hidden" name="action" value="add">
        <div class="col-md-3">
          <div class="label">Kategori</div>
          <select name="category" class="form-select" required>
            {''.join([f'<option>{c}</option>' for c in DEVICE_CATEGORIES])}
          </select>
        </div>
        <div class="col-md-3">
          <div class="label">Marka</div>
          <input class="form-control" name="brand" list="brand_suggestions" placeholder="Marka seçin veya yazın..." required>
          <datalist id="brand_suggestions">
            <option value="Lenovo">
            <option value="Dell">
            <option value="Acer">
            <option value="Philips">
            <option value="Lexmark">
            <option value="Epson">
            <option value="Brother">
            <option value="Pantum">
          </datalist>
        </div>
        <div class="col-md-3"><div class="label">Model (Opsiyonel)</div><input class="form-control" name="model"></div>
        <div class="col-md-3"><div class="label">Seri No</div><input class="form-control" name="serial_no" required></div>
        <div class="col-12"><button class="btn btn-primary w-100">Kaydet</button></div>
      </form>
    </div>

    <div class="card p-4">
      <h6>Liste</h6>
      <div class="table-responsive">
        <table class="table table-striped align-middle">
          <thead>
            <tr>
              <th>Kategori</th><th>Marka</th><th>Model</th><th>Seri No</th>
              <th>Durum</th><th>Kime Zimmetli</th><th></th>
            </tr>
          </thead>
          <tbody>{rows or '<tr><td colspan="7">Kayıt yok.</td></tr>'}</tbody>
        </table>
      </div>
    </div>
    """
    return render_template_string(BASE_HTML, content=content, message=message, is_admin=is_admin())

# =========================================================
# DEPO
# =========================================================
@app.route("/depo", methods=["GET", "POST"])
def depo():
    if not require_login():
        return redirect(url_for("login"))

    message = None
    search_q = normalize_spaces(request.args.get("q") or request.form.get("search_q") or "")

    if request.method == "POST":
        action = request.form.get("action")
        if action == "add_item":
            try:
                new_name = normalize_spaces(request.form.get("new_name") or "")
                init_qty = int(request.form.get("init_qty") or 0)
                if not new_name:
                    raise ValueError("Malzeme adı boş olamaz.")
                if Stock.query.filter_by(name=new_name).first():
                    raise ValueError("Bu malzeme zaten depoda kayıtlı.")
                db.session.add(Stock(name=new_name, qty=max(0, init_qty)))
                db.session.commit()
                message = f"'{new_name}' depoya başarıyla eklendi."
            except Exception as e:
                db.session.rollback()
                message = f"Hata: {e}"

        elif action == "delete_item":
            try:
                stock_id = int(request.form.get("stock_id"))
                st = Stock.query.get(stock_id)
                if st:
                    name_del = st.name
                    db.session.delete(st)
                    db.session.commit()
                    message = f"'{name_del}' depodan silindi."
            except Exception as e:
                db.session.rollback()
                message = f"Hata: {e}"

        elif action == "update_qty" or request.form.get("op"):
            try:
                name = request.form.get("name")
                qty = int(request.form.get("qty") or 1)
                op = (request.form.get("op") or "plus").strip().lower()
                qty = 1 if qty == 0 else abs(qty)
                delta = qty if op == "plus" else -qty
                st = Stock.query.filter_by(name=name).first()
                if not st:
                    raise ValueError("Kalem bulunamadı.")
                new_qty = st.qty + delta
                if new_qty < 0:
                    raise ValueError("Stok eksiye düşemez.")
                st.qty = new_qty
                db.session.commit()
                message = "Stok miktarı güncellendi."
            except Exception as e:
                db.session.rollback()
                message = f"Hata: {e}"

    def tr_lower(s):
        if not s:
            return ""
        return str(s).translate(str.maketrans({
            "İ":"i", "I":"ı", "Ş":"ş", "ş":"ş", "Ğ":"ğ", "ğ":"ğ", "Ü":"ü", "ü":"ü", "Ö":"ö", "ö":"ö", "Ç":"ç", "ç":"ç"
        })).lower()

    search_q_lower = tr_lower(search_q)
    all_st = Stock.query.all()
    all_st.sort(key=lambda x: (STOCK_ITEMS.index(x.name) if x.name in STOCK_ITEMS else 10**9, x.name))

    total_items = len(all_st)
    total_qty = sum(st.qty for st in all_st)
    critical_count = sum(1 for st in all_st if st.qty <= 2)

    filtered_st = []
    for st in all_st:
        if search_q_lower and search_q_lower not in tr_lower(st.name):
            continue
        filtered_st.append(st)

    rows = ""
    for st in filtered_st:
        badge_cls = "badge-soft"
        dot_html = ""
        if st.qty == 0:
            badge_cls = "badge-soft bg-danger text-white border-0"
            dot_html = '<span class="pulse-red-dot me-2" title="Stok Tüklendi! (0 Adet)"></span>'
        elif st.qty <= 2:
            badge_cls = "badge-soft badge-yellow"
            dot_html = '<span class="pulse-red-dot me-2" title="Kritik Stok! (Tükenmek Üzere)"></span>'

        rows += f"""
        <tr>
          <td class="fw-bold">{dot_html}{st.name}</td>
          <td><span class="{badge_cls} px-3 py-1 fs-6">{st.qty} Adet</span></td>
          <td>
            <form method="post" class="d-flex gap-2 align-items-center flex-wrap">
              <input type="hidden" name="action" value="update_qty">
              <input type="hidden" name="name" value="{st.name}">
              <input type="number" name="qty" class="form-control" style="max-width:90px" value="1" min="1" step="1">
              <button class="btn btn-sm btn-success fw-bold" name="op" value="plus">+ Ekle</button>
              <button class="btn btn-sm btn-danger fw-bold" name="op" value="minus">- Düş</button>
            </form>
          </td>
          <td class="text-end">
            <form method="post" style="display:inline" onsubmit="return confirm('{st.name} malzemesini depodan silmek istediğinize emin misiniz?');">
              <input type="hidden" name="action" value="delete_item">
              <input type="hidden" name="stock_id" value="{st.id}">
              <button class="btn btn-sm btn-outline-danger"><i class="fa fa-trash"></i> Sil</button>
            </form>
          </td>
        </tr>
        """

    content = f"""
    <div class="card p-4 mb-3">
      <div class="d-flex justify-content-between align-items-center flex-wrap gap-2">
        <div>
          <h3 class="mb-1">Depo ve Stok Yönetimi</h3>
          <div class="label">Sarf malzemeleri ve stok takip paneli.</div>
        </div>
        <div class="d-flex gap-3">
          <div class="card px-3 py-2 text-center mb-0" style="min-width: 120px;">
            <div class="label">Toplam Kalem</div>
            <div class="fs-4 fw-bold">{total_items}</div>
          </div>
          <div class="card px-3 py-2 text-center mb-0" style="min-width: 120px;">
            <div class="label">Toplam Stok</div>
            <div class="fs-4 fw-bold text-success">{total_qty}</div>
          </div>
          <div class="card px-3 py-2 text-center mb-0" style="min-width: 120px;">
            <div class="label">Kritik Stok</div>
            <div class="fs-4 fw-bold text-warning">{critical_count}</div>
          </div>
        </div>
      </div>
    </div>

    <!-- Ekleme ve Arama Paneli -->
    <div class="row g-3 mb-3">
      <div class="col-md-6">
        <div class="card p-3 h-100">
          <h6><i class="fa fa-plus-circle text-primary me-2"></i>Yeni Malzeme Ekle</h6>
          <form method="post" class="row g-2 mt-1">
            <input type="hidden" name="action" value="add_item">
            <div class="col-7">
              <input class="form-control" name="new_name" placeholder="Malzeme Adı (Örn: Cat6 Kablo)" required>
            </div>
            <div class="col-3">
              <input type="number" class="form-control" name="init_qty" placeholder="Adet" value="0" min="0">
            </div>
            <div class="col-2">
              <button class="btn btn-primary w-100 fw-bold">+</button>
            </div>
          </form>
        </div>
      </div>
      <div class="col-md-6">
        <div class="card p-3 h-100">
          <h6><i class="fa fa-search text-primary me-2"></i>Depoda Malzeme Ara</h6>
          <form method="get" action="/depo" class="row g-2 mt-1">
            <div class="col-8">
              <input class="form-control" name="q" value="{search_q}" placeholder="Malzeme adı ile ara...">
            </div>
            <div class="col-4 d-flex gap-1">
              <button class="btn btn-primary flex-fill fw-bold"><i class="fa fa-search"></i> Ara</button>
              <a href="/depo" class="btn btn-outline-secondary"><i class="fa fa-times"></i></a>
            </div>
          </form>
        </div>
      </div>
    </div>

    <div class="card p-4">
      <div class="d-flex justify-content-between align-items-center mb-3">
        <h6 class="mb-0">Stok Listesi {f'("{search_q}" aramasına ait {len(filtered_st)} sonuç)' if search_q else ''}</h6>
      </div>
      <div class="table-responsive">
        <table class="table table-striped table-hover align-middle depo-table">
          <thead>
            <tr>
              <th>Malzeme / Kalem Adı</th>
              <th>Mevcut Stok</th>
              <th>Miktar Güncelle (+ / -)</th>
              <th class="text-end">İşlem</th>
            </tr>
          </thead>
          <tbody>{rows or '<tr><td colspan="4" class="text-center py-4 text-muted">Aramaya uygun malzeme bulunamadı.</td></tr>'}</tbody>
        </table>
      </div>
    </div>
    """
    return render_base(content, message=message)

# =========================================================
# GEÇMİŞ
# =========================================================
@app.route("/gecmis")
def gecmis():
    if not require_login():
        return redirect(url_for("login"))
    items = History.query.order_by(History.id.desc()).limit(300).all()
    rows = ""
    for h in items:
        action_badge = '<span class="badge bg-success">Zimmet Verildi</span>' if h.action == "ZIMMET" else '<span class="badge bg-warning text-dark">Teslim Alındı</span>'
        alan_info = f"<strong>{h.alan_ad or ''}</strong> ({h.alan_sicil or '-'})"
        if h.alan_birim:
            alan_info += f"<br><small class='text-muted'>{h.alan_birim}</small>"
        
        teslim_info = f"{h.teslim_ad or ''} ({h.teslim_sicil or '-'})"

        rows += f"""
        <tr>
          <td><span class="badge-soft"><i class="fa fa-clock me-1"></i>{h.at}</span></td>
          <td>{action_badge}</td>
          <td>{alan_info}</td>
          <td>{teslim_info}</td>
          <td><strong>{h.category or ''}</strong> - {(h.brand or '')} {(h.model or '')}</td>
          <td><code>{h.serial_no or ''}</code></td>
        </tr>
        """

    personnel_list = get_all_personnel_list()
    personnel_opts = ""
    for sicil_val, name_val in personnel_list:
        personnel_opts += f'<option value="{sicil_val}">{name_val} (Sicil: {sicil_val})</option>\n'

    # Daire opsiyonları HTML'i
    daire_options_html = ""
    for grp_title, unit_names in BAM_UNITS_STRUCTURE.items():
        daire_options_html += f'<optgroup label="--- {grp_title} ---">\n'
        for uname in unit_names:
            daire_options_html += f'  <option value="{uname}">{uname}</option>\n'
        daire_options_html += '</optgroup>\n'

    content = f"""
    <div class="card p-4 mb-3">
      <h3>Zimmet Geçmişi ve Raporlar</h3>
      <div class="label">Geçmiş işlem kayıtları, Personel Zimmet Kartı ve Daire Zimmet Defterleri üretme paneli.</div>
    </div>

    <!-- Daire / Birim Genel Zimmet Defteri Oluştur Paneli (Option 4) -->
    <div class="card p-4 mb-3">
      <h6><i class="fa fa-book text-warning me-2"></i>🏢 Daire / Birim Genel Zimmet Defteri Oluştur (PDF)</h6>
      <div class="label mb-3">Seçeceğiniz daire veya ortak alanın tüm aktif cihazlarını listeleyen kurumsal Genel Zimmet Defteri çıktısı alır.</div>

      <form method="get" action="/admin/export/daire_defteri.pdf" class="row g-3 align-items-end">
        <div class="col-md-7">
          <div class="label mb-1">Diyarbakır BAM Daire / Birim Seçiniz</div>
          <select name="daire_adi" class="form-select" required style="font-size: 14px; font-weight: 700;">
            <option value="" selected disabled>-- Diyarbakır BAM Birimi Seçiniz --</option>
            {daire_options_html}
          </select>
        </div>
        <div class="col-md-5">
          <button class="btn btn-warning w-100 fw-bold py-2 text-dark"><i class="fa fa-file-pdf me-1"></i> Daire Defteri (PDF) İndir</button>
        </div>
      </form>
    </div>

    <!-- Personel Zimmet Kartı Oluştur Paneli -->
    <div class="card p-4 mb-3">
      <h6><i class="fa fa-id-card text-primary me-2"></i>Personel Zimmet Kartı Oluştur (PDF)</h6>
      <div class="label mb-3">Seçeceğiniz personelin üzerinde zimmetli aktif cihazları ve geçmiş teslimatlarını listeleyen kurumsal kart çıktısı alır.</div>

      <form method="get" action="/admin/export/personel_kart.pdf" class="row g-3 align-items-end">
        <div class="col-md-7">
          <div class="label mb-1">Listeden Personel Seçiniz</div>
          <select name="sicil" class="form-select" required style="font-size: 14px; font-weight: 600;">
            <option value="" selected disabled>-- Personel Seçiniz --</option>
            {personnel_opts or '<option value="" disabled>Kayıtlı personel bulunamadı</option>'}
          </select>
        </div>
        <div class="col-md-5">
          <button class="btn btn-primary w-100 fw-bold py-2"><i class="fa fa-file-pdf me-1"></i> Zimmet Kartı (PDF) İndir</button>
        </div>
      </form>
    </div>

    <div class="card p-4">
      <h6 class="mb-3"><i class="fa fa-list text-primary me-2"></i>Tarih ve Saatli İşlem Geçmişi (Son 300 Kayıt)</h6>
      <div class="table-responsive">
        <table class="table table-striped table-hover align-middle" style="font-size: 13px;">
          <thead>
            <tr>
              <th>Tarih / Saat</th>
              <th>İşlem Türü</th>
              <th>Zimmet Alan / İade Eden</th>
              <th>İşlemi Yapan (Bilgi İşlem)</th>
              <th>Cihaz / Marka Model</th>
              <th>Seri No</th>
            </tr>
          </thead>
          <tbody>{rows or '<tr><td colspan="6" class="text-center py-3 text-muted">Kayıtlı geçmiş bulunamadı.</td></tr>'}</tbody>
        </table>
      </div>
    </div>
    """
    return render_template_string(BASE_HTML, content=content, message=None, is_admin=is_admin())

# =========================================================
# API: ZİMMET EVENT LIST (arama: name/sicil)
# =========================================================
@app.route("/api/zimmet_events")
def api_zimmet_events():
    if not require_login():
        return jsonify({"ok": False, "error": "unauthorized"}), 401

    name_q = normalize_spaces(request.args.get("name") or "")
    sicil_q = normalize_spaces(request.args.get("sicil") or "")

    def tr_lower(s):
        if not s:
            return ""
        return str(s).translate(str.maketrans({
            "İ":"i", "I":"ı", "Ş":"ş", "ş":"ş", "Ğ":"ğ", "ğ":"ğ", "Ü":"ü", "ü":"ü", "Ö":"ö", "ö":"ö", "Ç":"ç", "ç":"ç"
        })).lower()

    name_q_lower = tr_lower(name_q)
    sicil_q_lower = tr_lower(sicil_q)

    # Fetch distinct histories
    rows = (
        db.session.query(History.at, History.alan_sicil, History.alan_ad)
        .filter(History.action == "ZIMMET")
        .distinct()
        .order_by(History.at.desc())
        .all()
    )

    events = []
    for at, sicil, ad in rows:
        ad_lower = tr_lower(ad)
        sicil_lower = tr_lower(sicil)

        if name_q_lower and name_q_lower not in ad_lower:
            continue
        if sicil_q_lower and sicil_q_lower not in sicil_lower:
            continue

        events.append({
            "at": at,
            "sicil": sicil,
            "ad": ad,
            "label": f"{at} | {ad or ''} ({sicil or ''})".strip()
        })

    events = events[:200]
    return jsonify({"ok": True, "events": events})

@app.route("/api/zimmet_event")
def api_zimmet_event():
    if not require_login():
        return jsonify({"ok": False, "error": "unauthorized"}), 401
    at = normalize_spaces(request.args.get("at") or "")
    sicil = normalize_spaces(request.args.get("sicil") or "")
    if not at or not sicil:
        return jsonify({"ok": False, "error": "missing_params"}), 400

    q = History.query.filter_by(action="ZIMMET", at=at, alan_sicil=sicil)
    first = q.first()
    if not first:
        return jsonify({"ok": False, "error": "not_found"}), 404

    devices = []
    for h in q.all():
        inv_id = h.inv_id
        if not inv_id and h.serial_no:
            it = InventoryItem.query.filter_by(serial_no=h.serial_no).first()
            if it:
                inv_id = it.id
        devices.append({
            "inv_id": inv_id,
            "category": h.category or "",
            "brand": h.brand or "",
            "model": h.model or "",
            "serial_no": h.serial_no or "",
        })

    return jsonify({
        "ok": True,
        "alan_ad": first.alan_ad or "",
        "alan_sicil": first.alan_sicil or "",
        "alan_unvan": first.alan_unvan or "",
        "alan_birim": first.alan_birim or "",
        "bilgisayar_adi": first.bilgisayar_adi or "",
        "devices": devices,
    })

# =========================================================
# ZİMMET / TESLİM-TESELLÜM EKRANI
# =========================================================
@app.route("/zimmet")
def zimmet_page():
    if not require_login():
        return redirect(url_for("login"))

    depodakiler = InventoryItem.query.filter_by(status="Depoda").order_by(InventoryItem.id.desc()).all()
    zimmetliler = InventoryItem.query.filter_by(status="Zimmetli").order_by(InventoryItem.id.desc()).all()

    def checklist(items, prefix):
        if not items:
            return "<div class='label'>Kayıt yok.</div>"

        # Teslim Tesellüm tarafı: kişiye göre grupla (aç/kapa)
        if prefix == "t":
            grouped = {}
            for it in items:
                key = f"{it.assigned_name or 'Bilinmeyen'} ({it.assigned_sicil or ''})"
                grouped.setdefault(key, []).append(it)

            parts = []
            for person, arr in grouped.items():
                inner = []
                for it in arr:
                    label = f"{it.category} • {it.brand} {(it.model or '')} • {it.serial_no}"
                    inner.append(f"""
      <div class="form-check check-tight">
        <input class="form-check-input" type="checkbox" name="{prefix}_inv_ids" value="{it.id}" id="{prefix}_{it.id}">
        <label class="form-check-label" for="{prefix}_{it.id}" style="font-weight:800">{label}</label>
      </div>
                    """)
                parts.append(f"""
    <details style="border:1px solid var(--border);border-radius:12px;padding:10px;margin-bottom:10px;background:rgba(255,255,255,.03)">
      <summary style="cursor:pointer;font-weight:900;color:#f9fafb">{person}</summary>
      <div class="mt-2">
        {''.join(inner)}
      </div>
    </details>
                """)
            return "\n".join(parts)

        # Zimmet tarafı: düz liste
        out = []
        for it in items:
            label = f"{it.category} • {it.brand} {(it.model or '')} • {it.serial_no}"
            out.append(f"""
      <div class="form-check check-tight">
        <input class="form-check-input" type="checkbox" name="{prefix}_inv_ids" value="{it.id}" id="{prefix}_{it.id}">
        <label class="form-check-label" for="{prefix}_{it.id}" style="font-weight:800">{label}</label>
      </div>
            """)
        return "\n".join(out)

    z_list = checklist(depodakiler, "z")
    t_list = checklist(zimmetliler, "t")

    teslim_ad = session.get("user","")
    teslim_sicil = session.get("sicil","")

    daire_options_html = ""
    for group_name, unit_list in BAM_UNITS_STRUCTURE.items():
        daire_options_html += f'<optgroup label="{group_name}">'
        for u in unit_list:
            daire_options_html += f'<option value="{u}">{u}</option>'
        daire_options_html += '</optgroup>'

    z_daire_list = checklist(depodakiler, "z_daire")

    # İçerik template (JS içinde { } kullanabilmek için f-string değil, placeholder ile)
    tpl = r"""
    <div class="card p-4 mb-3">
      <h3>Zimmet / Teslim Tesellüm</h3>
      <div class="label">Personel veya Daire/Ortak Alan bazında zimmetleme ve iade alma işlemleri.</div>
    </div>

    <ul class="nav nav-tabs mb-3" role="tablist">
      <li class="nav-item"><button class="nav-link active fw-bold" data-bs-toggle="tab" data-bs-target="#tab_z" type="button"><i class="fa fa-user me-1"></i> 1) Personel Zimmeti (Şahıs)</button></li>
      <li class="nav-item"><button class="nav-link fw-bold" data-bs-toggle="tab" data-bs-target="#tab_daire" type="button"><i class="fa fa-building me-1 text-warning"></i> 2) Daire / Ortak Alan Zimmeti (Diyarbakır BAM)</button></li>
      <li class="nav-item"><button class="nav-link fw-bold" data-bs-toggle="tab" data-bs-target="#tab_t" type="button"><i class="fa fa-rotate-left me-1"></i> 3) Teslim Tesellüm (İade Al)</button></li>
    </ul>

    <div class="tab-content">

      <!-- ================= ZIMMET TAB ================= -->
      <div class="tab-pane fade show active" id="tab_z">

        <div class="card p-4 mb-3">
          <h6>0) Kayıtlı Zimmetten Otomatik Getir</h6>
          <div class="label">İsim/sicil ile ara → seç → form ve cihaz satırları dolsun.</div>

          <div class="row g-2 align-items-end mt-2">
            <div class="col-md-4">
              <div class="label">İsim ile ara</div>
              <input class="form-control" id="z_search_name" placeholder="Ad soyad (örn: Murat)">
            </div>
            <div class="col-md-3">
              <div class="label">Sicil ile ara</div>
              <input class="form-control" id="z_search_sicil" placeholder="187665">
            </div>
            <div class="col-md-5 d-flex gap-2">
              <button type="button" class="btn btn-outline-primary flex-fill" onclick="zSearchName()">İsme Göre Ara</button>
              <button type="button" class="btn btn-outline-primary flex-fill" onclick="zSearchSicil()">Sicile Göre Ara</button>
              <button type="button" class="btn btn-outline-secondary" onclick="zReset()">Temizle</button>
            </div>
          </div>

          <div class="row g-3 align-items-end mt-2">
            <div class="col-md-8">
              <div class="label">Zimmet Kaydı</div>
              <select class="form-select" id="z_event_select">
                <option value="">Seçiniz...</option>
              </select>
            </div>
            <div class="col-md-4 d-grid">
              <button type="button" class="btn btn-success" style="background:var(--primary);border:none" onclick="zLoad()">Seçili Zimmeti Getir</button>
            </div>
          </div>
        </div>

        <form method="post" action="/pdf_zimmet" onsubmit="return submitManual('z')">
          <input type="hidden" name="manual_json" id="z_manual_json">

          <div class="card p-4 mb-3">
            <h6>1) Envanterden Seç (Depoda Olanlar)</h6>
            <div style="max-height:240px;overflow:auto;border:1px solid var(--border);border-radius:12px;padding:10px;background:var(--card)">
              __Z_LIST__
            </div>
          </div>

          <div class="card p-4 mb-3">
            <h6>2) Manuel Cihaz Satırı Ekle</h6>
            <div class="label mb-2">Envanter dışı / geçici kayıtlar için satır ekleyebilirsin.</div>
            <div id="z_rows"></div>
            <div class="d-flex gap-2 mt-2">
              <button type="button" class="btn btn-outline-primary" onclick="addRow('z_rows')">Satır Ekle</button>
              <button type="button" class="btn btn-outline-secondary" onclick="deleteSelectedRows('z_rows')">Satır Sil</button>
            </div>
          </div>

          <div class="card p-4">
            <div class="row g-3">
              <div class="col-12"><h6>Zimmet Alan</h6></div>
              <div class="col-md-4"><div class="label">Ad Soyad</div><input class="form-control" id="z_alan_ad" name="alan_ad" required></div>
              <div class="col-md-2"><div class="label">Sicil</div><input class="form-control" id="z_alan_sicil" name="alan_sicil" required></div>
              <div class="col-md-3"><div class="label">Ünvan</div><input class="form-control" id="z_alan_unvan" name="alan_unvan" required></div>
              <div class="col-md-3"><div class="label">Birim</div><input class="form-control" id="z_alan_birim" name="alan_birim" required></div>
              <div class="col-md-4"><div class="label">Bilgisayar Adı (Opsiyonel)</div><input class="form-control" id="z_pc" name="bilgisayar_adi"></div>

              <div class="col-12"><hr></div>
              <div class="col-12"><h6>Teslim Eden (Bilgi İşlem)</h6></div>
              <div class="col-md-4"><div class="label">Ad Soyad</div><input class="form-control" name="teslim_ad" value="__TESLIM_AD__" required></div>
              <div class="col-md-2"><div class="label">Sicil</div><input class="form-control" name="teslim_sicil" value="__TESLIM_SICIL__" required></div>
              <div class="col-md-6"><div class="label">Ünvan</div><input class="form-control" name="teslim_unvan" value="Bilgi İşlem Personeli" required></div>

              <div class="col-12"><button class="btn btn-primary w-100 fw-bold">Zimmet PDF Oluştur</button></div>
            </div>
          </div>
        </form>
      </div>

      <!-- ================= DAIRE ZIMMET TAB ================= -->
      <div class="tab-pane fade" id="tab_daire">
        <form method="post" action="/pdf_daire_zimmet" onsubmit="return submitManual('z_daire')">
          <input type="hidden" name="manual_json" id="z_daire_manual_json">

          <div class="card p-4 mb-3">
            <h6><i class="fa fa-building text-warning me-2"></i>1) Daire / Ortak Alan ve Makam Seçimi</h6>
            <div class="label mb-3">Cihazı şahsa değil, Diyarbakır BAM Daire, Duruşma Salonu veya İdari Bürolarına zimmetler. Evrak Daire Yazı İşleri Müdürü adına düzenlenir.</div>

            <div class="row g-3">
              <div class="col-md-6">
                <div class="label mb-1">Daire / Birim Seçiniz (veya Aşağıdan Manuel Yazınız)</div>
                <select name="daire_adi_select" id="daire_adi_select" class="form-select" onchange="if(this.value) document.getElementById('daire_adi_input').value=this.value;" style="font-weight:700">
                  <option value="" selected disabled>-- Diyarbakır BAM Birimi Seçiniz --</option>
                  __DAIRE_OPTIONS__
                </select>
              </div>
              <div class="col-md-6">
                <div class="label mb-1">Daire / Birim Adı (Seçiniz veya Manuel Yazınız)</div>
                <input class="form-control" name="daire_adi" id="daire_adi_input" placeholder="Örn: 1. Ceza Dairesi / Kalem / Sistem Odası" required style="font-weight:700">
              </div>
              <div class="col-md-6">
                <div class="label mb-1">Zimmetlenen Makam / Verilen Envanter Unvanı</div>
                <input class="form-control" name="makam_title" placeholder="Örn: Daire Başkanı / Üye Hakim Odası / C. Savcısı Odası / Masa 3" required>
              </div>
              <div class="col-md-6">
                <div class="label mb-1">Daire Yazı İşleri Müdürü Adı Soyadı</div>
                <input class="form-control" name="yazi_isleri_muduru_ad" placeholder="Örn: Mehmet ÖZTÜRK" required>
              </div>
              <div class="col-md-6">
                <div class="label mb-1">Daire Yazı İşleri Müdürü Sicil No</div>
                <input class="form-control" name="yazi_isleri_muduru_sicil" placeholder="Örn: 174829" required>
              </div>
            </div>
          </div>

          <div class="card p-4 mb-3">
            <h6>2) Envanterden Seç (Depoda Olanlar)</h6>
            <div style="max-height:240px;overflow:auto;border:1px solid var(--border);border-radius:12px;padding:10px;background:var(--card)">
              __Z_DAIRE_LIST__
            </div>
          </div>

          <div class="card p-4 mb-3">
            <h6>3) Manuel Cihaz Satırı Ekle (Opsiyonel)</h6>
            <div id="z_daire_rows"></div>
            <div class="d-flex gap-2 mt-2">
              <button type="button" class="btn btn-outline-primary" onclick="addRow('z_daire_rows')">Satır Ekle</button>
              <button type="button" class="btn btn-outline-secondary" onclick="deleteSelectedRows('z_daire_rows')">Satır Sil</button>
            </div>
          </div>

          <div class="card p-4">
            <div class="row g-3">
              <div class="col-12"><h6>Teslim Eden (Bilgi İşlem Müdürlüğü)</h6></div>
              <div class="col-md-4"><div class="label">Ad Soyad</div><input class="form-control" name="teslim_ad" value="__TESLIM_AD__" required></div>
              <div class="col-md-2"><div class="label">Sicil</div><input class="form-control" name="teslim_sicil" value="__TESLIM_SICIL__" required></div>
              <div class="col-md-6 d-flex align-items-end"><button class="btn w-100 fw-bold py-2" style="background:#f59e0b;border:none;color:#111827"><i class="fa fa-file-pdf me-1"></i> Daire Zimmet Tutanağı (PDF) Oluştur</button></div>
            </div>
          </div>
        </form>
      </div>

      <!-- ================= TESLIM TAB ================= -->
      <div class="tab-pane fade" id="tab_t">

        <div class="card p-4 mb-3">
          <h6>0) Kayıtlı Zimmetten Otomatik Getir</h6>
          <div class="label">İsim/sicil ile ara → seç → iade eden + cihaz satırları dolsun.</div>

          <div class="row g-2 align-items-end mt-2">
            <div class="col-md-4">
              <div class="label">İsim ile ara</div>
              <input class="form-control" id="t_search_name" placeholder="Ad soyad (örn: Murat)">
            </div>
            <div class="col-md-3">
              <div class="label">Sicil ile ara</div>
              <input class="form-control" id="t_search_sicil" placeholder="187665">
            </div>
            <div class="col-md-5 d-flex gap-2">
              <button type="button" class="btn btn-outline-primary flex-fill" onclick="tSearchName()">İsme Göre Ara</button>
              <button type="button" class="btn btn-outline-primary flex-fill" onclick="tSearchSicil()">Sicile Göre Ara</button>
              <button type="button" class="btn btn-outline-secondary" onclick="tReset()">Temizle</button>
            </div>
          </div>

          <div class="row g-3 align-items-end mt-2">
            <div class="col-md-8">
              <div class="label">Zimmet Kaydı</div>
              <select class="form-select" id="t_event_select">
                <option value="">Seçiniz...</option>
              </select>
            </div>
            <div class="col-md-4 d-grid">
              <button type="button" class="btn btn-outline-primary" onclick="tLoad()">Seçili Zimmeti Getir</button>
            </div>
          </div>
        </div>

        <form method="post" action="/pdf_teslim" onsubmit="return submitManual('t')">
          <input type="hidden" name="manual_json" id="t_manual_json">

          <div class="card p-4 mb-3">
            <h6>1) Zimmetli Cihazlardan Seç</h6>
            <div style="border:1px solid var(--border);border-radius:12px;padding:10px;background:var(--card)">
              __T_LIST__
            </div>
          </div>

          <div class="card p-4 mb-3">
            <h6>2) Manuel Cihaz Satırı Ekle</h6>
            <div class="label mb-2">Envanter dışı / geçici kayıtlar için satır ekleyebilirsin.</div>
            <div id="t_rows"></div>
            <div class="d-flex gap-2 mt-2">
              <button type="button" class="btn btn-outline-primary" onclick="addRow('t_rows')">Satır Ekle</button>
              <button type="button" class="btn btn-outline-secondary" onclick="deleteSelectedRows('t_rows')">Satır Sil</button>
            </div>
          </div>

          <div class="card p-4">
            <div class="row g-3">
              <div class="col-12"><h6>İade Eden</h6></div>
              <div class="col-md-4"><div class="label">Ad Soyad</div><input class="form-control" id="t_alan_ad" name="alan_ad" required></div>
              <div class="col-md-2"><div class="label">Sicil</div><input class="form-control" id="t_alan_sicil" name="alan_sicil" required></div>

              <div class="col-12"><hr></div>
              <div class="col-12"><h6>Teslim Alan (Bilgi İşlem)</h6></div>
              <div class="col-md-4"><div class="label">Ad Soyad</div><input class="form-control" name="teslim_ad" value="__TESLIM_AD__" required></div>
              <div class="col-md-2"><div class="label">Sicil</div><input class="form-control" name="teslim_sicil" value="__TESLIM_SICIL__" required></div>

              <div class="col-12"><button class="btn w-100" style="background:#f59e0b;border:none;color:#111827;font-weight:900">Teslim Tesellüm PDF Oluştur</button></div>
            </div>
          </div>
        </form>
      </div>

      </div>
    </div>

    <script>
    function esc(v){ return (v ?? "").toString().trim(); }

    function applyBrandMode(row){
      const cat = esc(row.querySelector('[data-k="category"]')?.value);
      const brandSel = row.querySelector('[data-k="brand_select"]');
      const brandInp = row.querySelector('[data-k="brand"]');
      if(!brandSel || !brandInp) return;

      const MAP = {
        "Kasa": ["Lenovo","Dell"],
        "Monitör": ["Lenovo","Dell"],
        "Yazıcı": ["Lexmark","Epson","Brother","Pantum"],
        "Tarayıcı": ["Canon","Avision","Ricoh"]
      };

      const opts = MAP[cat] || null;

      if(opts){
        brandSel.style.display = "";
        // seçenekleri yeniden kur
        brandSel.innerHTML = "";
        for(const o of opts){
          const op = document.createElement("option");
          op.value = o; op.textContent = o;
          brandSel.appendChild(op);
        }
        const opm = document.createElement("option");
        opm.value = "__manual__"; opm.textContent = "Elle gir";
        brandSel.appendChild(opm);

        const cur = esc(brandInp.value);
        if(opts.includes(cur)){
          brandSel.value = cur;
          brandInp.value = cur;
          brandInp.style.display = "none";
        }else{
          brandSel.value = "__manual__";
          brandInp.style.display = "";
        }
      }else{
        brandSel.style.display = "none";
        brandInp.style.display = "";
      }
    }

    function wireRow(row){
      const catSel = row.querySelector('[data-k="category"]');
      const brandSel = row.querySelector('[data-k="brand_select"]');
      const brandInp = row.querySelector('[data-k="brand"]');
      if(catSel){
        catSel.addEventListener("change", ()=> applyBrandMode(row));
      }
      if(brandSel){
        brandSel.addEventListener("change", ()=>{
          if(brandSel.value === "__manual__"){
            if(brandInp){
              brandInp.value = "";
              brandInp.style.display = "";
              brandInp.focus();
            }
          }else{
            if(brandInp){
              brandInp.value = brandSel.value;
              brandInp.style.display = "none";
            }
          }
        });
      }
      // initial
      applyBrandMode(row);
    }

    function addRow(containerId){
      const el = document.getElementById(containerId);
      const idx = Date.now().toString(36) + Math.random().toString(36).slice(2,6);
      const row = document.createElement("div");
      row.className = "card p-3 mb-2";
      row.setAttribute("data-row","1");
      row.innerHTML = `
        <div class="d-flex align-items-center gap-2 mb-2">
          <input type="checkbox" class="form-check-input" data-del="1" title="Sil">
          <div class="label">Manuel cihaz satırı</div>
        </div>
        <div class="row g-2">
          <div class="col-md-3"><div class="label">Kategori</div><select class="form-select" data-k="category"><option value="">Seçiniz...</option><option>Kasa</option><option>Monitör</option><option>Yazıcı</option><option>Tarayıcı</option></select></div>
          <div class="col-md-3"><div class="label">Marka</div><div class="d-flex gap-2"><select class="form-select" data-k="brand_select" style="max-width:170px; display:none"></select><input class="form-control" data-k="brand" placeholder="Marka"></div></div>
          <div class="col-md-3"><div class="label">Model (Opsiyonel)</div><input class="form-control" data-k="model" placeholder="Optiplex..."></div>
          <div class="col-md-3"><div class="label">Seri No</div><input class="form-control" data-k="serial_no" placeholder="SER123..."></div>
        </div>
      `;
      el.appendChild(row);
      wireRow(row);
    }

    function deleteSelectedRows(containerId){
      const el = document.getElementById(containerId);
      [...el.querySelectorAll('[data-row="1"]')].forEach(r=>{
        const cb = r.querySelector('input[type="checkbox"][data-del="1"]');
        if(cb && cb.checked) r.remove();
      });
    }

    function collectManual(containerId){
      const el = document.getElementById(containerId);
      const rows = [...el.querySelectorAll('[data-row="1"]')];
      const out = [];
      for(const r of rows){
        const get = (k)=> esc(r.querySelector(`[data-k="${k}"]`)?.value);
        const d = {category:get("category"), brand:get("brand"), model:get("model"), serial_no:get("serial_no")};
        if(d.category || d.brand || d.model || d.serial_no) out.push(d);
      }
      return out;
    }

    function submitManual(which){
      if(which === "z"){
        document.getElementById("z_manual_json").value = JSON.stringify(collectManual("z_rows"));
      }else if(which === "z_daire"){
        document.getElementById("z_daire_manual_json").value = JSON.stringify(collectManual("z_daire_rows"));
      }else{
        document.getElementById("t_manual_json").value = JSON.stringify(collectManual("t_rows"));
      }
      return true;
    }

    async function fetchEvents(name, sicil){
      const params = new URLSearchParams();
      if(name) params.set("name", name);
      if(sicil) params.set("sicil", sicil);
      const res = await fetch("/api/zimmet_events?" + params.toString());
      const js = await res.json();
      return js.ok ? js.events : [];
    }

    function fillSelect(selId, events){
      const sel = document.getElementById(selId);
      sel.innerHTML = `<option value="">Seçiniz...</option>`;
      for(const e of events){
        const opt = document.createElement("option");
        opt.value = JSON.stringify({at:e.at, sicil:e.sicil});
        opt.textContent = e.label;
        sel.appendChild(opt);
      }
    }

    function cleanQuery(val) {
      if(!val) return "";
      let s = String(val).trim();
      if(s.startsWith("Ad soyad") || s.includes("örn:")) return "";
      return s;
    }

    async function zSearchName(){
      const rawName = document.getElementById("z_search_name").value;
      const rawSicil = document.getElementById("z_search_sicil").value;
      const name = cleanQuery(rawName);
      const sicil = cleanQuery(rawSicil);
      const events = await fetchEvents(name, sicil);
      fillSelect("z_event_select", events);
      if(events.length === 1) {
        document.getElementById("z_event_select").selectedIndex = 1;
        zLoad();
      }
    }
    async function zSearchSicil(){
      return zSearchName();
    }
    function zReset(){
      document.getElementById("z_search_name").value="";
      document.getElementById("z_search_sicil").value="";
      fillSelect("z_event_select", []);
    }

    async function tSearchName(){
      const rawName = document.getElementById("t_search_name").value;
      const rawSicil = document.getElementById("t_search_sicil").value;
      const name = cleanQuery(rawName);
      const sicil = cleanQuery(rawSicil);
      const events = await fetchEvents(name, sicil);
      fillSelect("t_event_select", events);
      if(events.length === 1) {
        document.getElementById("t_event_select").selectedIndex = 1;
        tLoad();
      }
    }
    async function tSearchSicil(){
      return tSearchName();
    }
    function tReset(){
      document.getElementById("t_search_name").value="";
      document.getElementById("t_search_sicil").value="";
      fillSelect("t_event_select", []);
    }

    async function loadEvent(selId){
      const sel = document.getElementById(selId);
      if(!sel.value) return null;
      const v = JSON.parse(sel.value);
      const params = new URLSearchParams({at:v.at, sicil:v.sicil});
      const res = await fetch("/api/zimmet_event?" + params.toString());
      const js = await res.json();
      return js.ok ? js : null;
    }

    async function zLoad(){
      const data = await loadEvent("z_event_select");
      if(!data) return;
      document.getElementById("z_alan_ad").value = data.alan_ad || "";
      document.getElementById("z_alan_sicil").value = data.alan_sicil || "";
      document.getElementById("z_alan_unvan").value = data.alan_unvan || "";
      document.getElementById("z_alan_birim").value = data.alan_birim || "";
      document.getElementById("z_pc").value = data.bilgisayar_adi || "";

      // manuel satırlara cihazları bas (envanter checkboxlarını otomatik işaretlemiyoruz)
      const cont = document.getElementById("z_rows");
      cont.innerHTML = "";
      (data.devices || []).forEach(d=>{
        addRow("z_rows");
        const row = cont.lastElementChild;
        row.querySelector('[data-k="category"]').value = d.category || "";
        row.querySelector('[data-k="brand"]').value = d.brand || "";
        row.querySelector('[data-k="model"]').value = d.model || "";
        row.querySelector('[data-k="serial_no"]').value = d.serial_no || "";
        applyBrandMode(row);
      });
    }

    async function tLoad(){
      const data = await loadEvent("t_event_select");
      if(!data) return;
      document.getElementById("t_alan_ad").value = data.alan_ad || "";
      document.getElementById("t_alan_sicil").value = data.alan_sicil || "";

      const cont = document.getElementById("t_rows");
      cont.innerHTML = "";
      (data.devices || []).forEach(d=>{
        addRow("t_rows");
        const row = cont.lastElementChild;
        row.querySelector('[data-k="category"]').value = d.category || "";
        row.querySelector('[data-k="brand"]').value = d.brand || "";
        row.querySelector('[data-k="model"]').value = d.model || "";
        row.querySelector('[data-k="serial_no"]').value = d.serial_no || "";
        applyBrandMode(row);
      });
    }
    </script>
    """

    html = (tpl
        .replace("__Z_LIST__", z_list)
        .replace("__T_LIST__", t_list)
        .replace("__Z_DAIRE_LIST__", z_daire_list)
        .replace("__DAIRE_OPTIONS__", daire_options_html)
        .replace("__TESLIM_AD__", teslim_ad)
        .replace("__TESLIM_SICIL__", teslim_sicil)
    )

    return render_template_string(BASE_HTML, content=html, message=None, is_admin=is_admin())

# =========================================================
# PDF ROUTES
# =========================================================
@app.route("/pdf_zimmet", methods=["GET", "POST"])
def pdf_zimmet():
    if not require_login():
        return redirect(url_for("login"))
    if request.method == "GET":
        return redirect(url_for("view_pdf", type="zimmet"))

    try:
        # Form alanları
        alan_ad = normalize_spaces(request.form.get("alan_ad") or "")
        alan_sicil = normalize_spaces(request.form.get("alan_sicil") or "")
        alan_unvan = normalize_spaces(request.form.get("alan_unvan") or "")
        alan_birim = normalize_spaces(request.form.get("alan_birim") or "")
        bilgisayar_adi = normalize_spaces(request.form.get("bilgisayar_adi") or "")

        teslim_ad = normalize_spaces(request.form.get("teslim_ad") or session.get("user") or "")
        teslim_sicil = normalize_spaces(request.form.get("teslim_sicil") or session.get("sicil") or "")
        teslim_unvan = normalize_spaces(request.form.get("teslim_unvan") or "Bilgi İşlem Personeli")

        if not (alan_ad and alan_sicil and alan_unvan and alan_birim):
            raise ValueError("Zimmet alan bilgileri eksik.")
        if not (teslim_ad and teslim_sicil and teslim_unvan):
            raise ValueError("Teslim eden bilgileri eksik.")

        inv_ids = parse_checked_ids("z")
        manual = parse_manual_devices()

        devices: List[Dict] = []

        # Envanterden seçilenler
        for iid in inv_ids:
            it = InventoryItem.query.get(iid)
            if not it:
                continue
            devices.append({
                "inv_id": it.id,
                "category": it.category or "",
                "brand": it.brand or "",
                "model": it.model or "",
                "serial_no": it.serial_no or "",
            })

        # Manuel satırlar
        for d in manual:
            cat = normalize_spaces(d.get("category") or "")
            brand = normalize_spaces(d.get("brand") or "")
            model = normalize_spaces(d.get("model") or "")
            serial = normalize_spaces(d.get("serial_no") or "")
            if cat or brand or model or serial:
                devices.append({
                    "inv_id": None,
                    "category": cat,
                    "brand": brand,
                    "model": model,
                    "serial_no": serial,
                })

        if not devices:
            raise ValueError("Cihaz seçmediniz. (Envanterden işaretleyin veya manuel satır ekleyin.)")

        # DB güncelle: envanter cihazlarını zimmetle
        batch_at = now_str()
        for d in devices:
            if d.get("inv_id"):
                it = InventoryItem.query.get(int(d["inv_id"]))
                if not it:
                    continue
                it.status = "Zimmetli"
                it.assigned_name = alan_ad
                it.assigned_sicil = alan_sicil
                it.assigned_title = alan_unvan
                it.assigned_unit = alan_birim
                it.assigned_at = today_str()
                it.last_event = "ZIMMET"
                it.last_event_at = batch_at

            # History kaydı (manuel de yazılsın)
            db.session.add(History(
                at=batch_at,
                action="ZIMMET",
                inv_id=d.get("inv_id"),
                category=d.get("category"),
                brand=d.get("brand"),
                model=d.get("model"),
                serial_no=d.get("serial_no"),
                alan_ad=alan_ad,
                alan_sicil=alan_sicil,
                alan_unvan=alan_unvan,
                alan_birim=alan_birim,
                bilgisayar_adi=bilgisayar_adi,
                teslim_ad=teslim_ad,
                teslim_sicil=teslim_sicil,
                teslim_unvan=teslim_unvan,
            ))
        db.session.commit()

        pdf_bytes = build_zimmet_fis_pdf(
            alan_ad=alan_ad,
            alan_sicil=alan_sicil,
            alan_unvan=alan_unvan,
            alan_birim=alan_birim,
            bilgisayar_adi=bilgisayar_adi,
            teslim_ad=teslim_ad,
            teslim_sicil=teslim_sicil,
            teslim_unvan=teslim_unvan,
            devices=devices,
        )

        sicil = session.get("sicil") or "unknown"
        temp_filename = f"temp_{sicil}_zimmet.pdf"
        temp_path = os.path.join(APP_DIR, temp_filename)
        with open(temp_path, "wb") as f:
            f.write(pdf_bytes)

        return redirect(url_for("view_pdf_page", type="zimmet"))

    except Exception as e:
        db.session.rollback()
        return render_template_string(
            BASE_HTML,
            content="<div class='card p-4'><h5>Hata oluştu.</h5><pre>"+traceback.format_exc()+"</pre></div>",
            message=f"PDF üretilemedi: {e}",
        )


@app.route("/pdf_teslim", methods=["GET", "POST"])
def pdf_teslim():
    if not require_login():
        return redirect(url_for("login"))
    if request.method == "GET":
        return redirect(url_for("view_pdf", type="teslim"))

    try:
        # Form alanları
        alan_ad = normalize_spaces(request.form.get("alan_ad") or "")      # iade eden
        alan_sicil = normalize_spaces(request.form.get("alan_sicil") or "")

        teslim_ad = normalize_spaces(request.form.get("teslim_ad") or session.get("user") or "")
        teslim_sicil = normalize_spaces(request.form.get("teslim_sicil") or session.get("sicil") or "")

        if not (alan_ad and alan_sicil):
            raise ValueError("İade eden bilgileri eksik.")
        if not (teslim_ad and teslim_sicil):
            raise ValueError("Teslim alan bilgileri eksik.")

        inv_ids = parse_checked_ids("t")
        manual = parse_manual_devices()

        devices: List[Dict] = []

        # Envanterden seçilen (zimmetli) cihazlar: depoya al
        for iid in inv_ids:
            it = InventoryItem.query.get(iid)
            if not it:
                continue
            devices.append({
                "inv_id": it.id,
                "category": it.category or "",
                "brand": it.brand or "",
                "model": it.model or "",
                "serial_no": it.serial_no or "",
            })

        # Manuel satırlar
        for d in manual:
            cat = normalize_spaces(d.get("category") or "")
            brand = normalize_spaces(d.get("brand") or "")
            model = normalize_spaces(d.get("model") or "")
            serial = normalize_spaces(d.get("serial_no") or "")
            if cat or brand or model or serial:
                devices.append({
                    "inv_id": None,
                    "category": cat,
                    "brand": brand,
                    "model": model,
                    "serial_no": serial,
                })

        if not devices:
            raise ValueError("Cihaz seçmediniz. (Zimmetlilerden işaretleyin veya manuel satır ekleyin.)")

        batch_at = now_str()
        for d in devices:
            if d.get("inv_id"):
                it = InventoryItem.query.get(int(d["inv_id"]))
                if it:
                    it.status = "Depoda"
                    it.assigned_name = None
                    it.assigned_sicil = None
                    it.assigned_title = None
                    it.assigned_unit = None
                    it.assigned_at = None
                    it.last_event = "TESLIM_TESELLUM"
                    it.last_event_at = batch_at

            db.session.add(History(
                at=batch_at,
                action="TESLIM_TESELLUM",
                inv_id=d.get("inv_id"),
                category=d.get("category"),
                brand=d.get("brand"),
                model=d.get("model"),
                serial_no=d.get("serial_no"),
                alan_ad=alan_ad,
                alan_sicil=alan_sicil,
                teslim_ad=teslim_ad,
                teslim_sicil=teslim_sicil,
            ))
        db.session.commit()

        pdf_bytes = build_teslim_tesellum_pdf(
            iade_eden_ad=alan_ad,
            iade_eden_sicil=alan_sicil,
            teslim_alan_ad=teslim_ad,
            teslim_alan_sicil=teslim_sicil,
            devices=devices,
        )

        sicil = session.get("sicil") or "unknown"
        temp_filename = f"temp_{sicil}_teslim.pdf"
        temp_path = os.path.join(APP_DIR, temp_filename)
        with open(temp_path, "wb") as f:
            f.write(pdf_bytes)

        return redirect(url_for("view_pdf_page", type="teslim"))

    except Exception as e:
        db.session.rollback()
        return render_template_string(
            BASE_HTML,
            content="<div class='card p-4'><h5>Hata oluştu.</h5><pre>"+traceback.format_exc()+"</pre></div>",
            message=f"PDF üretilemedi: {e}",
        )


@app.route("/view_pdf", methods=["GET", "POST"])
def view_pdf():
    if not require_login():
        return redirect(url_for("login"))
    pdf_type = request.args.get("type", "zimmet")
    sicil = session.get("sicil") or "unknown"
    filename = f"temp_{sicil}_{pdf_type}.pdf"
    path = os.path.join(APP_DIR, filename)
    if not os.path.exists(path):
        return render_template_string(
            BASE_HTML,
            content="<div class='card p-4'><h5>Dosya Bulunamadı</h5><div class='label'>PDF belgesi sunucuda bulunamadı veya henüz oluşturulmadı. Lütfen formu tekrar gönderin.</div></div>",
            message=None,
            is_admin=is_admin()
        )
    return send_file(path, mimetype="application/pdf")


@app.route("/view_pdf_page")
def view_pdf_page():
    if not require_login():
        return redirect(url_for("login"))
    pdf_type = request.args.get("type", "zimmet")

    title_text = "Personel Zimmet Kartı" if pdf_type == "personel_kart" else ("Zimmet Fişi" if pdf_type == "zimmet" else "Teslim Tesellüm Belgesi")
    back_url = "/gecmis" if pdf_type == "personel_kart" else "/zimmet"

    html = f"""
    <!DOCTYPE html>
    <html lang="tr">
    <head>
      <meta charset="UTF-8">
      <title>{title_text} Önizleme</title>
      <link href="https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700;800&display=swap" rel="stylesheet">
      <style>
        :root {{
          --bg: #0b0f19;
          --card: #151b2c;
          --border: #232d45;
          --text: #e2e8f0;
          --primary: #6366f1;
          --success: #10b981;
        }}
        body {{
          margin: 0;
          padding: 0;
          background-color: var(--bg);
          color: var(--text);
          font-family: 'Inter', sans-serif;
          display: flex;
          flex-direction: column;
          height: 100vh;
          overflow: hidden;
        }}
        .footer-right {{
          position: fixed;
          left: 280px;
          bottom: 18px;
          color: var(--muted);
          font-size: 11px;
          font-weight: 600;
          z-index: 100;
        }}
        .header-bar {{
          height: 60px;
          background: var(--card);
          border-bottom: 1px solid var(--border);
          display: flex;
          align-items: center;
          justify-content: space-between;
          padding: 0 20px;
          box-sizing: border-box;
        }}
        .btn {{
          padding: 8px 16px;
          border-radius: 8px;
          font-weight: 600;
          font-size: 14px;
          cursor: pointer;
          text-decoration: none;
          display: inline-flex;
          align-items: center;
          gap: 6px;
          border: none;
          transition: all 0.2s;
        }}
        .btn-secondary {{
          background: rgba(255,255,255,0.07);
          color: var(--text);
          border: 1px solid var(--border);
        }}
        .btn-secondary:hover {{
          background: rgba(255,255,255,0.12);
        }}
        .btn-primary {{
          background: var(--primary);
          color: white;
        }}
        .btn-primary:hover {{
          background: #4f46e5;
        }}
        .btn-success {{
          background: var(--success);
          color: #111827;
          font-weight: 700;
        }}
        .btn-success:hover {{
          background: #059669;
        }}
        .title {{
          font-size: 16px;
          font-weight: 700;
        }}
        .actions {{
          display: flex;
          gap: 10px;
        }}
        .iframe-container {{
          flex: 1;
          position: relative;
        }}
        iframe {{
          width: 100%;
          height: 100%;
          border: none;
        }}
      </style>
    </head>
    <body>
      <div class="header-bar">
        <button class="btn btn-secondary" onclick="goBack()">← Geri Dön</button>
        <div class="title">{title_text} Önizleme</div>
        <div class="actions">
          <button class="btn btn-primary" onclick="printPdf()">Yazdır</button>
          <a href="/download_pdf?type={pdf_type}" target="_blank" class="btn btn-success"><i class="fa fa-download me-1"></i> İndir (Kaydet)</a>
        </div>
      </div>
      <div class="iframe-container">
        <iframe id="pdf_frame" src="/view_pdf?type={pdf_type}"></iframe>
      </div>

      <script>
        function goBack() {{
          window.location.href = "/zimmet";
        }}
        function printPdf() {{
          const frame = document.getElementById('pdf_frame');
          try {{
            frame.contentWindow.focus();
            frame.contentWindow.print();
          }} catch (e) {{
            window.print();
          }}
        }}
      </script>
    </body>
    </html>
    """
    return render_template_string(html)


@app.route("/download_pdf")
def download_pdf():
    if not require_login():
        return redirect(url_for("login"))
    pdf_type = request.args.get("type", "zimmet")
    sicil = session.get("sicil") or "unknown"
    filename = f"temp_{sicil}_{pdf_type}.pdf"
    path = os.path.join(APP_DIR, filename)
    if not os.path.exists(path):
        return "PDF bulunamadı.", 404
    return send_file(
        path,
        mimetype="application/octet-stream",
        as_attachment=True,
        download_name=f"{pdf_type}_fisi.pdf"
    )


@app.route("/admin/export/daire_defteri.pdf")
def export_daire_defteri_pdf():
    if not require_login():
        return redirect(url_for("login"))
    daire_adi = (request.args.get("daire_adi") or "").strip()
    if not daire_adi:
        return redirect(url_for("gecmis", msg="Lütfen geçerli bir daire/birim seçiniz."))
    try:
        pdf_bytes = build_daire_envanter_defteri_pdf(daire_adi)
        user_sicil = session.get("sicil") or "unknown"
        temp_filename = f"temp_{user_sicil}_daire_defteri.pdf"
        temp_path = os.path.join(APP_DIR, temp_filename)
        with open(temp_path, "wb") as f:
            f.write(pdf_bytes)

        return redirect(url_for("view_pdf_page", type="daire_defteri"))
    except Exception as e:
        return render_template_string(
            BASE_HTML,
            content="<div class='card p-4'><h5>Daire Defteri Üretilemedi</h5><pre>"+traceback.format_exc()+"</pre></div>",
            message=f"Hata: {e}",
            is_admin=is_admin()
        )



# =========================================================
# YEDEKLEME VEYA GERİ YÜKLEME SİSTEMİ
# =========================================================
BACKUP_DIR = os.path.join(APP_DIR, "backups")

def create_db_backup(prefix="otomatik") -> str:
    if not os.path.exists(BACKUP_DIR):
        os.makedirs(BACKUP_DIR, exist_ok=True)
    timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
    filename = f"zimmet_yedek_{prefix}_{timestamp}.db"
    dest_path = os.path.join(BACKUP_DIR, filename)
    import shutil
    shutil.copy2(DB_PATH, dest_path)
    return filename

def list_db_backups():
    if not os.path.exists(BACKUP_DIR):
        return []
    files = []
    for f in os.listdir(BACKUP_DIR):
        if f.endswith(".db"):
            fp = os.path.join(BACKUP_DIR, f)
            st = os.stat(fp)
            dt = datetime.fromtimestamp(st.st_mtime).strftime("%d/%m/%Y %H:%M:%S")
            size_kb = round(st.st_size / 1024, 1)
            files.append({
                "filename": f,
                "date": dt,
                "size": f"{size_kb} KB",
                "mtime": st.st_mtime
            })
    files.sort(key=lambda x: x["mtime"], reverse=True)
    return files

def restore_db_backup(filename: str) -> bool:
    fp = os.path.join(BACKUP_DIR, filename)
    if not os.path.exists(fp):
        return False
    import shutil
    create_db_backup(prefix="oncesi_guvenlik")
    db.session.remove()
    shutil.copy2(fp, DB_PATH)
    return True

def delete_db_backup(filename: str) -> bool:
    fp = os.path.join(BACKUP_DIR, filename)
    if os.path.exists(fp):
        os.remove(fp)
        return True
    return False

# =========================================================
# AYARLAR (ADMIN)
# =========================================================

@app.route("/ayarlar")
def ayarlar_page():
    if not require_login():
        return redirect(url_for("login"))
    if not is_admin():
        return render_template_string(BASE_HTML, content="<div class='card p-4'><h5>Yetkisiz</h5><div class='label'>Bu sayfa sadece admin içindir.</div></div>", message=None, is_admin=is_admin())

    msg = request.args.get("msg")

    # Kullanıcı Listesi
    try:
        users = User.query.order_by(User.name.asc()).all()
    except Exception:
        users = []

    user_rows = ""
    for u in users:
        is_current = u.sicil == session.get("sicil")
        delete_btn = ""
        if not is_current and u.sicil != ADMIN_SICIL:
            delete_btn = f"""
            <form method="post" action="/admin/delete_user" style="display:inline" onsubmit="return confirm('{u.name} kullanıcısını silmek istediğinize emin misiniz?');">
              <input type="hidden" name="sicil" value="{u.sicil}">
              <button class="btn btn-sm btn-outline-danger ms-1"><i class="fa fa-trash"></i> Sil</button>
            </form>
            """
        role_badge = '<span class="badge bg-primary">Admin</span>' if u.sicil == ADMIN_SICIL else '<span class="badge bg-secondary">Kullanıcı</span>'
        user_rows += f"""
        <tr>
          <td><strong>{u.name}</strong></td>
          <td><span class="badge-soft">{u.sicil}</span></td>
          <td>{role_badge}</td>
          <td class="text-end">
            <button type="button" class="btn btn-sm btn-outline-warning fw-bold" onclick="resetUserPw('{u.sicil}', '{u.name}')"><i class="fa fa-key"></i> Şifre Değiştir</button>
            {delete_btn}
          </td>
        </tr>
        """

    # Veritabanı Yedekleri Listesi
    try:
        backups = list_db_backups()
    except Exception:
        backups = []

    backup_rows = ""
    for b in backups:
        backup_rows += f"""
        <tr>
          <td><i class="fa fa-database text-primary me-2"></i><strong>{b['filename']}</strong></td>
          <td>{b['date']}</td>
          <td><span class="badge-soft">{b['size']}</span></td>
          <td class="text-end">
            <form method="post" action="/admin/restore_backup" style="display:inline" onsubmit="return confirm('{b['filename']} yedeğini geri yüklemek istediğinize emin misiniz? Mevcut veritabanı bu yedeğe dönecektir.');">
              <input type="hidden" name="filename" value="{b['filename']}">
              <button class="btn btn-sm btn-success fw-bold"><i class="fa fa-rotate-left"></i> Geri Yükle</button>
            </form>
            <form method="post" action="/admin/delete_backup" style="display:inline" onsubmit="return confirm('{b['filename']} yedeğini silmek istediğinize emin misiniz?');">
              <input type="hidden" name="filename" value="{b['filename']}">
              <button class="btn btn-sm btn-outline-danger ms-1"><i class="fa fa-trash"></i> Sil</button>
            </form>
          </td>
        </tr>
        """

    # Audit Log (Denetim İzi) Listesi
    try:
        audits = AuditLog.query.order_by(AuditLog.id.desc()).limit(200).all()
    except Exception:
        audits = []

    audit_rows = ""
    for a in audits:
        audit_rows += f"""
        <tr>
          <td><span class="badge-soft">{a.at}</span></td>
          <td><strong>{a.user_name or ''}</strong> ({a.user_sicil or '-'})</td>
          <td><span class="badge bg-secondary">{a.action}</span></td>
          <td>{a.details or ''}</td>
          <td><code>{a.ip_address or '-'}</code></td>
        </tr>
        """

    content = f"""
    <div class="card p-4 mb-3">
      <h3>Ayarlar ve Veri Yönetimi</h3>
      <div class="label">Kullanıcıları yönetebilir, verilerinizi yedekleyebilir veya geçmişe tek tıkla geri dönebilirsiniz.</div>
    </div>

    <!-- Kullanıcı Yönetimi Paneli -->
    <div class="card p-4 mb-3">
      <h6><i class="fa fa-users text-primary me-2"></i>Kullanıcı Yönetimi (Yeni Kullanıcı Ekle & Sil)</h6>
      <div class="label mb-3">Sisteme giriş yapabilecek yeni personel/kullanıcı ekleyebilir veya mevcut hesapları yönetebilirsiniz.</div>

      <form method="post" action="/admin/add_user" class="row g-2 mb-4 p-3 style-card-inner" style="background: rgba(255,255,255,0.02); border: 1px solid var(--border); border-radius: 12px;">
        <div class="col-md-4">
          <div class="label mb-1">Ad Soyad</div>
          <input class="form-control" name="name" placeholder="Örn: Ahmet YILMAZ" required>
        </div>
        <div class="col-md-3">
          <div class="label mb-1">Sicil No</div>
          <input class="form-control" name="sicil" placeholder="Örn: 198273" required>
        </div>
        <div class="col-md-3">
          <div class="label mb-1">Şifre</div>
          <input type="password" class="form-control" name="password" placeholder="******" required>
        </div>
        <div class="col-md-2 d-flex align-items-end">
          <button class="btn btn-primary w-100 fw-bold"><i class="fa fa-user-plus me-1"></i> Kullanıcı Ekle</button>
        </div>
      </form>

      <div class="table-responsive">
        <table class="table table-striped table-hover align-middle">
          <thead>
            <tr>
              <th>Ad Soyad</th>
              <th>Sicil No</th>
              <th>Yetki</th>
              <th class="text-end">İşlem</th>
            </tr>
          </thead>
          <tbody>
            {user_rows or '<tr><td colspan="4" class="text-center py-3 text-muted">Kayıtlı kullanıcı yok.</td></tr>'}
          </tbody>
        </table>
      </div>
    </div>

    <!-- Modal for Admin Reset Password -->
    <div class="modal fade" id="adminResetPwModal" tabindex="-1" aria-hidden="true">
      <div class="modal-dialog modal-dialog-centered">
        <div class="modal-content" style="background: var(--card); border: 1px solid var(--border); color: var(--text); border-radius: 16px;">
          <div class="modal-header" style="border-bottom: 1px solid var(--border);">
            <h5 class="modal-title fw-bold" id="adminResetModalTitle"><i class="fa fa-key me-2 text-primary"></i>Şifre Değiştir</h5>
            <button type="button" class="btn-close btn-close-white" data-bs-dismiss="modal"></button>
          </div>
          <form action="/admin/reset_user_pw" method="post">
            <input type="hidden" name="sicil" id="reset_user_sicil">
            <div class="modal-body">
              <div class="mb-3">
                <div class="label mb-1">Yeni Şifre</div>
                <input type="password" class="form-control" name="new_password" required placeholder="Yeni şifreyi giriniz">
              </div>
            </div>
            <div class="modal-footer" style="border-top: 1px solid var(--border);">
              <button type="button" class="btn btn-secondary" data-bs-dismiss="modal">İptal</button>
              <button type="submit" class="btn btn-primary fw-bold">Şifreyi Güncelle</button>
            </div>
          </form>
        </div>
      </div>
    </div>

    <script>
    function resetUserPw(sicil, name) {{
      document.getElementById('reset_user_sicil').value = sicil;
      document.getElementById('adminResetModalTitle').innerHTML = '<i class="fa fa-key me-2 text-primary"></i>' + name + ' için Şifre Değiştir';
      new bootstrap.Modal(document.getElementById('adminResetPwModal')).show();
    }}
    </script>

    <div class="card p-4 mb-3">
      <h6><i class="fa fa-file-export text-primary me-2"></i>Veri & Kayıt Dışa Aktar</h6>
      <div class="row g-3 mt-1">
        <div class="col-md-6">
          <div class="card p-3 h-100">
            <div class="label mb-2">Tüm Zimmet Kayıtları</div>
            <div class="d-flex gap-2 flex-wrap">
              <a class="btn btn-primary fw-bold" href="/admin/export/zimmet.xlsx" target="_blank"><i class="fa fa-file-excel me-1"></i> Excel (XLSX)</a>
              <a class="btn btn-outline-primary fw-bold" href="/admin/export/zimmet.pdf" target="_blank"><i class="fa fa-file-pdf me-1"></i> PDF Raporu</a>
            </div>
          </div>
        </div>
        <div class="col-md-6">
          <div class="card p-3 h-100">
            <div class="label mb-2">Tüm Teslim Tesellüm Kayıtları</div>
            <div class="d-flex gap-2 flex-wrap">
              <a class="btn btn-primary fw-bold" href="/admin/export/teslim.xlsx" target="_blank"><i class="fa fa-file-excel me-1"></i> Excel (XLSX)</a>
              <a class="btn btn-outline-primary fw-bold" href="/admin/export/teslim.pdf" target="_blank"><i class="fa fa-file-pdf me-1"></i> PDF Raporu</a>
            </div>
          </div>
        </div>
        <div class="col-md-6">
          <div class="card p-3 h-100">
            <div class="label mb-2">Tüm Envanter Cihaz Kayıtları</div>
            <div class="d-flex gap-2 flex-wrap">
              <a class="btn btn-primary fw-bold" href="/admin/export/envanter.xlsx" target="_blank"><i class="fa fa-file-excel me-1"></i> Excel (XLSX)</a>
              <a class="btn btn-outline-primary fw-bold" href="/admin/export/envanter.pdf" target="_blank"><i class="fa fa-file-pdf me-1"></i> PDF Raporu</a>
            </div>
          </div>
        </div>
        <div class="col-md-6">
          <div class="card p-3 h-100">
            <div class="label mb-2">Tüm Depo / Sarf Stok Kayıtları</div>
            <div class="d-flex gap-2 flex-wrap">
              <a class="btn btn-primary fw-bold" href="/admin/export/depo.xlsx" target="_blank"><i class="fa fa-file-excel me-1"></i> Excel (XLSX)</a>
              <a class="btn btn-outline-primary fw-bold" href="/admin/export/depo.pdf" target="_blank"><i class="fa fa-file-pdf me-1"></i> PDF Raporu</a>
            </div>
          </div>
        </div>
      </div>
    </div>

    <!-- Veritabanı Yedekleme ve Geri Yükleme Paneli -->
    <div class="card p-4 mb-3">
      <div class="d-flex justify-content-between align-items-center mb-3 flex-wrap gap-2">
        <div>
          <h6 class="mb-1"><i class="fa fa-box-archive text-primary me-2"></i>Otomatik Veritabanı Yedekleri</h6>
          <div class="label mb-2">Sistem her gün saat <strong>17:00</strong> zamanında otomatik yedek ve her geçmiş temizliğinde güvenlik yedeği alır.</div>
          <div class="alert alert-success py-2 mb-3" style="background: rgba(16, 185, 129, 0.15); border: 1px solid rgba(16, 185, 129, 0.3) !important; color: #6ee7b7; font-size: 13px;">
            <i class="fa fa-clock me-2"></i> <strong>Otomatik Günlük Yedekleme Aktif:</strong> Her gün saat 17:00 itibarıyla arka planda otomatik veritabanı yedeği alınır.
          </div>
        </div>
        <form method="post" action="/admin/backup_now">
          <button class="btn btn-success fw-bold"><i class="fa fa-download me-1"></i> Şimdi Manuel Yedek Al</button>
        </form>
      </div>

      <div class="table-responsive">
        <table class="table table-striped table-hover align-middle">
          <thead>
            <tr>
              <th>Yedek Dosyası</th>
              <th>Oluşturulma Tarihi</th>
              <th>Boyut</th>
              <th class="text-end">İşlem</th>
            </tr>
          </thead>
          <tbody>
            {backup_rows or '<tr><td colspan="4" class="text-center py-3 text-muted">Henüz kayıtlı veritabanı yedeği yok.</td></tr>'}
          </tbody>
        </table>
      </div>
    </div>

    <div class="card p-4 border-danger mb-3">
      <h6 class="text-danger"><i class="fa fa-triangle-exclamation me-2"></i>Zimmet Geçmişini Temizle</h6>
      <div class="label mb-2">Temizle butonuna bastığınızda veriler silinmeden önce <strong>otomatik tarihli güvenlik yedeği</strong> alınır.</div>
      <form method="post" action="/admin/clear_history" onsubmit="return confirm('Tüm zimmet ve teslim geçmişi temizlenecek! (Otomatik yedek oluşturulacaktır). Onaylıyor musunuz?');" class="mt-2">
        <button class="btn btn-danger fw-bold"><i class="fa fa-trash me-1"></i> Geçmişi Temizle (Güvenlik Yedeği İle)</button>
      </form>
    </div>

    <!-- Sistem Denetim İzi (Audit Log) Paneli -->
    <div class="card p-4">
      <h6><i class="fa fa-shield-halved text-primary me-2"></i>Sistem Denetim İzi (Audit Log - Güvenlik Günlüğü)</h6>
      <div class="label mb-3">Sistem üzerinde gerçekleştirilen tüm ekleme, silme, zimmet ve durum değişikliklerinin kronolojik kaydı.</div>

      <div class="table-responsive" style="max-height: 380px; overflow-y: auto;">
        <table class="table table-striped table-hover align-middle" style="font-size: 12px;">
          <thead style="position: sticky; top: 0; z-index: 2;">
            <tr>
              <th>Tarih / Saat</th>
              <th>İşlem Yapan Kullanıcı</th>
              <th>İşlem Türü</th>
              <th>Açıklama / Detay</th>
              <th>IP Adresi</th>
            </tr>
          </thead>
          <tbody>
            {audit_rows or '<tr><td colspan="5" class="text-center py-3 text-muted">Henüz kayıtlı denetim kaydı bulunmuyor.</td></tr>'}
          </tbody>
        </table>
      </div>
    </div>
    """

    return render_template_string(BASE_HTML, content=content, message=msg, is_admin=is_admin())


@app.route("/admin/add_user", methods=["POST"])
def admin_add_user():
    if not require_login() or not is_admin():
        return redirect(url_for("login"))
    name = format_tr_name(request.form.get("name") or "")
    sicil = (request.form.get("sicil") or "").strip()
    pw = request.form.get("password") or ""
    if not name or not sicil or not pw:
        return redirect(url_for("ayarlar_page", msg="Lütfen tüm kullanıcı alanlarını doldurunuz."))
    if User.query.filter_by(sicil=sicil).first():
        return redirect(url_for("ayarlar_page", msg=f"Sicil No '{sicil}' ile kayıtlı kullanıcı zaten var!"))
    try:
        u = User(name=name, sicil=sicil, password_hash=generate_password_hash(pw))
        db.session.add(u)
        db.session.commit()
        log_audit("KULLANICI_EKLENDI", f"Yeni kullanıcı: {name} ({sicil})")
        return redirect(url_for("ayarlar_page", msg=f"Yeni kullanıcı '{name}' ({sicil}) başarıyla eklendi."))
    except Exception as e:
        db.session.rollback()
        return redirect(url_for("ayarlar_page", msg=f"Kullanıcı ekleme hatası: {e}"))


@app.route("/admin/delete_user", methods=["POST"])
def admin_delete_user():
    if not require_login() or not is_admin():
        return redirect(url_for("login"))
    sicil = (request.form.get("sicil") or "").strip()
    if sicil == session.get("sicil"):
        return redirect(url_for("ayarlar_page", msg="O an aktif olan kendi hesabınızı silemezsiniz."))
    u = User.query.filter_by(sicil=sicil).first()
    if u:
        db.session.delete(u)
        db.session.commit()
        log_audit("KULLANICI_SILINDI", f"Kullanıcı silindi: {u.name} ({sicil})")
        return redirect(url_for("ayarlar_page", msg=f"Kullanıcı '{u.name}' ({sicil}) başarıyla silindi."))
    return redirect(url_for("ayarlar_page", msg="Kullanıcı bulunamadı."))


@app.route("/admin/download_envanter_sablon")
def download_envanter_sablon():
    if not require_login():
        return redirect(url_for("login"))
    try:
        import openpyxl
        from openpyxl.styles import PatternFill, Font, Alignment

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Envanter Yukleme"

        headers = ["Kategori", "Marka", "Model", "Seri No"]
        ws.append(headers)

        fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
        font = Font(color="FFFFFF", bold=True)
        for col_idx in range(1, 5):
            cell = ws.cell(row=1, column=col_idx)
            cell.fill = fill
            cell.font = font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        sample_rows = [
            ["Kasa", "Lenovo", "ThinkCentre M720", "SNO-10029384"],
            ["Monitör", "Dell", "P2419H", "SNO-99201948"],
            ["Yazıcı", "Lexmark", "MS823dn", "SNO-77182930"],
            ["Tarayıcı", "Fujitsu", "fi-7160", "SNO-44920192"],
        ]
        for r in sample_rows:
            ws.append(r)

        ws.column_dimensions["A"].width = 18
        ws.column_dimensions["B"].width = 20
        ws.column_dimensions["C"].width = 25
        ws.column_dimensions["D"].width = 25

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return send_file(
            buf,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            as_attachment=True,
            download_name="ornek_envanter_sablonu.xlsx"
        )
    except Exception as e:
        return f"Şablon üretme hatası: {e}", 500


@app.route("/admin/import_envanter_excel", methods=["POST"])
def admin_import_envanter_excel():
    if not require_login():
        return redirect(url_for("login"))
    if "excel_file" not in request.files:
        return redirect(url_for("envanter", msg="Lütfen bir Excel dosyası seçiniz."))

    file = request.files["excel_file"]
    if not file or not file.filename.endswith((".xlsx", ".xls")):
        return redirect(url_for("envanter", msg="Lütfen geçerli bir .xlsx veya .xls Excel dosyası yükleyiniz."))

    try:
        import openpyxl
        wb = openpyxl.load_workbook(file, data_only=True)
        sheet = wb.active

        rows = list(sheet.iter_rows(values_only=True))
        if not rows or len(rows) < 2:
            return redirect(url_for("envanter", msg="Excel dosyasında veri satırı bulunamadı."))

        header_row = [str(cell or "").strip().lower() for cell in rows[0]]

        cat_idx, brand_idx, model_idx, serial_idx = -1, -1, -1, -1

        for idx, h in enumerate(header_row):
            if any(k in h for k in ["kategori", "tür", "cihaz türü", "donanım"]):
                cat_idx = idx
            elif any(k in h for k in ["marka", "üretici", "brand"]):
                brand_idx = idx
            elif any(k in h for k in ["model", "tip"]):
                model_idx = idx
            elif any(k in h for k in ["seri", "serial", "sn", "s/n", "barkod"]):
                serial_idx = idx

        if cat_idx == -1: cat_idx = 0
        if brand_idx == -1: brand_idx = 1
        if model_idx == -1: model_idx = 2
        if serial_idx == -1: serial_idx = 3

        added_count = 0
        skipped_count = 0

        def clean_val(val):
            if val is None:
                return ""
            s = str(val).strip()
            if s.endswith(".0") and s[:-2].isdigit():
                s = s[:-2]
            return s

        for row_idx, row in enumerate(rows[1:], start=2):
            if not row or not any(row):
                continue

            category = clean_val(row[cat_idx] if cat_idx < len(row) else "")
            brand = clean_val(row[brand_idx] if brand_idx < len(row) else "")
            model = clean_val(row[model_idx] if model_idx < len(row) else "")
            serial = clean_val(row[serial_idx] if serial_idx < len(row) else "")

            if not brand and len(row) > 0:
                for cell_v in row:
                    cv = clean_val(cell_v)
                    if cv:
                        brand = cv
                        break

            if not category or category not in DEVICE_CATEGORIES:
                category = "Kasa"

            if not brand:
                brand = "Genel"

            if not serial:
                serial = f"OTO-SN-{datetime.now().strftime('%Y%m%d')}-{row_idx}"

            if InventoryItem.query.filter_by(serial_no=serial).first():
                skipped_count += 1
                continue

            db.session.add(InventoryItem(
                category=category,
                brand=brand,
                model=(model or None),
                serial_no=serial,
                status="Depoda",
                last_event="Excel Import",
                last_event_at=now_str()
            ))
            added_count += 1

        db.session.commit()
        log_audit("EXCEL_IMPORT", f"Excel'den {added_count} cihaz yüklendi ({skipped_count} mükerrer atlandı).")
        return redirect(url_for("envanter", msg=f"Toplu Aktarım Başarılı! {added_count} adet cihaz envantere eklendi. ({skipped_count} mükerrer seri numarası atlandı)"))
    except Exception as e:
        db.session.rollback()
        return redirect(url_for("envanter", msg=f"Excel okuma hatası: {e}"))


@app.route("/pdf_daire_zimmet", methods=["POST"])
def pdf_daire_zimmet():
    if not require_login():
        return redirect(url_for("login"))
    try:
        daire_adi = normalize_spaces(request.form.get("daire_adi") or "")
        makam_title = normalize_spaces(request.form.get("makam_title") or "")
        ym_ad = normalize_spaces(request.form.get("yazi_isleri_muduru_ad") or "")
        ym_sicil = normalize_spaces(request.form.get("yazi_isleri_muduru_sicil") or "")

        teslim_ad = normalize_spaces(request.form.get("teslim_ad") or session.get("user") or "")
        teslim_sicil = normalize_spaces(request.form.get("teslim_sicil") or session.get("sicil") or "")

        if not daire_adi or not makam_title:
            raise ValueError("Lütfen Daire/Birim ve Makam unvanını giriniz.")
        if not ym_ad or not ym_sicil:
            raise ValueError("Daire Yazı İşleri Müdürü adı ve sicil numarası zorunludur.")

        inv_ids = parse_checked_ids("z_daire")
        manual = parse_manual_devices()

        devices: List[Dict] = []
        for iid in inv_ids:
            it = InventoryItem.query.get(iid)
            if not it or it.status != "Depoda":
                continue
            devices.append({
                "inv_id": it.id,
                "category": it.category or "",
                "brand": it.brand or "",
                "model": it.model or "",
                "serial_no": it.serial_no or "",
            })

        for d in manual:
            cat = normalize_spaces(d.get("category") or "")
            brand = normalize_spaces(d.get("brand") or "")
            model = normalize_spaces(d.get("model") or "")
            serial = normalize_spaces(d.get("serial_no") or "")
            if cat or brand or model or serial:
                devices.append({
                    "inv_id": None,
                    "category": cat,
                    "brand": brand,
                    "model": model,
                    "serial_no": serial,
                })

        if not devices:
            raise ValueError("En az 1 cihaz seçmelisiniz.")

        batch_at = now_str()
        for d in devices:
            if d.get("inv_id"):
                it = InventoryItem.query.get(int(d["inv_id"]))
                if it:
                    it.status = "Zimmetli"
                    it.assigned_name = f"{ym_ad} (Y.İ.M - {makam_title})"
                    it.assigned_sicil = ym_sicil
                    it.assigned_title = makam_title
                    it.assigned_unit = daire_adi
                    it.assigned_at = today_str()
                    it.last_event = "DAIRE_ZIMMET"
                    it.last_event_at = batch_at

            db.session.add(History(
                at=batch_at,
                action="ZIMMET",
                inv_id=d.get("inv_id"),
                category=d.get("category"),
                brand=d.get("brand"),
                model=d.get("model"),
                serial_no=d.get("serial_no"),
                alan_ad=ym_ad,
                alan_sicil=ym_sicil,
                alan_unvan=makam_title,
                alan_birim=daire_adi,
                teslim_ad=teslim_ad,
                teslim_sicil=teslim_sicil,
            ))
        db.session.commit()

        log_audit("DAIRE_ZIMMET", f"{daire_adi} ({makam_title}) - Y.İ.M: {ym_ad} ({ym_sicil}) - {len(devices)} cihaz")

        pdf_bytes = build_daire_makam_zimmet_pdf(
            daire_adi=daire_adi,
            makam_title=makam_title,
            yazi_isleri_muduru_ad=ym_ad,
            yazi_isleri_muduru_sicil=ym_sicil,
            teslim_alan_ad=teslim_ad,
            teslim_alan_sicil=teslim_sicil,
            devices=devices
        )

        sicil_key = session.get("sicil") or "unknown"
        temp_filename = f"temp_{sicil_key}_zimmet.pdf"
        temp_path = os.path.join(APP_DIR, temp_filename)
        with open(temp_path, "wb") as f:
            f.write(pdf_bytes)

        return redirect(url_for("view_pdf_page", type="zimmet"))
    except Exception as e:
        db.session.rollback()
        return render_template_string(
            BASE_HTML,
            content="<div class='card p-4'><h5>Hata Oluştu</h5><pre>"+traceback.format_exc()+"</pre></div>",
            message=f"Daire zimmet belgesi üretilemedi: {e}"
        )


@app.route("/admin/reset_user_pw", methods=["POST"])
def admin_reset_user_pw():
    if not require_login() or not is_admin():
        return redirect(url_for("login"))
    sicil = (request.form.get("sicil") or "").strip()
    new_pw = request.form.get("new_password") or ""
    if not sicil or not new_pw:
        return redirect(url_for("ayarlar_page", msg="Geçersiz şifre sıfırlama talebi."))
    u = User.query.filter_by(sicil=sicil).first()
    if u:
        u.password_hash = generate_password_hash(new_pw)
        db.session.commit()
        return redirect(url_for("ayarlar_page", msg=f"'{u.name}' kullanıcısının şifresi başarıyla güncellendi."))
    return redirect(url_for("ayarlar_page", msg="Kullanıcı bulunamadı."))


@app.route("/admin/backup_now", methods=["POST"])
def admin_backup_now():
    if not require_login() or not is_admin():
        return redirect(url_for("login"))
    try:
        fname = create_db_backup(prefix="manuel")
        return redirect(url_for("ayarlar_page", msg=f"Yeni veritabanı yedeği oluşturuldu: {fname}"))
    except Exception as e:
        return redirect(url_for("ayarlar_page", msg=f"Yedekleme hatası: {e}"))


@app.route("/admin/restore_backup", methods=["POST"])
def admin_restore_backup():
    if not require_login() or not is_admin():
        return redirect(url_for("login"))
    fname = request.form.get("filename")
    if fname and restore_db_backup(fname):
        return redirect(url_for("ayarlar_page", msg=f"Veritabanı '{fname}' yedeğine başarıyla geri döndürüldü!"))
    return redirect(url_for("ayarlar_page", msg="Geri yükleme başarısız."))


@app.route("/admin/delete_backup", methods=["POST"])
def admin_delete_backup():
    if not require_login() or not is_admin():
        return redirect(url_for("login"))
    fname = request.form.get("filename")
    if fname and delete_db_backup(fname):
        return redirect(url_for("ayarlar_page", msg=f"Yedek dosyası '{fname}' silindi."))
    return redirect(url_for("ayarlar_page", msg="Silme başarısız."))


@app.route("/admin/clear_history", methods=["POST"])
def admin_clear_history():
    if not require_login():
        return redirect(url_for("login"))
    if not is_admin():
        return render_template_string(BASE_HTML, content="<div class='card p-4'><h5>Yetkisiz</h5></div>", message=None, is_admin=is_admin())

    try:
        fname = create_db_backup(prefix="temizlik_oncesi")
        History.query.delete()
        db.session.commit()
        return redirect(url_for("ayarlar_page", msg=f"Zimmet geçmişi temizlendi. (Silinmeden önceki yedek '{fname}' adıyla kaydedildi.)"))
    except Exception as e:
        db.session.rollback()
        return render_template_string(BASE_HTML, content="<div class='card p-4'><h5>Hata</h5><pre>"+traceback.format_exc()+"</pre></div>", message=str(e), is_admin=is_admin())


@app.route("/admin/export/zimmet.xlsx")
def admin_export_zimmet_xlsx():
    if not require_login() or not is_admin():
        return redirect(url_for("login"))
    try:
        data = export_history_xlsx("ZIMMET")
        buf = io.BytesIO(data)
        buf.seek(0)
        return send_file(
            buf,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            as_attachment=True,
            download_name="zimmet_kayitlari.xlsx"
        )
    except Exception as e:
        return render_template_string(BASE_HTML, content=f"<div class='card p-4'><h5>Dışa Aktarma Hatası</h5><pre>{traceback.format_exc()}</pre></div>", message=str(e), is_admin=is_admin())


@app.route("/admin/export/teslim.xlsx")
def admin_export_teslim_xlsx():
    if not require_login() or not is_admin():
        return redirect(url_for("login"))
    try:
        data = export_history_xlsx("TESLIM_TESELLUM")
        buf = io.BytesIO(data)
        buf.seek(0)
        return send_file(
            buf,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            as_attachment=True,
            download_name="teslim_tesellum_kayitlari.xlsx"
        )
    except Exception as e:
        return render_template_string(BASE_HTML, content=f"<div class='card p-4'><h5>Dışa Aktarma Hatası</h5><pre>{traceback.format_exc()}</pre></div>", message=str(e), is_admin=is_admin())


@app.route("/admin/export/envanter.xlsx")
def admin_export_envanter_xlsx():
    if not require_login() or not is_admin():
        return redirect(url_for("login"))
    try:
        data = export_envanter_xlsx()
        buf = io.BytesIO(data)
        buf.seek(0)
        return send_file(
            buf,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            as_attachment=True,
            download_name="tum_envanter_kayitlari.xlsx"
        )
    except Exception as e:
        return render_template_string(BASE_HTML, content=f"<div class='card p-4'><h5>Dışa Aktarma Hatası</h5><pre>{traceback.format_exc()}</pre></div>", message=str(e), is_admin=is_admin())


@app.route("/admin/export/depo.xlsx")
def admin_export_depo_xlsx():
    if not require_login() or not is_admin():
        return redirect(url_for("login"))
    try:
        data = export_depo_xlsx()
        buf = io.BytesIO(data)
        buf.seek(0)
        return send_file(
            buf,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            as_attachment=True,
            download_name="depo_stok_kayitlari.xlsx"
        )
    except Exception as e:
        return render_template_string(BASE_HTML, content=f"<div class='card p-4'><h5>Dışa Aktarma Hatası</h5><pre>{traceback.format_exc()}</pre></div>", message=str(e), is_admin=is_admin())


@app.route("/admin/export/zimmet.pdf")
def admin_export_zimmet_pdf():
    if not require_login() or not is_admin():
        return redirect(url_for("login"))
    try:
        pdf = export_history_pdf("ZIMMET", "Tüm Zimmet Kayıtları Raporu")
        buf = io.BytesIO(pdf)
        buf.seek(0)
        return send_file(buf, mimetype="application/pdf", as_attachment=True, download_name="zimmet_kayitlari_raporu.pdf")
    except Exception as e:
        return render_template_string(BASE_HTML, content=f"<div class='card p-4'><h5>Dışa Aktarma Hatası</h5><pre>{traceback.format_exc()}</pre></div>", message=str(e), is_admin=is_admin())


@app.route("/admin/export/teslim.pdf")
def admin_export_teslim_pdf():
    if not require_login() or not is_admin():
        return redirect(url_for("login"))
    try:
        pdf = export_history_pdf("TESLIM_TESELLUM", "Tüm Teslim Tesellüm Kayıtları Raporu")
        buf = io.BytesIO(pdf)
        buf.seek(0)
        return send_file(buf, mimetype="application/pdf", as_attachment=True, download_name="teslim_tesellum_kayitlari_raporu.pdf")
    except Exception as e:
        return render_template_string(BASE_HTML, content=f"<div class='card p-4'><h5>Dışa Aktarma Hatası</h5><pre>{traceback.format_exc()}</pre></div>", message=str(e), is_admin=is_admin())


@app.route("/admin/export/envanter.pdf")
def admin_export_envanter_pdf():
    if not require_login() or not is_admin():
        return redirect(url_for("login"))
    try:
        pdf = export_envanter_pdf()
        buf = io.BytesIO(pdf)
        buf.seek(0)
        return send_file(buf, mimetype="application/pdf", as_attachment=True, download_name="envanter_cihaz_raporu.pdf")
    except Exception as e:
        return render_template_string(BASE_HTML, content=f"<div class='card p-4'><h5>Dışa Aktarma Hatası</h5><pre>{traceback.format_exc()}</pre></div>", message=str(e), is_admin=is_admin())


@app.route("/admin/export/depo.pdf")
def admin_export_depo_pdf():
    if not require_login() or not is_admin():
        return redirect(url_for("login"))
    try:
        pdf = export_depo_pdf()
        buf = io.BytesIO(pdf)
        buf.seek(0)
        return send_file(buf, mimetype="application/pdf", as_attachment=True, download_name="depo_stok_raporu.pdf")
    except Exception as e:
        return render_template_string(BASE_HTML, content=f"<div class='card p-4'><h5>Dışa Aktarma Hatası</h5><pre>{traceback.format_exc()}</pre></div>", message=str(e), is_admin=is_admin())

# =========================================================
# YENİ MODÜLLER: ANALİTİK, DAİRE İNCELEME & OTOMATİK YEDEKLEME
# =========================================================

@app.route("/api/analytics")
def api_analytics():
    if not require_login():
        return jsonify({"ok": False, "error": "Unauthorized"}), 401
    try:
        # 1. Daire Bazlı Cihaz Dağılımı (Zimmetli olanlar)
        items = InventoryItem.query.filter_by(status="Zimmetli").all()
        unit_map = {}
        for it in items:
            u = (it.assigned_unit or "Belirtilmemiş").strip()
            c = (it.category or "Diğer").strip()
            if u not in unit_map:
                unit_map[u] = {"Kasa": 0, "Monitör": 0, "Yazıcı": 0, "Tarayıcı": 0}
            if c in unit_map[u]:
                unit_map[u][c] += 1
            else:
                unit_map[u]["Kasa"] += 1

        sorted_units = sorted(unit_map.items(), key=lambda x: sum(x[1].values()), reverse=True)[:8]
        u_labels = [x[0] for x in sorted_units]
        kasa = [x[1].get("Kasa", 0) for x in sorted_units]
        mon = [x[1].get("Monitör", 0) for x in sorted_units]
        yaz = [x[1].get("Yazıcı", 0) for x in sorted_units]
        tar = [x[1].get("Tarayıcı", 0) for x in sorted_units]

        # 2. Aylık İşlem / Sarf Tüketim Hacmi
        hist_items = History.query.all()
        m_map = {}
        for h in hist_items:
            try:
                m = datetime.strptime(h.at, "%d.%m.%Y %H:%M:%S").strftime("%Y-%m")
            except Exception:
                continue
            m_map[m] = m_map.get(m, 0) + 1
        
        m_labels = sorted(m_map.keys())[-6:] if m_map else [datetime.now().strftime("%Y-%m")]
        m_data = [m_map.get(m, 0) for m in m_labels]

        return jsonify({
            "ok": True,
            "units": {"labels": u_labels, "kasa": kasa, "monitor": mon, "yazici": yaz, "tarayici": tar},
            "monthly": {"labels": m_labels, "data": m_data}
        })
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500


@app.route("/daire_incele")
def daire_incele():
    if not require_login():
        return redirect(url_for("login"))
    
    selected_unit = request.args.get("unit") or "1. Ceza Dairesi"
    
    # Birim Seçim Listesi (HTML Optgroup)
    daire_options = ""
    for g_name, u_list in BAM_UNITS_STRUCTURE.items():
        daire_options += f'<optgroup label="{g_name}">'
        for u in u_list:
            sel = "selected" if u.strip().lower() == selected_unit.strip().lower() else ""
            daire_options += f'<option value="{u}" {sel}>{u}</option>'
        daire_options += '</optgroup>'

    # Seçilen birime ait zimmetli cihazlar
    all_invs = InventoryItem.query.all()
    items = [it for it in all_invs if (it.assigned_unit and selected_unit.strip().lower() in it.assigned_unit.strip().lower() and it.status == "Zimmetli")]
    
    personnel_set = set()
    cat_counts = {"Kasa": 0, "Monitör": 0, "Yazıcı": 0, "Tarayıcı": 0}
    for it in items:
        if it.assigned_name:
            personnel_set.add(it.assigned_name)
        if it.category in cat_counts:
            cat_counts[it.category] += 1

    table_rows = ""
    for idx, it in enumerate(items, 1):
        badge_cls = "bg-primary" if it.category == "Kasa" else ("bg-info text-dark" if it.category == "Monitör" else ("bg-warning text-dark" if it.category == "Yazıcı" else "bg-success"))
        table_rows += f"""
        <tr>
          <td class="text-center fw-bold">{idx}</td>
          <td><span class="badge {badge_cls} px-2 py-1">{it.category}</span></td>
          <td><strong style="color:var(--heading);">{it.brand}</strong> <span class="muted">{it.model or ''}</span></td>
          <td><code>{it.serial_no}</code></td>
          <td><strong style="color:var(--primary);">{it.assigned_name or '-'}</strong> <small class="muted">({it.assigned_sicil or ''})</small></td>
          <td>{it.assigned_at or '-'}</td>
        </tr>
        """

    content = f"""
    <div class="card p-4 mb-3" style="background: linear-gradient(135deg, rgba(245, 158, 11, 0.1) 0%, rgba(15, 23, 42, 0.4) 100%); border: 1px solid rgba(245, 158, 11, 0.25);">
      <div class="d-flex align-items-center justify-content-between flex-wrap gap-3">
        <div>
          <h3 class="mb-0 fw-bold"><i class="fa fa-building text-warning me-2"></i> Daire & Birim Bazlı İnceleme Paneli</h3>
          <div class="label mt-1" style="color: #fde68a;">Diyarbakır BAM bünyesindeki birimlerin aktif zimmetli tüm donanım ve personel haritası.</div>
        </div>
        <a href="/admin/export/daire_toplu.pdf?unit={selected_unit}" class="btn btn-danger btn-lg fw-bold shadow-sm"><i class="fa fa-file-pdf me-2"></i> Toplu Zimmet Raporu (PDF) İndir</a>
      </div>
    </div>

    <div class="card p-4 mb-3">
      <form method="get" action="/daire_incele" class="row g-3 align-items-center">
        <div class="col-md-9">
          <label class="label mb-1 fw-bold" style="color: var(--heading);">İncelenecek Daire / İdari Büro Seçin</label>
          <select name="unit" class="form-select form-select-lg fw-bold" style="background: var(--input-bg); color: var(--input-text);" onchange="this.form.submit()">
            {daire_options}
          </select>
        </div>
        <div class="col-md-3 text-end pt-4">
          <button type="submit" class="btn btn-primary btn-lg w-100 fw-bold"><i class="fa fa-filter me-1"></i> Detayları Getir</button>
        </div>
      </form>
    </div>

    <div class="row g-3 mb-3">
      <div class="col-md-3 col-6"><div class="card p-3 text-center"><div class="label mb-1">Birimdeki Zimmetli Cihaz</div><h4 class="mb-0 fw-bold text-warning">{len(items)}</h4></div></div>
      <div class="col-md-3 col-6"><div class="card p-3 text-center"><div class="label mb-1">Zimmetli Personel Sayısı</div><h4 class="mb-0 fw-bold text-primary">{len(personnel_set)}</h4></div></div>
      <div class="col-md-3 col-6"><div class="card p-3 text-center"><div class="label mb-1">Kasa / Monitör</div><h4 class="mb-0 fw-bold text-info">{cat_counts['Kasa']} / {cat_counts['Monitör']}</h4></div></div>
      <div class="col-md-3 col-6"><div class="card p-3 text-center"><div class="label mb-1">Yazıcı / Tarayıcı</div><h4 class="mb-0 fw-bold text-success">{cat_counts['Yazıcı']} / {cat_counts['Tarayıcı']}</h4></div></div>
    </div>

    <div class="card p-4">
      <h5 class="fw-bold mb-3" style="color: var(--heading);"><i class="fa fa-list me-2 text-warning"></i> {selected_unit} Zimmetli Donanım Listesi</h5>
      <div class="table-responsive">
        <table class="table align-middle">
          <thead>
            <tr>
              <th class="text-center" style="width:50px">#</th>
              <th>Kategori</th>
              <th>Marka / Model</th>
              <th>Seri Numarası</th>
              <th>Kullanan Personel (Sicil)</th>
              <th>Zimmet Tarihi</th>
            </tr>
          </thead>
          <tbody>
            {table_rows or '<tr><td colspan="6" class="text-center muted py-4"><i class="fa fa-info-circle me-1"></i> Bu dairede zimmetli aktif cihaz bulunmamaktadır.</td></tr>'}
          </tbody>
        </table>
      </div>
    </div>
    """
    return render_base(content)


@app.route("/admin/export/daire_toplu.pdf")
def admin_export_daire_toplu_pdf():
    if not require_login():
        return redirect(url_for("login"))
    unit = request.args.get("unit") or "1. Ceza Dairesi"
    try:
        pdf = build_daire_envanter_defteri_pdf(unit)
        buf = io.BytesIO(pdf)
        buf.seek(0)
        safe_unit_name = re.sub(r'[^a-zA-Z0-9]', '_', unit).strip('_')
        return send_file(buf, mimetype="application/pdf", as_attachment=True, download_name=f"{safe_unit_name}_toplu_zimmet_raporu.pdf")
    except Exception as e:
        return render_template_string(BASE_HTML, content=f"<div class='card p-4'><h5>Dışa Aktarma Hatası</h5><pre>{traceback.format_exc()}</pre></div>", message=str(e), is_admin=is_admin())


@app.route("/admin/export/hek_tutanak.pdf")
def admin_export_hek_tutanak_pdf():
    if not require_login():
        return redirect(url_for("login"))
    try:
        hek_items = InventoryItem.query.filter_by(status="Hek / Hurda").all()
        dev_list = []
        for it in hek_items:
            dev_list.append({
                "category": it.category or "Donanım",
                "brand": it.brand or "",
                "model": it.model or "",
                "serial_no": it.serial_no or "-",
                "reason": "Ekonomik Ömrünü Tamamlamış / Arızalı"
            })
        pdf = build_hek_hurda_pdf(dev_list)
        buf = io.BytesIO(pdf)
        buf.seek(0)
        return send_file(buf, mimetype="application/pdf", as_attachment=True, download_name="hek_hurda_ayirma_tutanagi.pdf")
    except Exception as e:
        return render_template_string(BASE_HTML, content=f"<div class='card p-4'><h5>Dışa Aktarma Hatası</h5><pre>{traceback.format_exc()}</pre></div>", message=str(e), is_admin=is_admin())


def send_db_backup_email(to_email=None) -> Tuple[bool, str]:
    import smtplib
    from email.mime.multipart import MIMEMultipart
    from email.mime.base import MIMEBase
    from email.mime.text import MIMEText
    from email import encoders

    smtp_server = os.environ.get("SMTP_SERVER", "smtp.gmail.com")
    smtp_port = int(os.environ.get("SMTP_PORT", 587))
    smtp_user = os.environ.get("SMTP_USER", "")
    smtp_pass = os.environ.get("SMTP_PASS", "")
    target_email = to_email or os.environ.get("BACKUP_EMAIL", "irven90@gmail.com")

    if not smtp_user or not smtp_pass:
        return False, "SMTP e-posta sunucusu veya şifresi ortam değişkenlerinde tanımlanmamış (SMTP_USER / SMTP_PASS)."

    try:
        msg = MIMEMultipart()
        msg['From'] = smtp_user
        msg['To'] = target_email
        msg['Subject'] = f"Diyarbakır BAM Envanter - Otomatik Veritabanı Yedeği ({today_str()})"

        body = f"""Sayın Murat İRVEN,

Diyarbakır Bölge Adliye Mahkemesi Envanter ve Zimmet Sistemi otomatik veritabanı yedeği ekte yer almaktadır.

Tarih: {now_str()}
Sistem: Diyarbakır BAM Bilgi İşlem Müdürlüğü
"""
        msg.attach(MIMEText(body, 'plain', 'utf-8'))

        if os.path.exists(DB_PATH):
            with open(DB_PATH, "rb") as f:
                part = MIMEBase("application", "octet-stream")
                part.set_payload(f.read())
            encoders.encode_base64(part)
            part.add_header("Content-Disposition", f"attachment; filename=Diyarbakir_BAM_Zimmet_Yedek_{today_str()}.db")
            msg.attach(part)

        server = smtplib.SMTP(smtp_server, smtp_port)
        server.starttls()
        server.login(smtp_user, smtp_pass)
        server.send_message(msg)
        server.quit()
        return True, f"Veritabanı yedeği başarıyla {target_email} adresine e-posta olarak gönderildi."
    except Exception as e:
        return False, f"E-Posta gönderim hatası: {e}"


@app.route("/admin/send_db_backup")
def admin_send_db_backup():
    if not require_login() or not is_admin():
        return redirect(url_for("login"))
    ok, msg = send_db_backup_email()
    alert_cls = "alert-success" if ok else "alert-warning"
    return render_template_string(
        BASE_HTML,
        content=f"""
        <div class="card p-4">
          <h4 class="fw-bold mb-3"><i class="fa fa-envelope-open-text text-info me-2"></i> Otomatik Veritabanı Yedek E-Postası</h4>
          <div class="alert {alert_cls}">{msg}</div>
          <div class="mt-3">
            <a href="/ayarlar" class="btn btn-primary fw-bold"><i class="fa fa-arrow-left me-1"></i> Ayarlar Paneline Dön</a>
          </div>
        </div>
        """,
        is_admin=is_admin()
    )


with app.app_context():
    try:
        init_db_once()
    except Exception as e:
        print("init_db_once error:", e)

# =========================================================
# MAIN
# =========================================================
if __name__ == "__main__":
    local_ip = get_local_ip()
    print("\n" + "="*70)
    print(" ZİMMET SİSTEMİ SUNUCUSU AKTİF")
    print(f" Yerel bilgisayardan erişim: http://127.0.0.1:5000")
    print(f" Ağdaki diğer cihazlar için: http://{local_ip}:5000")
    print("="*70 + "\n")
    port = int(os.environ.get("PORT", 5000))
    app.run(host="0.0.0.0", port=port, debug=False)
